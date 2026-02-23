"""assembleur_core.py
Noyau 'pur' (sans Tk) : géométrie + structures de base.
"""

import math
import datetime as _dt
from typing import Optional, List, Dict
import re

import numpy as np
from dataclasses import dataclass

from shapely.geometry import Polygon as _ShPoly, LineString as _ShLine
from shapely.ops import unary_union as _sh_union

import xml.etree.ElementTree as _ET


def _tri_shape(P: Dict[str, np.ndarray]):
    """Triangle → Polygon (brut, sans rétrécissement)."""
    poly = _ShPoly([tuple(map(float, P["O"])),
                    tuple(map(float, P["B"])),
                    tuple(map(float, P["L"]))])
    # 'buffer(0)' nettoie les très petites imprécisions numériques
    return poly.buffer(0)


def _group_shape_from_nodes(nodes, last_drawn):
    """Union BRUTE des triangles listés dans 'nodes' (sans rétrécissement)."""
    polys = []
    for nd in nodes:
        tid = nd.get("tid")
        if tid is None or not (0 <= tid < len(last_drawn)):
            continue
        polys.append(_tri_shape(last_drawn[tid]["pts"]))
    if not polys:
        return None
    # nettoie et unionne
    return _sh_union([p.buffer(0) for p in polys])


def _build_local_triangle(OB: float, OL: float, BL: float) -> dict:
    """
    Construit un triangle local avec O=(0,0), B=(OB,0), L=(x,y) à partir des 3 longueurs.
    """
    x = (OL*OL - BL*BL + OB*OB) / (2*OB)
    y2 = max(0.0, OL*OL - x*x)
    y = math.sqrt(y2)
    return {
        "O": np.array([0.0,     0.0], dtype=float),
        "B": np.array([float(OB), 0.0], dtype=float),
        "L": np.array([float(x),  float(y)], dtype=float),
    }


class ScenarioAssemblage:
    """
    Représente un scénario d'assemblage (manuel ou automatique).
    Un scénario contient :
      - un nom lisible (ex. 'Scénario manuel', 'Auto #1'),
      - un type de source ('manual' / 'auto'),
      - éventuellement l'identifiant de l'algorithme qui l'a généré,
      - la liste des triangles d'origine (ids),
      - l'état géométrique : triangles dessinés + groupes.

    Pour le scénario manuel, last_drawn / groups sont des références
    directes vers les structures runtime du viewer (pas une copie).
    Pour les scénarios automatiques, on utilisera plutôt des copies.
    """
    def __init__(self, name: str, source_type: str = "manual", algo_id: Optional[str] = None, tri_ids: Optional[List[int]] = None):
        self.name: str = name
        self.source_type: str = source_type      # "manual" ou "auto"
        self.algo_id: Optional[str] = algo_id    # id de l'algo (pour les auto)
        self.tri_ids: List[int] = list(tri_ids) if tri_ids is not None else []

        # État géométrique associé au scénario
        self.last_drawn: List[Dict] = []         # même structure que viewer._last_drawn
        self.groups: Dict[int, Dict] = {}        # même structure que viewer.groups
        self.topoWorld = TopologyWorld()         # TopologyWorld

        # --- Vue & Carte ---
        # Vue: zoom + offset (pan). Stockée pour retrouver exactement l'affichage au switch.
        # Carte: état du fond (fichier + worldRect + visibilité/opacité + scale affiché).
        # Pour les scénarios automatiques, la carte peut être partagée au niveau du viewer.
        self.view_state: Dict[str, float] = {}   # {"zoom":..., "offset_x":..., "offset_y":...}
        self.map_state: Dict[str, object] = {}   # {"path":.., "x0":.., "y0":.., "w":.., "h":.., "visible":.., "opacity":.., "scale":..}

        # Statut global du scénario (utile pour les simulations auto)
        self.status: str = "complete"            # "complete", "pruned", etc.
        self.created_at: _dt.datetime = _dt.datetime.now()


# =============================================================================
# TopologyModel v4.3 (Core) – Modèle objet V1 (sans Tk) – IDs lisibles (Phase 2)
# =============================================================================
#
# Convention d’identifiants lisibles :
# - Element (triangle instance) : "T<triRank:02d>"         ex: "T01"
# - Node atomique (par sommet) : "T<triRank:02d}:N<idx>"   ex: "T01:N0"
# - Group (DSU)                : "G<k:03d>"                ex: "G001"
# - Attachment                 : "A<k:03d>"                ex: "A003"
#
# Remarques :
# - Les IDs sont des strings pour une lisibilité maximale (TopoDump XML).
# - La règle “plus récent gagne” utilise createdOrder (int monotone interne),
#   et PAS l’ordre lexicographique des IDs.
# - t est fourni par l’UI en référentiel monde (km).
# - Les points internes sont purement conceptuels (pas de points physiques sur les arêtes).
#
# Couverture V1 :
# - vertex↔vertex
# - vertex↔edge (t obligatoire)
# - edge↔edge (mapping "direct"|"reverse") + coverage total [0,1] sur les deux arêtes
#
# Dégrouper/Undo (MVP) :
# - suppression d’attaches + rebuild complet depuis la liste des attachments.
#


class TopologyNodeType:
    """Types de nœuds topologiques (métier) et priorité associée.

    Rôle :
    - Représenter le 'type' métier d’un nœud : Ouverture (O), Base (B), Lumière (L).
    - Fournir une fonction de rang pour la canonisation DSU : L > B > O.

    Ne fait PAS :
    - Ne stocke aucune donnée d’instance (simple namespace de constantes).

    V1 :
    - La règle de priorité L > B > O est figée par la spec v4.2.
    """
    OUVERTURE = "O"
    BASE = "B"
    LUMIERE = "L"

    @staticmethod
    def rank(node_type: str) -> int:
        if node_type == TopologyNodeType.LUMIERE:
            return 3
        if node_type == TopologyNodeType.BASE:
            return 2
        return 1


class TopologyFeatureType:
    """Types de features référencées par TopologyFeatureRef.

    Rôle :
    - Distinguer explicitement les deux natures de cible manipulées par la topologie :
      - vertex (sommet d’un élément)
      - edge (arête d’un élément)

    Ne fait PAS :
    - Ne contient aucune logique topo ; c’est un namespace de constantes.

    V1 :
    - Suffisant pour vertex↔vertex, vertex↔edge et edge↔edge.
    """
    VERTEX = "vertex"
    EDGE = "edge"


class TopologyFeatureRef:
    """Référence typée et sérialisable vers une feature topologique.

    Rôle :
    - Pointer de manière stable vers une feature (Vertex ou Edge) dans un élément donné :
      (feature_type, element_id, index).
    - Permettre de stocker/rejouer des TopologyAttachment sans dépendre d’objets Python en mémoire.
    - Servir directement à la sérialisation (TopoDump XML) : export lisible et diffable.

    Ne fait PAS :
    - Ne porte aucune décision topologique (pas de fusion, pas de merge).
    - Ne dépend pas du canvas Tk ; aucun calcul géométrique.

    Note :
    - Le champ optionnel t est une commodité de debug/trace. Dans la spec v4.2,
      t est fourni par l’UI (référentiel monde, km) et consommé par le Core.

    V1 :
    - Utilisée comme brique de base pour les attachments et l’export TopoDump.
    """
    def __init__(self, feature_type: str, element_id: str, index: int, t: float | None = None):
        self.feature_type = str(feature_type)
        self.element_id = str(element_id)
        self.index = int(index)
        self.t = float(t) if t is not None else None

    def to_string(self) -> str:
        if self.t is None:
            return f"{self.feature_type}({self.element_id},{self.index})"
        return f"{self.feature_type}({self.element_id},{self.index},t={self.t:.6f})"


class TopologyAttachment:
    """Trace métier d’une intention topologique binaire (action rejouable).

    Rôle :
    - Représenter UNE action topologique atomique entre deux features (A ↔ B).
    - Constituer la source de vérité pour reconstruire la topologie (rebuild depuis la liste d’attaches).
    - Être sérialisable (TopoDump XML) et diffable (par scénario).

    Ne fait PAS :
    - Ne modifie pas directement le modèle : l’application est effectuée par TopologyWorld.
    - Ne stocke pas de coordonnées (la géométrie est hors topo ; t est fourni par l’UI).

    V1 :
    - Stockage + export.
    - Supporte vertex↔vertex, vertex↔edge et edge↔edge (mapping direct|reverse).
    - La reconstruction (rebuild) depuis la liste d’attaches est la stratégie de référence.
    """
    def __init__(self, attachment_id: str | None, kind: str,
                 feature_a: TopologyFeatureRef, feature_b: TopologyFeatureRef,
                 params: dict | None = None, source: str = "manual"):
        self.attachment_id = None if attachment_id is None else str(attachment_id)
        self.kind = str(kind)
        self.feature_a = feature_a
        self.feature_b = feature_b
        self.params = dict(params) if params is not None else {}
        self.source = str(source)


class TopologyVertex:
    """Sommet d’un élément (polygone), labellisé (ville), rattaché à un node atomique.

    Rôle :
    - Porter le label (nom de ville) utilisé pour l’affichage et les clés de matching.
    - Référencer le node atomique d’origine (node_id) et son type (O/B/L) à la création.
    - Fournir une identité stable : (element_id, vertex_index).

    Ne fait PAS :
    - Ne calcule pas la topologie ; il ne fait que référencer le node d’origine.
    - Ne dépend pas de la géométrie (pas de coordonnées).

    V1 :
    - Les unions (DSU) opèrent au niveau TopologyWorld via node_id (atomique) -> node canonique.
    """
    def __init__(self, element_id: str, vertex_index: int, label: str, node_id: str, node_type: str):
        self.element_id = str(element_id)
        self.vertex_index = int(vertex_index)
        self.label = str(label)
        self.node_id = str(node_id)
        self.node_type = str(node_type)


class TopologyCoverageInterval:
    """Intervalle de couverture sur une arête (portion interne), exprimé en t.

    Rôle :
    - Modéliser des portions d’arêtes devenues internes (collage edge↔edge, etc.),
      sous forme d’intervalles [t0, t1].
    - Permettre le calcul du boundary comme complément des coverages.

    Ne fait PAS :
    - Ne décide pas quand une couverture apparaît : cela dépend des attachments et des règles métier.

    V1 :
    - Structure utilisée.
    - Un collage edge↔edge produit un coverage total [0.0, 1.0] sur chacune des deux arêtes collées.
    - Le boundary sera calculé comme le complément de l’union des coverages sur [0,1].
    """
    def __init__(self, t0: float, t1: float):
        self.t0 = float(t0)
        self.t1 = float(t1)


class TopologyEdge:
    """Arête paramétrique d’un élément (Option A) : endpoints + coverages.

    Rôle :
    - Représenter une arête orientée (v_start -> v_end) avec des endpoints uniquement.
    - Fournir des clés de matching (edgeLabelKey) à partir des labels de sommets.
    - Porter des coverages (intervalles internes) pour le calcul du boundary.

    Ne fait PAS :
    - Ne fait pas de DSU : la canonisation des nodes est dans TopologyWorld.
    - Ne dépend pas du canvas Tk ; aucune logique d’affichage.

    V1 :
    - Les split points sont strictement conceptuels (pas de points physiques).
    - Le calcul complet du boundary sera implémenté ultérieurement.
    """
    def __init__(self, element_id: str, edge_index: int, v_start: TopologyVertex, v_end: TopologyVertex,
                 edge_length_km: float):
        self.element_id = str(element_id)
        self.edge_index = int(edge_index)
        self.v_start = v_start
        self.v_end = v_end
        self.edge_length_km = float(edge_length_km)

        self.coverages: list[TopologyCoverageInterval] = []

    def edge_id(self) -> str:
        return f"{self.element_id}:E{self.edge_index}"

    def edge_label_key(self) -> tuple[str, str]:
        a = self.v_start.label
        b = self.v_end.label
        return (a, b) if a <= b else (b, a)

    def labels_display(self) -> str:
        return f"{self.v_start.label}–{self.v_end.label}"


class TopologyPose2D:
    """Pose 2D d'un élément dans le référentiel monde (rotation + translation).

    Rôle :
    - Porter la transformation Monde <- Local, sous forme (R, T) :
        p_world = R @ p_local + T
    - Être la SEULE représentation des coordonnées monde persistées dans le modèle.
      Les coordonnées monde des sommets/points sont toujours dérivées via cette pose.

    Ne fait PAS :
    - Ne modifie pas la géométrie intrinsèque (longueurs / angles) de l'élément.
    - Ne décide pas de la pose (elle est calculée/ajustée par l'UI ou les règles d'attachments).
    """
    def __init__(self, R: np.ndarray | None = None, T: np.ndarray | None = None):
        self.R = np.array(R, float) if R is not None else np.eye(2, dtype=float)
        self.T = np.array(T, float) if T is not None else np.zeros(2, dtype=float)

    def theta_rad(self) -> float:
        # R = [[c, -s],[s, c]]
        return float(math.atan2(self.R[1, 0], self.R[0, 0]))

    def to_dict(self) -> dict:
        return {
            "R": self.R.tolist(),
            "T": self.T.tolist(),
            "thetaRad": self.theta_rad(),
        }

    def compose(self, other: "TopologyPose2D") -> "TopologyPose2D":
        """Compose deux poses : self ∘ other.
        Si p' = other(p) et p'' = self(p'), alors compose() retourne la pose p'' = (self ∘ other)(p).
        """
        R = self.R @ other.R
        T = (self.R @ other.T) + self.T
        return TopologyPose2D(R=R, T=T)

    @staticmethod
    def identity() -> "TopologyPose2D":
        return TopologyPose2D(R=np.eye(2, dtype=float), T=np.zeros(2, dtype=float))


class TopologyElementPose2D:
    """Pose 2D d'un élément : rotation (det=+1) + translation + réflexion locale.

    Décision de modèle (session) :
    - Les coordonnées locales des éléments (vertex_local_xy) ne sont jamais modifiées.
    - La pose monde est portée par l'élément (pas par les groupes).

    Convention stable (local -> world) :
      - si mirrored == False : p_world = R @ p_local + T
      - si mirrored == True  : p_world = R @ (M @ p_local) + T
        avec M = diag(-1, +1) dans le repère local.

    Note : on impose R ~ rotation pure (det ~ +1) et on stocke toute réflexion
    dans le flag mirrored.
    """

    def __init__(self, R: np.ndarray | None = None, T: np.ndarray | None = None, mirrored: bool = False):
        self.R = np.array(R, float) if R is not None else np.eye(2, dtype=float)
        self.T = np.array(T, float) if T is not None else np.zeros(2, dtype=float)
        self.mirrored = bool(mirrored)

    @staticmethod
    def identity() -> "TopologyElementPose2D":
        return TopologyElementPose2D(R=np.eye(2, dtype=float), T=np.zeros(2, dtype=float), mirrored=False)

    @staticmethod
    def mirror_matrix() -> np.ndarray:
        return np.array([[1.0, 0.0], [0.0, -1.0]], dtype=float)

    def local_to_world(self, p_local: np.ndarray) -> np.ndarray:
        p = np.array(p_local, dtype=float)
        if self.mirrored:
            p = self.mirror_matrix() @ p
        return (self.R @ p) + self.T


class TopologyElement:
    """Élément topologique : polygone (triangle = cas particulier n=3).

    Rôle :
    - Regrouper les labels de sommets (villes) et leurs types (O/B/L) dans un cycle ordonné.
    - Porter les longueurs d’arêtes en km (monde) nécessaires à la politique de fusion sur t.
    - Servir de conteneur pour les TopologyVertex et TopologyEdge instanciés par TopologyWorld.

    Ne fait PAS :
    - Ne crée pas lui-même les nodes : TopologyWorld crée un node atomique par sommet.
    - Ne dépend pas du canvas Tk (pixels/zoom/pan).

    Notes (v4.3) :
    - La 'vérité' géométrique du triangle est intrinsèque (longueurs / angles).
    - La pose monde (rotation + translation) permet de dériver des coordonnées monde si nécessaire.
    - Les coordonnées monde ne sont pas une vérité métier ; elles sont dérivées.

    V1 :
    - edge_lengths_km peut être fourni (recommandé). À défaut, une valeur par défaut est posée
      et remplacée ensuite par la géométrie monde.
    """
    def __init__(self, element_id: str, name: str,
                 vertex_labels: list[str], vertex_types: list[str],
                 edge_lengths_km: list[float],
                 intrinsic_sides_km: dict[str, float] | None = None,
                 local_frame: dict | None = None,
                 vertex_local_xy: dict[int, tuple[float, float]] | None = None,
                 meta: dict | None = None):
        self.element_id = str(element_id)
        self.name = str(name)
        self.meta = dict(meta) if meta is not None else {}

        if len(vertex_labels) < 3:
            raise ValueError("TopologyElement: un polygone doit avoir au moins 3 sommets")
        if len(vertex_labels) != len(vertex_types):
            raise ValueError("TopologyElement: vertex_types doit avoir la même taille que vertex_labels")
        if len(edge_lengths_km) != len(vertex_labels):
            raise ValueError("TopologyElement: edge_lengths_km doit avoir la même taille que les sommets")

        self.vertex_labels = [str(x) for x in vertex_labels]
        self.vertex_types = [str(x) for x in vertex_types]
        self.edge_lengths_km = [float(x) for x in edge_lengths_km]

        self.intrinsic_sides_km: dict[str, float] = dict(intrinsic_sides_km) if intrinsic_sides_km is not None else {}

        # --- Repère local (coordonnées relatives) ---
        # Convention recommandée (v4.3+):
        # - origine au sommet Ouverture (O) -> (0,0)
        # - axe X aligné sur le segment O->B (si présent)
        # Les coordonnées locales sont déterministes et dérivées de la géométrie intrinsèque.
        self.local_frame: dict = dict(local_frame) if local_frame is not None else {
            "origin": TopologyNodeType.OUVERTURE,
            "xAxis": "O->B",
            "units": "km",
        }

        # vertex_local_xy : coordonnées relatives des sommets (par index), exprimées dans local_frame.
        # Ex: { idxO:(0,0), idxB:(OB,0), idxL:(x,y) }
        self.vertex_local_xy: dict[int, tuple[float, float]] = dict(vertex_local_xy) if vertex_local_xy is not None else {}

        if not self.vertex_local_xy:
            self._try_build_default_local_coords()

        self.vertexes: list[TopologyVertex] = []
        self.edges: list[TopologyEdge] = []

        # --- Pose monde (par élément) ---
        # IMPORTANT: les coordonnées locales (vertex_local_xy) restent figées.
        # La position/orientation en monde est exclusivement portée par cette pose.
        self.pose: TopologyElementPose2D = TopologyElementPose2D.identity()

    # --- Pose API (élément) ---
    def get_pose(self) -> tuple[np.ndarray, np.ndarray, bool]:
        return (np.array(self.pose.R, float), np.array(self.pose.T, float), bool(self.pose.mirrored))

    def set_pose(self, R: np.ndarray, T: np.ndarray, mirrored: bool | None = None) -> None:
        self.pose.R = np.array(R, float)
        self.pose.T = np.array(T, float)
        if mirrored is not None:
            self.pose.mirrored = bool(mirrored)

    def localToWorld(self, p_local: np.ndarray | tuple[float, float]) -> np.ndarray:
        return self.pose.local_to_world(np.array(p_local, dtype=float))

    def _try_build_default_local_coords(self):
        """Construit des coordonnées locales canoniques si l'élément est un triangle O/B/L.

        Stratégie V1 :
        - On détecte les indices des sommets de type O, B, L.
        - On reconstruit les longueurs OB, OL, BL depuis edge_lengths_km (cycle).
        - On pose O=(0,0), B=(OB,0), L=(x,y) (y>=0) via la loi des cosinus.

        Si l'élément n'est pas un triangle O/B/L (hors-scope V1), la méthode ne fait rien.
        Si l'élément EST un triangle O/B/L mais que les longueurs nécessaires sont absentes,
        c'est un problème de données et on lève une exception explicite (pas de masquage).
        """
        if len(self.vertex_labels) != 3:
            return

        # Hors-scope V1 : si on n'a pas exactement un O, un B et un L, on ne force pas.
        if (TopologyNodeType.OUVERTURE not in self.vertex_types) or \
           (TopologyNodeType.BASE not in self.vertex_types) or \
           (TopologyNodeType.LUMIERE not in self.vertex_types):
            return

        # indices par type (première occurrence)
        idxO = self.vertex_types.index(TopologyNodeType.OUVERTURE)
        idxB = self.vertex_types.index(TopologyNodeType.BASE)
        idxL = self.vertex_types.index(TopologyNodeType.LUMIERE)

        # map des longueurs par paire non orientée (triangle)
        n = 3
        pair_len: dict[frozenset, float] = {}
        for i in range(n):
            j = (i + 1) % n
            pair_len[frozenset((i, j))] = float(self.edge_lengths_km[i])

        def d(i, j) -> float:
            return float(pair_len.get(frozenset((i, j)), 0.0))

        OB = d(idxO, idxB)
        OL = d(idxO, idxL)
        BL = d(idxB, idxL)
        if OB <= 0 or OL <= 0 or BL <= 0:
            raise ValueError(
                f"TopologyElement({self.element_id}): longueurs invalides pour repère local O/B/L "
                f"(OB={OB}, OL={OL}, BL={BL})"
            )

        P = _build_local_triangle(OB=OB, OL=OL, BL=BL)

        # Convention import historique : si orient == "CW", miroir par rapport à l'axe X local.
        # (équivalent à P["L"][1] *= -1 dans l'ancien code Tk)
        orient = str(self.meta.get("orient", "")).strip().upper()
        if orient == "CW":
            P["L"][1] = -float(P["L"][1])

        # Remap vers les indices réels
        self.vertex_local_xy[idxO] = (float(P["O"][0]), float(P["O"][1]))
        self.vertex_local_xy[idxB] = (float(P["B"][0]), float(P["B"][1]))
        self.vertex_local_xy[idxL] = (float(P["L"][0]), float(P["L"][1]))

        return

    def n_vertices(self) -> int:
        return len(self.vertex_labels)


class TopologyGroup:
    """Groupe topologique : composante connexe d’éléments (unité de manipulation).

    Rôle :
    - Représenter un ensemble d’éléments reliés par des attachments (groupe d’assemblage).
    - Porter la liste des elementIds et des attachmentIds associés au groupe canonique.

    Ne fait PAS :
    - Ne calcule pas la connectivité : TopologyWorld gère l’union-find des groupes.
    - Ne stocke pas de géométrie.

    V1 :
    - Structure légère : l’objectif est d’avoir un conteneur stable pour l’export TopoDump.
    """
    def __init__(self, group_id: str):
        self.group_id = str(group_id)
        self.element_ids: list[str] = []
        self.attachment_ids: list[str] = []


@dataclass
class ConceptNodeInfo:
    concept_id: str
    members: set[str]
    sumx: float = 0.0
    sumy: float = 0.0
    count: int = 0

    def add_occ(self, x: float, y: float) -> None:
        self.sumx += float(x)
        self.sumy += float(y)
        self.count += 1

    def world_xy(self) -> tuple[float, float]:
        if self.count <= 0:
            return (0.0, 0.0)
        return (self.sumx / self.count, self.sumy / self.count)


@dataclass
class ConceptEdgeInfo:
    a: str
    b: str
    occurrences: list[dict]


@dataclass
class TopologyBoundaries:
    cycle: list[str] | None = None
    edges: set[tuple[str, str]] | None = None
    index: dict[str, int] | None = None
    orientation: str | None = None

    def clear(self) -> None:
        self.cycle = None
        self.edges = None
        self.index = None
        self.orientation = None

    def inverse(self) -> None:
        """Inverse le cycle boundary (reverse order) et réoriente les edges."""
        if not self.cycle:
            return
        cycle = list(reversed(self.cycle))
        self.cycle = cycle
        self.index = {n: i for i, n in enumerate(cycle)}
        if self.edges is not None:
            self.edges = {(b, a) for (a, b) in self.edges}
        if self.orientation == "cw":
            self.orientation = "ccw"
        elif self.orientation == "ccw":
            self.orientation = "cw"

    def compute(self, world: "TopologyWorld", gid: str, orientation: str) -> None:
        """
        Overview
        --------
        Calcule le contour (boundary) d’un graphe conceptuel plan (un groupe topologique),
        sous la forme d’un cycle de nœuds conceptuels ordonné (CW/CCW). Ce cycle sert ensuite
        de référence unique pour reconstruire des rings géométriques et effectuer des découpes
        (arcs) lors des simulations d’anti-chevauchement.

        Entrées
        -------
        world : TopologyWorld
            Monde topo contenant le cache conceptuel (nodes/edges + coordonnées monde).
        gid : str
            Identifiant de groupe (canonisé en interne).
        orientation : str
            Orientation attendue pour le cycle final: "cw" ou "ccw".

        Sorties
        -------
        None
            Remplit les champs de l’instance:
            - self.cycle : list[str]
            - self.edges : set[tuple[str,str]]
            - self.index : dict[str,int]
            - self.orientation : str
            En cas de graphe dégénéré, produit un cycle vide et des structures vides.

        Traitement
        ----------
        1) Normalise l’orientation demandée ("cw"/"ccw"), récupère le cache conceptuel du groupe.
        2) Construit l’adjacence non orientée `neighbors` à partir des arêtes conceptuelles.
           - Cas dégénérés: graphe vide ou nœuds de degré < 2 → boundary vide.
        3) Récupère les coordonnées monde (barycentres) des nœuds conceptuels.
        4) Pour chaque nœud, ordonne ses voisins par angle polaire autour du nœud.
           - Produit `sorted_neighbors` et un index local `pos_index` pour le suivi de face.
        5) Énumère les "darts" (u→v) et parcourt les faces via une règle de tournage:
           - En CW : à l’arrivée en v, prendre le voisin précédent autour de v.
           - En CCW : à l’arrivée en v, prendre le voisin suivant autour de v.
           Chaque parcours génère un cycle candidat (face) si on se referme sur le dart initial.
        6) Filtre les faces:
           - longueur >= 3
           - aire signée non nulle (|area| > eps)
           - cycle simple (pas de nœud répété)
        7) Sélectionne la face "extérieure" par aire absolue maximale.
        8) Normalise le cycle choisi (rotation) pour démarrer sur le nœud le plus "bas-gauche"
           (min x puis min y), afin d’avoir un cycle stable/diffable.
        9) Construit:
           - `cycle` (liste des nœuds),
           - `edges` (ensemble des arêtes du contour),
           - `index` (nodeId → position dans le cycle),
           - `orientation` (orientation mesurée via l’aire).
           Si l’orientation mesurée ne correspond pas à celle demandée, inverse le cycle.
        """
        orient = TopologyChemins._normalizeOrientation(orientation)
        gid = world.find_group(str(gid))
        c = world.ensureConceptGeom(gid)

        neighbors: dict[str, set[str]] = {}
        for (a, b) in c.edges.keys():
            neighbors.setdefault(a, set()).add(b)
            neighbors.setdefault(b, set()).add(a)

        if not neighbors:
            self.cycle = []
            self.edges = set()
            self.index = {}
            self.orientation = orient
            return

        for n, adj in neighbors.items():
            if len(adj) < 2:
                self.cycle = []
                self.edges = set()
                self.index = {}
                self.orientation = orient
                return

        coords = {}
        for n in neighbors.keys():
            coords[n] = world.getConceptNodeWorldXY(n, gid)

        sorted_neighbors: dict[str, list[str]] = {}
        pos_index: dict[str, dict[str, int]] = {}
        for u, adj in neighbors.items():
            ux, uy = coords[u]
            ordered = sorted(
                list(adj),
                key=lambda v: math.atan2(float(coords[v][1]) - float(uy), float(coords[v][0]) - float(ux))
            )
            sorted_neighbors[u] = ordered
            pos_index[u] = {v: i for i, v in enumerate(ordered)}

        def _next_dart(u: str, v: str) -> tuple[str, str]:
            nb = sorted_neighbors.get(v, [])
            if not nb:
                raise ValueError(f"[Boundary] missing neighbors for node {v}")
            i = pos_index[v].get(u, None)
            if i is None:
                raise ValueError(f"[Boundary] missing neighbor {u} around {v}")
            if orient == "cw":
                w = nb[(i - 1) % len(nb)]
            else:
                w = nb[(i + 1) % len(nb)]
            return (v, w)

        darts = []
        for u, adj in neighbors.items():
            for v in adj:
                if u != v:
                    darts.append((u, v))

        visited: set[tuple[str, str]] = set()
        faces: list[tuple[list[str], float]] = []
        max_steps = len(darts) + 1
        eps_area = 1e-12

        def _signed_area(cycle: list[str]) -> float:
            area = 0.0
            for i in range(len(cycle)):
                x0, y0 = coords[cycle[i]]
                x1, y1 = coords[cycle[(i + 1) % len(cycle)]]
                area += (float(x0) * float(y1)) - (float(x1) * float(y0))
            return 0.5 * area

        for d0 in darts:
            if d0 in visited:
                continue
            local_darts: list[tuple[str, str]] = []
            local_set: set[tuple[str, str]] = set()
            nodes = [d0[0]]
            cur = d0
            steps = 0
            closed = False
            while True:
                if cur in visited:
                    break
                if cur in local_set:
                    break
                local_darts.append(cur)
                local_set.add(cur)
                nodes.append(cur[1])
                steps += 1
                if steps > max_steps:
                    break
                nxt = _next_dart(cur[0], cur[1])
                if nxt == d0:
                    closed = True
                    break
                cur = nxt
            if not closed:
                continue
            if nodes and nodes[-1] == nodes[0]:
                nodes = nodes[:-1]
            if len(nodes) < 3:
                continue
            area = _signed_area(nodes)
            if abs(area) <= eps_area:
                continue
            visited.update(local_darts)
            faces.append((nodes, area))

        if not faces:
            self.cycle = []
            self.edges = set()
            self.index = {}
            self.orientation = orient
            return

        def _is_simple(nodes: list[str]) -> bool:
            if not nodes:
                return False
            if nodes[-1] == nodes[0]:
                nodes = nodes[:-1]
            if len(nodes) < 3:
                return False
            if len(nodes) != len(set(nodes)):
                return False
            return True

        faces_simple = [(face, area) for (face, area) in faces if _is_simple(face)]
        if not faces_simple:
            self.cycle = []
            self.edges = set()
            self.index = {}
            self.orientation = orient
            return

        outer = None
        outer_area = None
        for face, area in faces_simple:
            if outer_area is None or abs(area) > abs(outer_area):
                outer = face
                outer_area = area

        if outer is None or outer_area is None:
            self.cycle = []
            self.edges = set()
            self.index = {}
            self.orientation = orient
            return

        cycle = list(outer)
        best_idx = 0
        best_x, best_y = coords[cycle[0]]
        for i in range(1, len(cycle)):
            x, y = coords[cycle[i]]
            if (x < best_x) or (x == best_x and y < best_y):
                best_idx = i
                best_x, best_y = x, y
        cycle = cycle[best_idx:] + cycle[:best_idx]

        edges = set()
        if len(cycle) >= 2:
            for i in range(len(cycle)):
                a = cycle[i]
                b = cycle[(i + 1) % len(cycle)]
                k = (a, b) if a < b else (b, a)
                if k not in c.edges:
                    raise ValueError(f"[Boundary] missing concept edge for {a}-{b}")
                edges.add((a, b))

        self.cycle = cycle
        self.edges = edges
        self.index = {n: i for i, n in enumerate(cycle)}

        signed = _signed_area(cycle)
        computed = "cw" if signed < 0.0 else "ccw"
        self.orientation = computed

        # Si l'orientation finale du graph n'est finalement pas celle demandée en entrée,
        # on inverse le cycle
        if computed != orient:
            self.inverse()
            self.orientation = orient   # contrainte forte


@dataclass
class BoundarySegment:
    conceptA: str
    conceptB: str
    elementId: str
    edgeIndex: int
    fromNodeId: str
    toNodeId: str
    t0: float
    t1: float


@dataclass
class ConceptGroupCache:
    """Cache du modèle conceptuel pour un groupe canonique."""
    graphValid: bool = False
    geomValid: bool = False
    nodes: dict[str, ConceptNodeInfo] = None
    edges: dict[tuple[str, str], ConceptEdgeInfo] = None
    nodeOccurrencesByCid: dict[str, list[tuple[str, tuple[float, float]]]] = None
    boundaryCycle: list[str] | None = None
    boundaryEdges: set[tuple[str, str]] | None = None
    boundaryIndex: dict[str, int] | None = None
    boundaryOrientation: str | None = None
    boundaries: TopologyBoundaries | None = None

    def __post_init__(self) -> None:
        if self.nodes is None:
            self.nodes = {}
        if self.edges is None:
            self.edges = {}
        if self.nodeOccurrencesByCid is None:
            self.nodeOccurrencesByCid = {}
        if self.boundaries is None:
            self.boundaries = TopologyBoundaries()


class TopologyCheminTriplet:
    """Triplet de chemin A-O-B (structure + geometrie calculee a la demande)."""

    def __init__(self, nodeA: str, nodeO: str, nodeB: str) -> None:
        """Construit un triplet combinatoire avec geometrie invalide par defaut."""
        self.nodeA = str(nodeA)
        self.nodeO = str(nodeO)
        self.nodeB = str(nodeB)
        if len({self.nodeA, self.nodeO, self.nodeB}) != 3:
            raise ValueError("[Chemins] triplet invalide: nodes non distincts")

        # Etat normatif a la creation combinatoire.
        self.azOA = 0.0
        self.azOB = 0.0
        self.distOA_km = 0.0
        self.distOB_km = 0.0
        self.angleDeg = 0.0
        self.isGeometrieValide = False

    def calculerGeometrie(
        self,
        world: "TopologyWorld",
        groupId: str,
        orientationUser: str,
        baliseRefName: str,
    ) -> None:
        """Calcule azimuts, distances et angle orienté à partir du world courant.

        azOA/azOB sont exprimés RELATIVEMENT à l'axe de référence défini par:
            axe 0° = (O -> baliseRefName)
        """
        # Balise de référence utilisée par le Core pour définir l'axe 0°
        gid = str(groupId)
        orient = TopologyChemins._normalizeOrientation(orientationUser)
        balise_name = str(baliseRefName or "").strip()
        if not balise_name:
            raise ValueError("[Chemins][Triplet] baliseRefName invalide (vide)")
        if not world.hasBalise(balise_name):
            raise ValueError(f"[Chemins][Triplet] balise de reference introuvable: '{balise_name}'")

        # --- Points du triplet (World XY) ---
        xA, yA = world.getConceptNodeWorldXY(self.nodeA, gid)
        xO, yO = world.getConceptNodeWorldXY(self.nodeO, gid)
        xB, yB = world.getConceptNodeWorldXY(self.nodeB, gid)

        # --- Balise de référence (World XY) ---
        xR, yR = world.getBaliseWorldXY(balise_name)

        # Axe de référence = O -> R
        dxOR = float(xR - xO)
        dyOR = float(yR - yO)
        azAxis = world.azimutDegFromDxDy(dxOR, dyOR)

        # NOTE (debug / cohérence avec ton point 4) :
        # azRef = azimut(Balise -> O) = world.azimutDegFromDxDy(xO-xR, yO-yR)
        # (on ne l'utilise pas pour la définition de l'axe 0°, juste pour référence si besoin)

        # --- Vecteurs OA / OB ---
        dxOA = float(xA - xO)
        dyOA = float(yA - yO)
        dxOB = float(xB - xO)
        dyOB = float(yB - yO)

        azOA_north = world.azimutDegFromDxDy(dxOA, dyOA)
        azOB_north = world.azimutDegFromDxDy(dxOB, dyOB)

        # Azimuts RELATIFS à l'axe O->R (pliés dans [0°, 180°])
        azOA_rel = float((azOA_north - azAxis + 360.0) % 360.0)
        azOB_rel = float((azOB_north - azAxis + 360.0) % 360.0)

        # Distances inchangées (World = km dans ton modèle actuel)
        self.distOA_km = float(math.hypot(dxOA, dyOA))
        self.distOB_km = float(math.hypot(dxOB, dyOB))

        angle = (azOB_rel - azOA_rel + 360.0) % 360.0
        if orient == "ccw":
            angle = float((360.0 - angle) % 360.0)
        self.angleDeg = angle

        # On repasse à 180° pour la persistance des azimuts (apres le calcul du delta)
        if azOA_rel > 180.0:
            azOA_rel = 360.0 - azOA_rel
        if azOB_rel > 180.0:
            azOB_rel = 360.0 - azOB_rel

        self.azOA = azOA_rel
        self.azOB = azOB_rel

        self.isGeometrieValide = True

    @staticmethod
    def getMesuresSpecs() -> list[dict]:
        """Spécifications UI des mesures disponibles pour l'affichage des triplets.
            Clés attendues côté UI :
            - azOA, azOB, angle, distOA, distOB
        """
        return [
             {"key": "azOA",   "label": "Az OA°",        "kind": "angle"},
             {"key": "azOB",   "label": "Az OB°",        "kind": "angle"},
             {"key": "angle",  "label": "(A, O, B)°",    "kind": "angle"},
             {"key": "distOA", "label": "Dist [OA]",     "kind": "distance"},
             {"key": "distOB", "label": "Dist [OB]",     "kind": "distance"},
        ]


class TopologyChemins:
    """Chemin unique du scenario : snapshot, mask, ordre, triplets et geometrie."""

    def __init__(self, world: "TopologyWorld") -> None:
        """Initialise un conteneur de chemin attache a un TopologyWorld."""
        self._world = world
        self._resetUndefined()

    def _resetUndefined(self) -> None:
        self.isDefined: bool = False
        self.groupId: str | None = None
        self.startNodeId: str | None = None
        self.orientationUser: str | None = None
        self.borderSnapshotNodes: list[str] = []
        self.selectionMask: list[bool] = []
        self.pathNodesOrdered: list[str] = []
        self.triplets: list[TopologyCheminTriplet] = []

    @staticmethod
    def _normalizeOrientation(orientationUser: str) -> str:
        orient = str(orientationUser).strip().lower()
        if orient not in ("cw", "ccw"):
            raise ValueError(f"[Chemins] orientationUser invalide: {orientationUser}")
        return orient

    @staticmethod
    def _buildPathNodesOrdered(
        borderSnapshotNodes: list[str],
        selectionMask: list[bool],
        orientationUser: str,
        boundaryOrientation: str,
    ) -> list[str]:
        """Construit pathNodesOrdered sans modifier snapshot ni mask."""
        if len(selectionMask) != len(borderSnapshotNodes):
            raise ValueError("[Chemins] mask/snapshot tailles differentes")
        if orientationUser not in ("cw", "ccw"):
            raise ValueError(f"[Chemins] orientationUser invalide: {orientationUser}")
        if boundaryOrientation not in ("cw", "ccw"):
            raise ValueError(f"[Chemins] boundaryOrientation invalide: {boundaryOrientation}")

        # Invariant spec: filtrage dans l'ordre snapshot, puis inversion eventuelle.
        filtered = [
            str(borderSnapshotNodes[i])
            for i in range(len(borderSnapshotNodes))
            if bool(selectionMask[i])
        ]
        if orientationUser == boundaryOrientation:
            return filtered
        # Inversion "pivot A": le start node (index 0 du snapshot filtre) reste en tete.
        return filtered[:1] + list(reversed(filtered[1:]))

    @staticmethod
    def _buildTriplets(pathNodesOrdered: list[str]) -> list[TopologyCheminTriplet]:
        """Construit les triplets combinatoires (pas de 2, couverture complete)."""
        L = [str(n) for n in list(pathNodesOrdered)]
        n = len(L)
        if n < 3:
            raise ValueError("[Chemins] calculerTriplets: pathNodesOrdered < 3")

        covered: set[str] = set()
        target: set[str] = set(L)
        triplets: list[TopologyCheminTriplet] = []
        i = 0
        while covered != target:
            A = L[i % n]
            O = L[(i + 1) % n]
            B = L[(i + 2) % n]
            if len({A, O, B}) != 3:
                raise RuntimeError(f"[Chemins] triplet degenere: ({A},{O},{B})")
            triplets.append(TopologyCheminTriplet(nodeA=A, nodeO=O, nodeB=B))
            covered.add(A)
            covered.add(O)
            covered.add(B)
            i += 2
        return triplets

    def _assertDefinedInvariants(self) -> None:
        """Verifie les invariants d'etat pour les phases suivantes."""
        if not self.isDefined:
            if self.groupId is not None or self.startNodeId is not None or self.orientationUser is not None:
                raise RuntimeError("[Chemins] invariant brise: undefined avec ids non nuls")
            if self.borderSnapshotNodes or self.selectionMask or self.pathNodesOrdered or self.triplets:
                raise RuntimeError("[Chemins] invariant brise: undefined avec contenu")
            return

        if self.groupId is None or self.startNodeId is None or self.orientationUser is None:
            raise RuntimeError("[Chemins] invariant brise: chemin defini incomplet")
        if len(self.borderSnapshotNodes) < 3:
            raise RuntimeError("[Chemins] invariant brise: snapshot trop court")
        if self.borderSnapshotNodes[0] != self.startNodeId:
            raise RuntimeError("[Chemins] invariant brise: snapshot non pivote sur startNodeId")
        if len(self.selectionMask) != len(self.borderSnapshotNodes):
            raise RuntimeError("[Chemins] invariant brise: mask/snapshot tailles differentes")
        if not all(isinstance(v, bool) for v in self.selectionMask):
            raise RuntimeError("[Chemins] invariant brise: selectionMask doit etre booleen")
        if len(self.pathNodesOrdered) < 3:
            raise RuntimeError("[Chemins] invariant brise: pathNodesOrdered trop court")
        boundary_orientation = self._world.getBoundaryOrientation(str(self.groupId))
        expected_path = self._buildPathNodesOrdered(
            self.borderSnapshotNodes,
            self.selectionMask,
            str(self.orientationUser),
            boundary_orientation,
        )
        if self.pathNodesOrdered != expected_path:
            raise RuntimeError("[Chemins] invariant brise: pathNodesOrdered incoherent")
        covered: set[str] = set()
        for t in self.triplets:
            if len({t.nodeA, t.nodeO, t.nodeB}) != 3:
                raise RuntimeError("[Chemins] invariant brise: triplet non distinct")
            covered.add(t.nodeA)
            covered.add(t.nodeO)
            covered.add(t.nodeB)
        if not set(self.pathNodesOrdered).issubset(covered):
            raise RuntimeError("[Chemins] invariant brise: couverture triplets incomplete")

    def _requireDefined(self) -> None:
        if not self.isDefined:
            raise ValueError("[Chemins] chemin non defini")

    def creerDepuisGroupe(self, groupId: str, startNodeId: str, orientationUser: str, baliseRefName: str) -> None:
        """Cree un chemin depuis un groupe + start node + orientation utilisateur."""
        orient = self._normalizeOrientation(orientationUser)
        gid = self._world.find_group(str(groupId))
        start = self._world.find_node(str(startNodeId))

        self._world.computeBoundary(gid)
        cycle = self._world.getBoundaryCycle(gid, start)
        if len(cycle) < 3:
            raise ValueError("[Chemins] boundary invalide: moins de 3 nodes")

        # Construction en local puis commit final: aucun rollback try/except requis.
        snapshotLocal = [str(n) for n in cycle]
        maskLocal = [True] * len(snapshotLocal)
        boundary_orientation = self._world.getBoundaryOrientation(gid)
        pathLocal = self._buildPathNodesOrdered(snapshotLocal, maskLocal, orient, boundary_orientation)
        if len(pathLocal) < 3:
            raise ValueError("[Chemins] pathNodesOrdered invalide: moins de 3 nodes")
        tripletsLocal = self._buildTriplets(pathLocal)
        for t in tripletsLocal:
            t.calculerGeometrie(self._world, gid, orient, baliseRefName)

        self.isDefined = True
        self.groupId = gid
        self.startNodeId = start
        self.orientationUser = orient
        self.borderSnapshotNodes = snapshotLocal
        self.selectionMask = maskLocal
        self.pathNodesOrdered = pathLocal
        self.triplets = tripletsLocal
        self._assertDefinedInvariants()

    def appliquerEdition(self, orientationUser: str, selectionMask: list[bool], baliseRefName: str) -> None:
        """Applique orientation + mask sur le snapshot puis recalcule path/triplets/geometrie."""
        self._requireDefined()
        orient = self._normalizeOrientation(orientationUser)
        mask = [bool(v) for v in list(selectionMask)]
        if len(mask) != len(self.borderSnapshotNodes):
            raise ValueError("[Chemins] selectionMask incompatible avec snapshot")

        # Snapshot immuable en edition: seul le mask et le sens utilisateur changent.
        snapshotLocal = list(self.borderSnapshotNodes)
        boundary_orientation = self._world.getBoundaryOrientation(str(self.groupId))
        pathLocal = self._buildPathNodesOrdered(snapshotLocal, mask, orient, boundary_orientation)
        if len(pathLocal) < 3:
            raise ValueError("[Chemins] edition invalide: moins de 3 nodes selectionnes")
        tripletsLocal = self._buildTriplets(pathLocal)
        for t in tripletsLocal:
            t.calculerGeometrie(self._world, str(self.groupId), orient, baliseRefName)

        self.orientationUser = orient
        self.selectionMask = mask
        self.pathNodesOrdered = pathLocal
        self.triplets = tripletsLocal
        self._assertDefinedInvariants()

    def calculerPathNodesOrdered(self) -> None:
        """Recalcule pathNodesOrdered depuis snapshot/mask/orientationUser."""
        self._requireDefined()
        boundary_orientation = self._world.getBoundaryOrientation(str(self.groupId))
        self.pathNodesOrdered = self._buildPathNodesOrdered(
            self.borderSnapshotNodes,
            self.selectionMask,
            str(self.orientationUser),
            boundary_orientation,
        )

    def calculerTriplets(self) -> None:
        """Recalcule les triplets combinatoires a partir de pathNodesOrdered."""
        self._requireDefined()
        self.triplets = self._buildTriplets(self.pathNodesOrdered)

    def calculerGeometrieTriplets(self, baliseRefName: str) -> None:
        """Met a jour la geometrie de tous les triplets sans changer la combinatoire."""
        self._requireDefined()
        gid = str(self.groupId)
        for t in self.triplets:
            t.calculerGeometrie(self._world, gid, str(self.orientationUser), baliseRefName)

    def exportXlsx(self, filePath: str, columns: list[dict], scenarioName: str) -> None:
        """
        Exporte les triplets du chemin courant en .xlsx.

        - filePath: chemin cible du fichier .xlsx
        - columns: liste ordonnee des colonnes visibles cote UI.
          Chaque entree contient au minimum:
            - key: cle de mesure (ex: "azOA", "azOB", "angle", "distOA", "distOB", ou "triplet")
            - label: libelle UI (pour l'en-tete Excel)
            - kind: "angle" | "distance" | "text"
        - scenarioName: utilise uniquement pour le titre si necessaire (optionnel V1).
        """
        self._requireDefined()

        target_path = str(filePath).strip()
        if not target_path:
            raise ValueError("[CheminsExportXlsx] filePath vide")
        if not target_path.lower().endswith(".xlsx"):
            raise ValueError(f"[CheminsExportXlsx] extension invalide: {target_path}")

        raw_columns = list(columns or [])
        if not raw_columns:
            raise ValueError("[CheminsExportXlsx] columns vide")

        # Import local: l'export XLSX est optionnel dans le reste du Core.
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter

        value_specs_by_key = {
            "azOA": {"kind": "angle", "attr": "azOA"},
            "azOB": {"kind": "angle", "attr": "azOB"},
            "angle": {"kind": "angle", "attr": "angleDeg"},
            "distOA": {"kind": "distance", "attr": "distOA_km"},
            "distOB": {"kind": "distance", "attr": "distOB_km"},
        }

        resolved_columns: list[dict] = []
        for i, col in enumerate(raw_columns):
            if not isinstance(col, dict):
                raise TypeError(f"[CheminsExportXlsx] colonne invalide (dict attendu): index={i}")
            key = str(col.get("key", "") or "").strip()
            label = str(col.get("label", "") or "").strip()
            kind = str(col.get("kind", "") or "").strip().lower()

            if not key:
                raise ValueError(f"[CheminsExportXlsx] colonne sans key: index={i}")
            if not label:
                raise ValueError(f"[CheminsExportXlsx] colonne sans label: key={key}")

            if key == "triplet":
                if kind != "text":
                    raise ValueError(f"[CheminsExportXlsx] kind invalide pour triplet: {kind}")
                resolved_columns.append({
                    "key": key,
                    "label": label,
                    "kind": kind,
                    "attr": None,
                })
                continue

            if key not in value_specs_by_key:
                raise ValueError(f"[CheminsExportXlsx] key inconnue: {key}")

            expected = value_specs_by_key[key]
            expected_kind = str(expected["kind"])
            if kind != expected_kind:
                raise ValueError(
                    f"[CheminsExportXlsx] kind incoherent pour {key}: attendu={expected_kind}, recu={kind}"
                )

            resolved_columns.append({
                "key": key,
                "label": label,
                "kind": kind,
                "attr": str(expected["attr"]),
            })

        _ = str(scenarioName or "")

        wb = Workbook()
        ws = wb.active
        ws.title = "Chemins"
        bold_font = Font(bold=True)

        for col_idx, col in enumerate(resolved_columns, start=1):
            header_cell = ws.cell(row=1, column=col_idx, value=str(col["label"]))
            header_cell.font = bold_font

        for row_idx, t in enumerate(self.triplets, start=2):
            if not bool(t.isGeometrieValide):
                raise RuntimeError(
                    f"[CheminsExportXlsx] triplet sans geometrie valide: ({t.nodeA},{t.nodeO},{t.nodeB})"
                )

            for col_idx, col in enumerate(resolved_columns, start=1):
                key = str(col["key"])
                if key == "triplet":
                    triplet_str = (
                        f"{self._world.getNodeLabel(t.nodeA)} - "
                        f"{self._world.getNodeLabel(t.nodeO)} - "
                        f"{self._world.getNodeLabel(t.nodeB)}"
                    )
                    ws.cell(row=row_idx, column=col_idx, value=triplet_str)
                    continue

                attr_name = str(col["attr"])
                raw_value = getattr(t, attr_name)
                value = float(raw_value)
                if not math.isfinite(value):
                    raise RuntimeError(
                        f"[CheminsExportXlsx] valeur non finie: key={key}, value={raw_value}"
                    )
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.number_format = "0.00"

        for col_idx in range(1, len(resolved_columns) + 1):
            max_width = 0
            for row_idx in range(1, ws.max_row + 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                cell_text = "" if cell_value is None else str(cell_value)
                max_width = max(max_width, len(cell_text))
            ws.column_dimensions[get_column_letter(col_idx)].width = float(max_width + 2)

        wb.save(target_path)

    def _saveToXml(self, parent: _ET.Element) -> None:
        chemins_el = _ET.SubElement(parent, "chemins", {
            "isDefined": "1" if bool(self.isDefined) else "0",
        })
        if not self.isDefined:
            return

        self._assertDefinedInvariants()

        chemins_el.set("groupId", str(self.groupId))
        chemins_el.set("startNodeId", str(self.startNodeId))
        chemins_el.set("orientationUser", str(self.orientationUser))

        border_el = _ET.SubElement(chemins_el, "borderSnapshot")
        for nid in self.borderSnapshotNodes:
            _ET.SubElement(border_el, "n", {"id": str(nid)})

        mask_el = _ET.SubElement(chemins_el, "selectionMask")
        for selected in self.selectionMask:
            _ET.SubElement(mask_el, "m", {"v": "1" if bool(selected) else "0"})

        path_el = _ET.SubElement(chemins_el, "pathNodesOrdered")
        for nid in self.pathNodesOrdered:
            _ET.SubElement(path_el, "n", {"id": str(nid)})

        triplets_el = _ET.SubElement(chemins_el, "triplets")
        for t in self.triplets:
            if not bool(t.isGeometrieValide):
                raise RuntimeError(
                    f"[CheminsSave] triplet sans geometrie valide: ({t.nodeA},{t.nodeO},{t.nodeB})"
                )
            _ET.SubElement(triplets_el, "t", {
                "a": str(t.nodeA),
                "o": str(t.nodeO),
                "b": str(t.nodeB),
                "azOA": self._world._format_chemin_angle(t.azOA),
                "azOB": self._world._format_chemin_angle(t.azOB),
                "distOA_km": self._world._format_chemin_distance(t.distOA_km),
                "distOB_km": self._world._format_chemin_distance(t.distOB_km),
                "angleDeg": self._world._format_chemin_angle(t.angleDeg),
            })

    def _loadFromXml(self, cheminsNode: _ET.Element | None) -> None:
        if cheminsNode is None:
            self._resetUndefined()
            self._assertDefinedInvariants()
            return

        raw_defined = str(cheminsNode.get("isDefined", "") or "").strip()
        if raw_defined == "0":
            self._resetUndefined()
            self._assertDefinedInvariants()
            return
        if raw_defined != "1":
            raise ValueError("[CheminsLoad] isDefined invalide (attendu 0/1)")

        groupId = str(cheminsNode.get("groupId", "") or "").strip()
        startNodeId = str(cheminsNode.get("startNodeId", "") or "").strip()
        orientationUser = str(cheminsNode.get("orientationUser", "") or "").strip().lower()
        if not groupId or not startNodeId:
            raise ValueError("[CheminsLoad] groupId/startNodeId manquants")
        orient = self._normalizeOrientation(orientationUser)

        border_el = cheminsNode.find("borderSnapshot")
        if border_el is None:
            raise ValueError("[CheminsLoad] borderSnapshot manquant")
        newSnapshot: list[str] = []
        for n_el in border_el.findall("n"):
            nid = str(n_el.get("id", "") or "").strip()
            if not nid:
                raise ValueError("[CheminsLoad] borderSnapshot: id manquant")
            newSnapshot.append(nid)

        mask_el = cheminsNode.find("selectionMask")
        if mask_el is None:
            raise ValueError("[CheminsLoad] selectionMask manquant")
        newMask: list[bool] = []
        for m_el in mask_el.findall("m"):
            v = str(m_el.get("v", "") or "").strip()
            if v not in ("0", "1"):
                raise ValueError("[CheminsLoad] selectionMask: valeur invalide (attendu 0/1)")
            newMask.append(v == "1")
        if len(newMask) != len(newSnapshot):
            raise ValueError("[CheminsLoad] selectionMask/borderSnapshot tailles differentes")

        path_el = cheminsNode.find("pathNodesOrdered")
        if path_el is None:
            raise ValueError("[CheminsLoad] pathNodesOrdered manquant")
        newPath: list[str] = []
        for n_el in path_el.findall("n"):
            nid = str(n_el.get("id", "") or "").strip()
            if not nid:
                raise ValueError("[CheminsLoad] pathNodesOrdered: id manquant")
            newPath.append(nid)
        if len(newPath) < 3:
            raise ValueError("[CheminsLoad] pathNodesOrdered trop court")

        triplets_el = cheminsNode.find("triplets")
        if triplets_el is None:
            raise ValueError("[CheminsLoad] triplets manquants")
        newTriplets: list[TopologyCheminTriplet] = []
        for t_el in triplets_el.findall("t"):
            a = str(t_el.get("a", "") or "").strip()
            o = str(t_el.get("o", "") or "").strip()
            b = str(t_el.get("b", "") or "").strip()
            if not a or not o or not b:
                raise ValueError("[CheminsLoad] triplet avec ids vides")
            azOA = t_el.get("azOA")
            azOB = t_el.get("azOB")
            distOA = t_el.get("distOA_km")
            distOB = t_el.get("distOB_km")
            angleDeg = t_el.get("angleDeg")
            if azOA is None or azOB is None or distOA is None or distOB is None or angleDeg is None:
                raise ValueError("[CheminsLoad] triplet incomplet (valeurs géométriques manquantes)")
            t = TopologyCheminTriplet(nodeA=a, nodeO=o, nodeB=b)
            t.azOA = float(azOA)
            t.azOB = float(azOB)
            t.distOA_km = float(distOA)
            t.distOB_km = float(distOB)
            t.angleDeg = float(angleDeg)
            t.isGeometrieValide = True
            newTriplets.append(t)

        boundary_orientation = None
        for cand in ("cw", "ccw"):
            expected = self._buildPathNodesOrdered(newSnapshot, newMask, orient, cand)
            if expected == newPath:
                boundary_orientation = cand
                break
        if boundary_orientation is None:
            raise ValueError("[CheminsLoad] pathNodesOrdered incoherent")
        self._world._concept_cache(groupId).boundaryOrientation = boundary_orientation

        self.isDefined = True
        self.groupId = groupId
        self.startNodeId = startNodeId
        self.orientationUser = orient
        self.borderSnapshotNodes = newSnapshot
        self.selectionMask = newMask
        self.pathNodesOrdered = newPath
        self.triplets = newTriplets
        self._assertDefinedInvariants()

    def recalculerChemin(self, baliseRefName: str) -> None:
        """Reconstruit snapshot/mask/path/triplets depuis le contour courant en conservant start/sens."""
        self._requireDefined()

        oldSnapshot = list(self.borderSnapshotNodes)
        oldMask = [bool(v) for v in list(self.selectionMask)]
        oldStart = str(self.startNodeId)
        oldOrient = self._normalizeOrientation(str(self.orientationUser))

        gid = str(self.groupId)
        self._world.computeBoundary(gid)
        try:
            newCycle = self._world.getBoundaryCycle(gid, oldStart)
        except Exception as e:
            raise ValueError("[Chemins] recalcul invalide: startNodeId absent du nouveau contour") from e
        if len(newCycle) < 3:
            raise ValueError("[Chemins] recalcul invalide: boundary < 3 nodes")
        oldStartCanon = str(self._world.find_node(oldStart))
        newCycleCanon = {str(self._world.find_node(str(n))) for n in newCycle}
        if oldStartCanon not in newCycleCanon:
            raise ValueError("[Chemins] recalcul invalide: startNodeId absent du nouveau contour")

        newSnapshot = [str(n) for n in list(newCycle)]

        # Rejoue les désactivations sur les mêmes ConceptNodeId canoniques.
        oldMaskFalseByCanonicalNode: set[str] = set()
        for i, nodeId in enumerate(oldSnapshot):
            if not bool(oldMask[i]):
                oldMaskFalseByCanonicalNode.add(str(self._world.find_node(str(nodeId))))

        newMask: list[bool] = []
        for nodeId in newSnapshot:
            canonNodeId = str(self._world.find_node(str(nodeId)))
            newMask.append(canonNodeId not in oldMaskFalseByCanonicalNode)

        if sum(1 for v in newMask if bool(v)) < 3:
            raise ValueError("[Chemins] recalcul invalide: moins de 3 nodes actifs")

        boundaryOrientation = self._world.getBoundaryOrientation(gid)
        newPath = self._buildPathNodesOrdered(newSnapshot, newMask, oldOrient, boundaryOrientation)
        if len(newPath) < 3:
            raise ValueError("[Chemins] recalcul invalide: pathNodesOrdered < 3")

        newTriplets = self._buildTriplets(newPath)
        for t in newTriplets:
            t.calculerGeometrie(self._world, gid, oldOrient, baliseRefName)

        oldPath = list(self.pathNodesOrdered)
        oldTriplets = list(self.triplets)
        try:
            self.borderSnapshotNodes = newSnapshot
            self.selectionMask = newMask
            self.orientationUser = oldOrient
            self.pathNodesOrdered = newPath
            self.triplets = newTriplets
            self._assertDefinedInvariants()
        except Exception:
            self.borderSnapshotNodes = oldSnapshot
            self.selectionMask = oldMask
            self.orientationUser = oldOrient
            self.pathNodesOrdered = oldPath
            self.triplets = oldTriplets
            raise

    def supprimerChemin(self) -> None:
        """Supprime le chemin courant et restaure l'etat undefined."""
        if not self.isDefined:
            return
        self._resetUndefined()
        self._assertDefinedInvariants()

    def getTriplets(self) -> list[TopologyCheminTriplet]:
        """Retourne une copie superficielle de la liste des triplets."""
        return list(self.triplets)

    def getTripletAt(self, i: int) -> TopologyCheminTriplet:
        """Retourne le triplet d'index i avec controle de borne strict."""
        idx = int(i)
        if idx < 0 or idx >= len(self.triplets):
            raise IndexError(f"[Chemins] index triplet hors borne: {i}")
        return self.triplets[idx]

    def getTripletSegmentsWorld(self, i: int) -> tuple[tuple[float, float, float, float], tuple[float, float, float, float]]:
        """Retourne les segments monde (O->A) et (O->B) du triplet i."""
        self._requireDefined()
        gid = str(self.groupId)
        t = self.getTripletAt(i)
        xA, yA = self._world.getConceptNodeWorldXY(t.nodeA, gid)
        xO, yO = self._world.getConceptNodeWorldXY(t.nodeO, gid)
        xB, yB = self._world.getConceptNodeWorldXY(t.nodeB, gid)
        return (
            (float(xO), float(yO), float(xA), float(yA)),
            (float(xO), float(yO), float(xB), float(yB)),
        )


class TopologyWorld:
    """Racine métier du modèle topologique (Core, sans Tk).

    Rôle :
    - Détenir l’ensemble des éléments, groupes, nodes et attachments.
    - Implémenter les DSU non destructifs :
      - nodes : canonisation métier (L > B > O, puis plus récent ID max),
      - groupes : canonisation par plus récent (ID max).
    - Gérer les attachments et la canonisation DSU (sans points physiques sur les arêtes).
    - Fournir un export TopoDump XML manuel (par scénario), lisible et diffable.

    Ne fait PAS :
    - Ne dépend jamais du canvas Tk.
    - Ne calcule pas t (fourni par l’UI en référentiel monde, km).

    V1 :
    - Pose les fondations (classes + DSU + export TopoDump).
    - Applique vertex↔vertex, vertex↔edge et edge↔edge (mapping direct|reverse) au niveau DSU + coverages.
    - Dégrouper/undo = suppression d’attaches puis rebuild complet (petits scénarios).
    """
    def __init__(self):
        self.fusion_distance_km: float = 1.0

        # Repère "monde" pour les calculs d'azimut (0°=Nord, sens horaire).
        # - True  : Y augmente vers le bas (repère image/écran)  -> dyNord = -dy
        # - False : Y augmente vers le Nord (repère carto)        -> dyNord =  dy
        self.worldYAxisDown: bool = False

        # --- cache topo conceptuelle (PAR GROUPE) ---
        self._concept_by_gid: dict[str, ConceptGroupCache] = {}

        self.groups: dict[str, TopologyGroup] = {}
        self.elements: dict[str, TopologyElement] = {}
        self.element_to_group: dict[str, str] = {}
        self.attachments: dict[str, TopologyAttachment] = {}

        self._node_parent: dict[str, str] = {}
        self._node_type: dict[str, str] = {}
        self._node_members: dict[str, list[str]] = {}
        self._node_created_order: dict[str, int] = {}

        self._group_parent: dict[str, str] = {}
        self._group_members: dict[str, list[str]] = {}
        self._group_created_order: dict[str, int] = {}

        self._created_counter_nodes = 0
        self._created_counter_groups = 0
        self._created_counter_attachments = 0
        self._topoTxDepth = 0
        self._topoTxTouchedGroups: set[str] = set()
        self._topoTxOrientation = "cw"
        self._isImportingSnapshot = False
        self.topologyChemins = TopologyChemins(self)

        self.balisesWorld: dict[str, tuple[float, float]] = {}

    # ------------------------------------------------------------------
    # Topologie conceptuelle (MVP)
    # - concept node = find_node(canon)
    # - position concept node = moyenne des occurrences monde (conceptuelles)
    # - concept edge = segment entre 2 concept nodes consécutifs sur une arête
    # ------------------------------------------------------------------
    def invalidateConceptGraph(self, group_id: str | None = None) -> None:
        """Invalide le cache conceptuel (graph + geom)."""
        if group_id is None:
            for c in self._concept_by_gid.values():
                c.graphValid = False
                c.geomValid = False
            return
        gid = self.find_group(str(group_id))
        c = self._concept_by_gid.get(gid)
        if c is not None:
            c.graphValid = False
            c.geomValid = False

    def invalidateConceptGeom(self, group_id: str | None = None) -> None:
        """Invalide uniquement la géométrie conceptuelle (world coords)."""
        if group_id is None:
            for c in self._concept_by_gid.values():
                c.geomValid = False
            return
        gid = self.find_group(str(group_id))
        c = self._concept_by_gid.get(gid)
        if c is not None:
            c.geomValid = False

    def beginTopoTransaction(self) -> None:
        self._topoTxDepth += 1
        if self._topoTxDepth == 1:
            self._topoTxTouchedGroups.clear()

    def commitTopoTransaction(self) -> None:
        if self._topoTxDepth <= 0:
            raise ValueError("[TopoTx] commit without begin")
        self._topoTxDepth -= 1
        if self._topoTxDepth > 0:
            return
        for gid in sorted(self._topoTxTouchedGroups):
            self.recomputeConceptAndBoundary(gid)
        self._topoTxTouchedGroups.clear()

    def _markTopoTouched(self, group_id: str) -> None:
        gid = self.find_group(str(group_id))
        if self._topoTxDepth > 0:
            self._topoTxTouchedGroups.add(gid)
            return

    def recomputeConceptAndBoundary(self, group_id: str) -> None:
        gid = self.find_group(str(group_id))
        self.invalidateConceptGraph(gid)
        self.ensureConceptGeom(gid)
        self.computeBoundary(gid)

    def toConceptNodeId(self, node_id: str) -> str:
        return self.find_node(str(node_id))

    def _concept_cache(self, group_id: str) -> ConceptGroupCache:
        gid = self.find_group(str(group_id))
        c = self._concept_by_gid.get(gid)
        if c is None:
            c = ConceptGroupCache()
            self._concept_by_gid[gid] = c
        return c

    def _requireConceptCache(self, group_id: str) -> ConceptGroupCache:
        gid = self.find_group(str(group_id))
        c = self._concept_cache(gid)
        if not c.graphValid:
            raise ValueError(f"[ConceptCache] not computed (gid={gid})")
        return c

    def ensureConceptGraph(self, group_id: str) -> ConceptGroupCache:
        gid = self.find_group(str(group_id))
        c = self._concept_cache(gid)
        if not c.graphValid:
            self._buildConceptCacheForGroup(gid)
        return c

    def ensureConceptGeom(self, group_id: str) -> ConceptGroupCache:
        gid = self.find_group(str(group_id))
        c = self.ensureConceptGraph(gid)
        if not c.geomValid:
            self._refreshConceptGeom(gid)
        return c

    @staticmethod
    def _element_id_from_atomic_node_id(node_id: str) -> str | None:
        s = str(node_id)
        if ":N" not in s:
            return None
        return s.rsplit(":N", 1)[0]

    def _infer_group_for_node(self, node_id: str) -> str | None:
        cn = self.find_node(str(node_id))
        mem = self.node_members(cn)
        if not mem:
            return None
        eid = self._element_id_from_atomic_node_id(mem[0])
        if not eid:
            return None
        return self.element_to_group.get(str(eid))

    def _iter_group_elements(self, group_id: str) -> list["TopologyElement"]:
        gid = self.find_group(str(group_id))
        g = self.groups.get(gid)
        if g is None:
            return []
        out: list[TopologyElement] = []
        for eid in g.element_ids:
            el = self.elements.get(str(eid))
            if el is not None:
                out.append(el)
        return out

    def assertNoPhysicalSplitPoints(self) -> None:
        for el in self.elements.values():
            for edge in el.edges:
                if hasattr(edge, "points_on_edge"):
                    raise RuntimeError(
                        f"[A1] Physical edge split points are forbidden: elementId={el.element_id} edgeId={edge.edge_id()}"
                    )

    def buildDerivedSplitPointsForPhysEdge(self, element_id: str, edge_index: int,
                                           eps_t: float = 1e-9) -> list[dict]:
        edge = self.get_edge(element_id, edge_index)
        edge_start = str(edge.v_start.node_id)
        edge_end = str(edge.v_end.node_id)
        if edge_start == edge_end:
            raise ValueError(f"[DP][{edge.edge_id()}] edge endpoints are identical")
        start_cn = self.find_node(str(edge.v_start.node_id))
        end_cn = self.find_node(str(edge.v_end.node_id))
        dp: list[dict] = [
            {"t": 0.0, "nodeCanon": start_cn, "source": "endpoint"},
            {"t": 1.0, "nodeCanon": end_cn, "source": "endpoint"},
        ]

        for att in self.attachments.values():
            if str(att.kind) != "vertex-edge":
                continue
            if att.feature_a.feature_type == TopologyFeatureType.VERTEX and att.feature_b.feature_type == TopologyFeatureType.EDGE:
                vRef, eRef = att.feature_a, att.feature_b
            elif att.feature_a.feature_type == TopologyFeatureType.EDGE and att.feature_b.feature_type == TopologyFeatureType.VERTEX:
                vRef, eRef = att.feature_b, att.feature_a
            else:
                continue
            if str(eRef.element_id) != str(element_id) or int(eRef.index) != int(edge_index):
                continue
            t_val = att.params.get("t", None)
            if t_val is None:
                raise ValueError(f"[DP][{edge.edge_id()}] vertex-edge missing t id={att.attachment_id}")
            edge_from = att.params.get("edgeFrom", None)
            if edge_from is None or str(edge_from).strip() == "":
                raise ValueError(f"[DP][{edge.edge_id()}] vertex-edge missing edgeFrom id={att.attachment_id}")
            edge_from = str(edge_from)
            if edge_from == edge_start:
                t_phys = float(t_val)
            elif edge_from == edge_end:
                t_phys = 1.0 - float(t_val)
            else:
                raise ValueError(
                    f"[DP][{edge.edge_id()}] edgeFrom not on edge (edgeFrom={edge_from}) id={att.attachment_id}"
                )
            if t_phys < 0.0 - eps_t or t_phys > 1.0 + eps_t:
                raise ValueError(
                    f"[DP][{edge.edge_id()}] vertex-edge t out of [0,1] (t={t_phys:.6f}) id={att.attachment_id}"
                )
            if abs(t_phys) <= eps_t or abs(t_phys - 1.0) <= eps_t:
                continue
            v = self.get_vertex(vRef.element_id, vRef.index)
            node_canon = self.find_node(v.node_id)
            if node_canon == start_cn or node_canon == end_cn:
                continue
            dp.append({
                "t": float(t_phys),
                "nodeCanon": node_canon,
                "source": "vertex-edge",
                "attachmentId": att.attachment_id,
            })

        dp.sort(key=lambda p: float(p.get("t", 0.0)))
        merged: list[dict] = []
        for sp in dp:
            if not merged:
                merged.append(sp)
                continue
            t_prev = float(merged[-1]["t"])
            t_cur = float(sp["t"])
            if abs(t_cur - t_prev) <= eps_t:
                if str(sp["nodeCanon"]) != str(merged[-1]["nodeCanon"]):
                    raise ValueError(
                        f"[DP][{edge.edge_id()}] conflicting splitpoints at t={t_cur} "
                        f"nodeA={merged[-1]['nodeCanon']} nodeB={sp['nodeCanon']}"
                    )
                continue
            merged.append(sp)

        if not merged:
            raise ValueError(f"[DP][{edge.edge_id()}] DP missing endpoints")
        if abs(float(merged[0]["t"]) - 0.0) > eps_t or abs(float(merged[-1]["t"]) - 1.0) > eps_t:
            raise ValueError(f"[DP][{edge.edge_id()}] DP missing endpoints")
        return merged

    def buildConceptOccurrencesForPhysEdge(self, element_id: str, edge_index: int,
                                           eps_t: float = 1e-9) -> list[dict]:
        edge = self.get_edge(element_id, edge_index)
        dp = self.buildDerivedSplitPointsForPhysEdge(element_id, edge_index, eps_t=eps_t)
        from_node_id = str(edge.v_start.node_id)
        to_node_id = str(edge.v_end.node_id)
        if from_node_id == to_node_id:
            raise ValueError(f"[C1][{edge.edge_id()}] from/to are identical")
        occs: list[dict] = []
        for i in range(len(dp) - 1):
            t0 = float(dp[i]["t"])
            t1 = float(dp[i + 1]["t"])
            if t1 <= t0 + eps_t:
                raise ValueError(f"[C1][{edge.edge_id()}] null segment in DP t0={t0} t1={t1}")
            occs.append({
                "elementId": str(edge.element_id),
                "edgeId": str(edge.edge_id()),
                "fromNodeId": from_node_id,
                "toNodeId": to_node_id,
                "t0": t0,
                "t1": t1,
            })
        return occs

    def _buildConceptCacheForGroup(self, group_id: str) -> None:
        gid = self.find_group(str(group_id))
        c = self._concept_cache(gid)

        nodes: dict[str, ConceptNodeInfo] = {}
        edges: dict[tuple[str, str], ConceptEdgeInfo] = {}
        node_occ: dict[str, list[tuple[str, tuple[float, float]]]] = {}

        eps_t = float(getattr(self, "concept_split_epsilon_t", 1e-6))
        for el in self._iter_group_elements(gid):
            for edge in getattr(el, "edges", []):
                dp = self.buildDerivedSplitPointsForPhysEdge(el.element_id, edge.edge_index, eps_t=eps_t)
                from_node_id = str(edge.v_start.node_id)
                to_node_id = str(edge.v_end.node_id)

                # nodes : occurrences locales (pour refresh geom)
                for sp in dp:
                    cn = str(sp["nodeCanon"])
                    info = nodes.get(cn)
                    if info is None:
                        info = ConceptNodeInfo(concept_id=cn, members=set(self.node_members(cn)))
                        nodes[cn] = info
                    p_local = self._localPointOnEdge(el, edge, float(sp["t"]))
                    node_occ.setdefault(cn, []).append(
                        (str(el.element_id), (float(p_local[0]), float(p_local[1])))
                    )

                # edges : segments consecutifs canonises
                for i in range(len(dp) - 1):
                    n0 = str(dp[i]["nodeCanon"])
                    n1 = str(dp[i + 1]["nodeCanon"])
                    t0 = float(dp[i]["t"])
                    t1 = float(dp[i + 1]["t"])
                    if n0 == n1:
                        continue
                    k = (n0, n1) if n0 < n1 else (n1, n0)
                    ce = edges.get(k)
                    if ce is None:
                        ce = ConceptEdgeInfo(a=k[0], b=k[1], occurrences=[])
                        edges[k] = ce
                    ce.occurrences.append({
                        "elementId": str(el.element_id),
                        "edgeId": str(edge.edge_id()),
                        "fromNodeId": from_node_id,
                        "toNodeId": to_node_id,
                        "t0": float(t0),
                        "t1": float(t1),
                    })
        c.nodes = nodes
        c.edges = edges
        c.nodeOccurrencesByCid = node_occ
        c.graphValid = True
        c.geomValid = False
        self._refreshConceptGeom(gid)

    def buildConceptCacheIfNeeded(self, group_id: str) -> None:
        gid = self.find_group(str(group_id))
        c = self._concept_cache(gid)
        if not c.graphValid:
            self._buildConceptCacheForGroup(gid)

    def _refreshConceptGeom(self, group_id: str) -> None:
        gid = self.find_group(str(group_id))
        c = self._concept_cache(gid)
        for info in c.nodes.values():
            info.sumx = 0.0
            info.sumy = 0.0
            info.count = 0
        for cid, occs in c.nodeOccurrencesByCid.items():
            info = c.nodes.get(cid)
            if info is None:
                continue
            for element_id, local_xy in occs:
                el = self.elements.get(str(element_id))
                if el is None:
                    raise ValueError(f"[ConceptGeom] unknown elementId={element_id} for cid={cid}")
                p_world = el.localToWorld(np.array(local_xy, dtype=float))
                info.add_occ(p_world[0], p_world[1])
        c.geomValid = True

    def getConceptNodeWorldXY(self, node_id: str, group_id: str | None = None) -> tuple[float, float]:
        """Retourne la position monde d'un concept node (x, y).
        Garantit un couple de floats pour un node résolu via le groupe.
        Ne modifie pas la topologie ni les attachments.
        Ne définit pas le groupe si l'inférence échoue.
        Peut retourner (0,0) si l'information n'est pas disponible.
        """
        gid = group_id if group_id is not None else self._infer_group_for_node(node_id)
        if gid is None:
            return (0.0, 0.0)
        cn = self.find_node(str(node_id))
        c = self.ensureConceptGeom(gid)
        info = c.nodes.get(cn)
        if info is None:
            return (0.0, 0.0)
        return info.world_xy()

    def getConceptRays(self, node_id: str, group_id: str | None = None) -> list[dict]:
        """
        Retourne les rayons (arcs) conceptuels incident au nœud conceptuel canonique.

        Entrées :
        - node_id (str) : identifiant d’un nœud (atomique ou déjà canonique). La canonisation DSU
        est appliquée en interne.
        - group_id (str | None) : si fourni, force le groupe ; sinon le groupe est inféré depuis node_id.

        Sortie :
        - list[dict] : liste de rayons, chacun au format :
            {
            "otherNodeId": str,  # id canonique du nœud voisin
            "azDeg": float       # azimut en degrés (0° = Nord, sens horaire), normalisé [0..360)
            }

        Garanties :
        - les voisins sont limités au groupe résolu (fourni ou inféré),
        - la liste est triée de manière stable (azDeg puis otherNodeId),
        - n’écrit aucun texte d’affichage et ne modifie ni topologie ni caches.
        """
        gid = group_id if group_id is not None else self._infer_group_for_node(node_id)
        if gid is None:
            return []
        cn = self.find_node(str(node_id))
        p0 = self.getConceptNodeWorldXY(cn, gid)
        c = self.ensureConceptGeom(gid)
        out: list[dict] = []
        for (a, b), ce in c.edges.items():
            if cn != a and cn != b:
                continue
            other = b if cn == a else a
            p1 = self.getConceptNodeWorldXY(other, gid)
            dx = float(p1[0] - p0[0])
            dy = float(p1[1] - p0[1])
            out.append({
                "otherNodeId": str(other),
                "azDeg": self.azimutDegFromDxDy(dx, dy),
            })
        out.sort(key=lambda r: (float(r.get("azDeg", 0.0)), str(r.get("otherNodeId", ""))))
        return out

    def computeBoundary(self, group_id: str) -> None:
        """
        Overview
        --------
        Calcule et met en cache le boundary (cycle du contour) d’un groupe topologique.
        Cette méthode est le point d’entrée "métier" : elle déclenche le calcul dans
        `TopologyBoundaries`, puis recopie les résultats dans le cache conceptuel du groupe.

        Entrées
        -------
        group_id : str
            Identifiant de groupe (canonisé en interne).

        Sorties
        -------
        None
            Met à jour le cache conceptuel du groupe:
            - boundaryCycle
            - boundaryEdges
            - boundaryIndex
            - boundaryOrientation

        Traitement
        ----------
        1) Canonise l’identifiant de groupe.
        2) Assure que la géométrie conceptuelle (nodes/edges + barycentres) est disponible.
        3) Lance `c.boundaries.compute(...)` avec l’orientation de transaction topo courante.
        4) Recopie dans le cache du groupe:
           - `boundaryCycle`, `boundaryEdges`, `boundaryIndex`, `boundaryOrientation`.
        """
        gid = self.find_group(str(group_id))
        c = self.ensureConceptGeom(gid)
        c.boundaries.compute(self, gid, orientation=self._topoTxOrientation)
        c.boundaryCycle = c.boundaries.cycle
        c.boundaryEdges = c.boundaries.edges
        c.boundaryIndex = c.boundaries.index
        c.boundaryOrientation = c.boundaries.orientation

    def getBoundaryCycle(self, group_id: str, startNodeId: str) -> list[str]:
        gid = self.find_group(str(group_id))
        c = self._concept_cache(gid)
        if not c.boundaryCycle:
            raise ValueError("[Boundary] not computed")
        start = str(startNodeId)
        if start not in c.boundaryIndex:
            raise ValueError("[Boundary] not computed")
        idx = c.boundaryIndex[start]
        return c.boundaryCycle[idx:] + c.boundaryCycle[:idx]

    def getBoundaryOrientation(self, group_id: str) -> str:
        gid = self.find_group(str(group_id))
        c = self._concept_cache(gid)
        if not c.boundaryOrientation:
            raise ValueError("[Boundary] not computed")
        return str(c.boundaryOrientation)

    def getBoundaryNeighbors(self, group_id: str, nodeId: str) -> tuple[str, str]:
        gid = self.find_group(str(group_id))
        c = self._concept_cache(gid)
        if not c.boundaryCycle:
            raise ValueError("[Boundary] not computed")
        node = str(nodeId)
        if node not in c.boundaryIndex:
            raise ValueError("[Boundary] not computed")
        idx = c.boundaryIndex[node]
        prev_node = c.boundaryCycle[idx - 1]
        next_node = c.boundaryCycle[(idx + 1) % len(c.boundaryCycle)]
        return (prev_node, next_node)

    def isBoundaryEdge(self, group_id: str, a: str, b: str) -> bool:
        gid = self.find_group(str(group_id))
        c = self._concept_cache(gid)
        if not c.boundaryEdges:
            raise ValueError("[Boundary] not computed")
        return (str(a), str(b)) in c.boundaryEdges or (str(b), str(a)) in c.boundaryEdges

    def getBoundarySegments(self, group_id: str) -> list[BoundarySegment]:
        gid = self.find_group(str(group_id))
        c = self._concept_cache(gid)
        if not c.boundaryCycle or not c.boundaryIndex:
            raise ValueError("[Boundary] not computed")
        if not c.edges:
            raise ValueError("[Boundary] concept edges missing")

        cycle = c.boundaryCycle
        if len(cycle) < 2:
            return []

        segments: list[BoundarySegment] = []
        for i in range(len(cycle)):
            ca = str(cycle[i])
            cb = str(cycle[(i + 1) % len(cycle)])
            k = (ca, cb) if ca < cb else (cb, ca)
            ce = c.edges.get(k)
            if ce is None:
                raise ValueError(f"[Boundary] missing concept edge for {ca}-{cb}")
            if not ce.occurrences:
                raise ValueError(f"[Boundary] missing occurrences for {ca}-{cb}")
            occ = ce.occurrences[0]

            if "elementId" not in occ:
                raise ValueError(f"[Boundary] missing elementId for {ca}-{cb}")
            if "edgeId" not in occ:
                raise ValueError(f"[Boundary] missing edgeId for {ca}-{cb}")
            if "t0" not in occ or "t1" not in occ:
                raise ValueError(f"[Boundary] missing t0/t1 for {ca}-{cb}")
            if "fromNodeId" not in occ or "toNodeId" not in occ:
                raise ValueError(f"[Boundary] missing fromNodeId/toNodeId for {ca}-{cb}")

            element_id = str(occ["elementId"])
            edge_id = str(occ["edgeId"])
            t0 = float(occ["t0"])
            t1 = float(occ["t1"])
            from_node_id = str(occ["fromNodeId"])
            to_node_id = str(occ["toNodeId"])
            if from_node_id == to_node_id:
                raise ValueError(f"[Boundary] from/to identical for {edge_id}")

            if ":E" not in edge_id:
                raise ValueError(f"[Boundary] invalid edgeId format: {edge_id}")
            edge_head, edge_tail = edge_id.rsplit(":E", 1)
            if not edge_head or not edge_tail:
                raise ValueError(f"[Boundary] invalid edgeId format: {edge_id}")
            try:
                edge_index = int(edge_tail)
            except ValueError as exc:
                raise ValueError(f"[Boundary] invalid edgeId index: {edge_id}") from exc

            if edge_head != element_id:
                raise ValueError(f"[Boundary] edgeId does not match elementId: {edge_id}")

            edge = self.get_edge(element_id, edge_index)
            if str(edge.v_start.node_id) != from_node_id or str(edge.v_end.node_id) != to_node_id:
                raise ValueError(f"[Boundary] occurrence from/to mismatch for {edge_id}")

            eps_t = float(getattr(self, "concept_split_epsilon_t", 1e-6))
            dp = self.buildDerivedSplitPointsForPhysEdge(element_id, edge_index, eps_t=eps_t)

            def _concept_at_t(t_val: float) -> str:
                for sp in dp:
                    if abs(float(sp["t"]) - t_val) <= eps_t:
                        return str(sp["nodeCanon"])
                raise ValueError(f"[Boundary] missing split point at t={t_val} for {edge_id}")

            n0 = _concept_at_t(t0)
            n1 = _concept_at_t(t1)

            if n0 == ca and n1 == cb:
                pass
            elif n0 == cb and n1 == ca:
                t0, t1 = (1.0 - t1), (1.0 - t0)
                from_node_id, to_node_id = to_node_id, from_node_id
            else:
                raise ValueError(f"[Boundary] occurrence orientation mismatch for {ca}-{cb}: {n0}-{n1}")

            if t0 < 0.0 or t0 > 1.0 or t1 < 0.0 or t1 > 1.0 or t0 >= t1:
                raise ValueError(f"[Boundary] invalid t0/t1 for {edge_id}: t0={t0} t1={t1}")

            segments.append(BoundarySegment(
                conceptA=ca,
                conceptB=cb,
                elementId=element_id,
                edgeIndex=edge_index,
                fromNodeId=from_node_id,
                toNodeId=to_node_id,
                t0=t0,
                t1=t1,
            ))

        return segments

    def _build_ring_from_boundary_cycle(
        self,
        group_id: str,
        eps_world: float = 1e-9,
    ) -> tuple[list[np.ndarray], list[str], dict[str, int]] | None:
        """
        Overview
        --------
        Reconstruit un "ring" géométrique (points monde) à partir du boundary conceptuel
        d’un groupe. L’objectif est d’obtenir une représentation polygonale simple du contour
        (un point physique par nœud conceptuel du cycle) utilisée ensuite par la simulation
        d’anti-chevauchement.

        Entrées
        -------
        group_id : str
            Identifiant de groupe (canonisé en interne).
        eps_world : float
            Seuil (monde) pour rejeter des doublons consécutifs (points trop proches).

        Sorties
        -------
        tuple | None
            - None si le ring ne peut pas être construit (cycle trop court, occurrences manquantes,
              doublons consécutifs, collisions d’ids).
            - Sinon un triplet:
              (ring_pts, ring_nodes, index)
              - ring_pts  : list[np.ndarray] (points monde, ring non fermé)
              - ring_nodes: list[str]        (ids conceptuels dans le même ordre)
              - index     : dict[str,int]    (nodeId → position)

        Traitement
        ----------
        1) Vérifie que le boundary du groupe est déjà calculé (cycle présent dans le cache).
        2) Assure que les occurrences physiques (nodeOccurrencesByCid) sont disponibles.
        3) Pour chaque nœud conceptuel du cycle:
           - prend une occurrence physique (element_id, local_xy),
           - convertit en coordonnées monde via la pose de l’élément,
           - alimente `ring_pts` (np.ndarray 2D) et `ring_nodes` (ids conceptuels).
        4) Effectue une validation minimale:
           - taille >= 3,
           - pas de doublons consécutifs (distance <= eps_world) le long du cycle.
        5) Construit `index` (nodeId → position) et rejette si collisions.
        """
        gid = self.find_group(str(group_id))
        c = self._concept_cache(gid)
        if not c.boundaryCycle:
            raise ValueError("[Boundary] not computed")

        cycle = list(c.boundaryCycle)
        if len(cycle) < 3:
            return None

        # Assure que les occurrences physiques sont prêtes
        c = self.ensureConceptGeom(gid)

        ring_pts: list[np.ndarray] = []
        ring_nodes: list[str] = []
        for cn in cycle:
            occs = c.nodeOccurrencesByCid.get(str(cn))
            if not occs:
                return None
            element_id, local_xy = occs[0]
            el = self.elements.get(str(element_id))
            if el is None:
                return None
            p_world = el.localToWorld(np.array(local_xy, dtype=float))
            ring_pts.append(np.array([float(p_world[0]), float(p_world[1])], dtype=float))
            ring_nodes.append(str(cn))

        if len(ring_pts) < 3:
            return None

        # Validation minimale : pas de doublons consécutifs
        for i in range(len(ring_pts) - 1):
            dx = float(ring_pts[i + 1][0] - ring_pts[i][0])
            dy = float(ring_pts[i + 1][1] - ring_pts[i][1])
            if (dx * dx + dy * dy) <= float(eps_world) * float(eps_world):
                return None

        index = {str(n): i for i, n in enumerate(ring_nodes)}
        if len(index) != len(ring_nodes):
            return None
        return (ring_pts, ring_nodes, index)

    def _arcForward(self, pts: list[np.ndarray], i: int, j: int) -> list[np.ndarray]:
        """Arc avant sur un ring non fermé: i -> j en sens direct."""
        if i == j:
            return [pts[i]]
        if i < j:
            return pts[i:j + 1]
        return pts[i:] + pts[:j + 1]

    def _arcReverse(self, pts: list[np.ndarray], i: int, j: int) -> list[np.ndarray]:
        """Arc arrière sur un ring non fermé: i -> j en sens inverse."""
        return list(reversed(self._arcForward(pts, j, i)))

    def _concatArcs(self, a: list[np.ndarray], b: list[np.ndarray]) -> list[np.ndarray]:
        """Concatène deux arcs. b est inversé, puis on évite le doublon au point de jonction."""
        if not a or not b:
            return []
        b_rev = list(reversed(b))
        return list(a[:-1]) + b_rev[:-1]

    def _resolveJunctionsMobDest(
        self,
        attachments: list[TopologyAttachment],
        nodesMob: list[str],
        ptsMob: list[np.ndarray],
        indexMob: dict[str, int],
        nodesDest: list[str],
        ptsDest: list[np.ndarray],
        indexDest: dict[str, int],
    ) -> tuple[dict, dict] | None:
        """
        Overview
        --------
        Interprète une liste d’attachments (edge-edge ou vertex-edge) et en déduit deux jonctions
        topologiques J0 et J1 entre un ring "mobile" et un ring "destination". Ces jonctions
        sont la base de toute la simulation: elles définissent les deux points (ou point+point sur arête)
        qui permettent de calculer une pose rigide et de découper les deux rings en arcs.

        Entrées
        -------
        attachments : list[TopologyAttachment]
            Attachments candidats (edge-edge ou vertex-edge selon le cas).
        nodesMob, ptsMob, indexMob : (list[str], list[np.ndarray], dict[str,int])
            Ring mobile (concept node ids + points monde + index).
        nodesDest, ptsDest, indexDest : (list[str], list[np.ndarray], dict[str,int])
            Ring destination (concept node ids + points monde + index).

        Sorties
        -------
        tuple[dict, dict] | None
            - None si les attachments ne correspondent pas à un cas supporté ou si des
              informations nécessaires manquent dans les rings.
            - Sinon (J0, J1) : deux dictionnaires décrivant les jonctions mob/dest.
              Chaque jonction expose typiquement:
              - mobType / destType : "node" ou "edge"
              - mobId / destId     : id conceptuel (node) ou référence (edge)
              - mobPt / destPt     : np.ndarray 2D (si type "node", lu depuis le ring)
              Des champs additionnels (t, dir, etc.) peuvent être présents selon le cas vertex-edge.

        Traitement
        ----------
        1) Normalise la liste d’attachments, rejette les cas vides ou non supportés.
        2) Cas A : edge-edge (un seul attachment)
           - Identifie quelles extrémités d’arêtes appartiennent au ring mobile et au ring destination.
           - Applique le mapping (direct/reverse) pour obtenir deux jonctions "node-node":
             J0 = (mob0 ↔ dest0) et J1 = (mob1 ↔ dest1).
        3) Cas B : vertex-edge (deux attachments: un vertex-vertex + un vertex-edge)
           - Le vertex-vertex définit J0 (node-node) et fixe la référence d’orientation.
           - Le vertex-edge définit un point sur une arête (t + edgeFrom):
             - calcule t côté edge (sens) et le point monde correspondant,
             - récupère une direction unitaire de l’arête depuis le ring (référencée côté J0),
             - décide si la partie "edge" appartient au ring mobile ou au ring destination,
               et construit J1 comme (edge ↔ node) du bon côté.
        4) Complète les champs `mobPt` / `destPt` pour les jonctions de type "node"
           en lisant les coordonnées dans les rings (indexMob / indexDest).
        5) Retourne (J0, J1) prêts pour le calcul de pose et la découpe en arcs.
        """

        atts = list(attachments or [])
        if not atts:
            return None

        def _edge_t_from_params(edge: TopologyEdge, t_val: float, edge_from: str) -> float | None:
            if edge_from == str(edge.v_start.node_id):
                return float(t_val)
            if edge_from == str(edge.v_end.node_id):
                return 1.0 - float(t_val)
            return None

        def _node_pt(nid: str, pts: list[np.ndarray], idx: dict[str, int]) -> np.ndarray | None:
            if nid not in idx:
                return None
            return pts[idx[nid]]

        def _edge_dir_from_ring(edge: TopologyEdge, ptsRing: list, indexRing: dict[str, int], refNodeId: str):
            """
            Return a unit direction vector for edgeId, oriented so it starts from refNodeId.
            refNodeId MUST be one of the edge endpoints in the ring (concept node id).
            """
            a = str(self.find_node(edge.v_start.node_id))
            b = str(self.find_node(edge.v_end.node_id))

            if a not in indexRing or b not in indexRing:
                return None

            pa = ptsRing[indexRing[a]]
            pb = ptsRing[indexRing[b]]

            v = pb - pa
            n = float((v[0] ** 2 + v[1] ** 2) ** 0.5)
            if n <= 1e-12:
                return None
            dir_ab = v / n

            # orienter le vecteur au départ du refNode
            if refNodeId == a:
                return dir_ab
            if refNodeId == b:
                return -dir_ab

            # refNode n'est pas un endpoint => incohérent
            return None

        # Cas A: edge-edge (1 attachment)
        if len(atts) == 1 and str(atts[0].kind) == "edge-edge":
            att = atts[0]
            if att.feature_a is None or att.feature_b is None:
                return None
            if att.feature_a.feature_type != TopologyFeatureType.EDGE or att.feature_b.feature_type != TopologyFeatureType.EDGE:
                return None

            eA = self.get_edge(att.feature_a.element_id, att.feature_a.index)
            eB = self.get_edge(att.feature_b.element_id, att.feature_b.index)
            a0 = self.find_node(str(eA.v_start.node_id))
            a1 = self.find_node(str(eA.v_end.node_id))
            b0 = self.find_node(str(eB.v_start.node_id))
            b1 = self.find_node(str(eB.v_end.node_id))

            mapping = str(att.params.get("mapping", "") or "").strip().lower()
            if mapping not in ("direct", "reverse"):
                return None

            if a0 in indexMob and a1 in indexMob and b0 in indexDest and b1 in indexDest:
                mob0, mob1 = a0, a1
                dest0, dest1 = b0, b1
            elif b0 in indexMob and b1 in indexMob and a0 in indexDest and a1 in indexDest:
                mob0, mob1 = b0, b1
                dest0, dest1 = a0, a1
            else:
                return None

            if mapping == "reverse":
                dest0, dest1 = dest1, dest0

            J0 = {"mobType": "node", "mobId": str(mob0), "destType": "node", "destId": str(dest0)}
            J1 = {"mobType": "node", "mobId": str(mob1), "destType": "node", "destId": str(dest1)}

        # Cas B: vertex-edge (2 attachments)
        elif len(atts) == 2:
            att_vv = None
            att_ve = None
            for att in atts:
                if str(att.kind) == "vertex-vertex":
                    att_vv = att
                elif str(att.kind) == "vertex-edge":
                    att_ve = att
            if att_vv is None or att_ve is None:
                return None

            # vertex-vertex => J0
            if att_vv.feature_a is None or att_vv.feature_b is None:
                return None
            vA = self.get_vertex(att_vv.feature_a.element_id, att_vv.feature_a.index)
            vB = self.get_vertex(att_vv.feature_b.element_id, att_vv.feature_b.index)
            nA = self.find_node(str(vA.node_id))
            nB = self.find_node(str(vB.node_id))
            if nA in indexMob and nB in indexDest:
                J0 = {"mobType": "node", "mobId": str(nA), "destType": "node", "destId": str(nB)}
            elif nB in indexMob and nA in indexDest:
                J0 = {"mobType": "node", "mobId": str(nB), "destType": "node", "destId": str(nA)}
            else:
                return None

            # vertex-edge => J1 (edge side)
            if att_ve.feature_a is None or att_ve.feature_b is None:
                return None
            if att_ve.feature_a.feature_type == TopologyFeatureType.VERTEX and att_ve.feature_b.feature_type == TopologyFeatureType.EDGE:
                v_ref = att_ve.feature_a
                e_ref = att_ve.feature_b
            elif att_ve.feature_a.feature_type == TopologyFeatureType.EDGE and att_ve.feature_b.feature_type == TopologyFeatureType.VERTEX:
                v_ref = att_ve.feature_b
                e_ref = att_ve.feature_a
            else:
                return None

            edge = self.get_edge(e_ref.element_id, e_ref.index)
            t_raw = att_ve.params.get("t", None)
            edge_from = att_ve.params.get("edgeFrom", None)
            if t_raw is None or edge_from is None or str(edge_from).strip() == "":
                return None
            t_val = float(t_raw)

            t_edge = _edge_t_from_params(edge, t_val, str(edge_from))
            if t_edge is None:
                return None

            el = self.elements.get(str(e_ref.element_id))
            if el is None:
                return None
            p_local = self._localPointOnEdge(el, edge, float(t_edge))
            p_world = el.localToWorld(p_local)

            ref_node = self.find_node(str(edge.v_start.node_id)) if t_edge <= 0.5 else self.find_node(str(edge.v_end.node_id))
            # on trouve les vecteurs edge_mob et edge_dest avec une référence coté J0
            edge_dir_mob = _edge_dir_from_ring(edge=edge, ptsRing=ptsMob, indexRing=indexMob, refNodeId=str(J0["mobId"]))
            edge_dir_dest = _edge_dir_from_ring(edge=edge, ptsRing=ptsDest, indexRing=indexDest, refNodeId=str(J0["destId"]))

            v_node = self.find_node(str(self.get_vertex(v_ref.element_id, v_ref.index).node_id))
            if edge_dir_mob is not None and ref_node in indexMob:
                J1 = {"mobType": "edge", "mobId": edge.edge_id(), "mobPt": p_world, "mobDir": edge_dir_mob,
                      "destType": "node", "destId": str(v_node)}
            elif edge_dir_dest is not None and ref_node in indexDest:
                J1 = {"mobType": "node", "mobId": str(v_node),
                      "destType": "edge", "destId": edge.edge_id(), "destPt": p_world, "destDir": edge_dir_dest}
            else:
                return None
        else:
            return None

        # Remplir points depuis rings
        for j in (J0, J1):
            if j.get("mobType") == "node":
                p = _node_pt(j.get("mobId", ""), ptsMob, indexMob)
                if p is None:
                    return None
                j["mobPt"] = p
            if j.get("destType") == "node":
                p = _node_pt(j.get("destId", ""), ptsDest, indexDest)
                if p is None:
                    return None
                j["destPt"] = p
        return (J0, J1)

    def _computeRigidTransform(self, J0: dict, J1: dict) -> tuple[np.ndarray, np.ndarray] | None:
        """Calcule (R,T) tel que R*mob + T = dest entre J0 et J1."""
        p0m = J0.get("mobPt")
        p1m = J1.get("mobPt")
        p0d = J0.get("destPt")
        p1d = J1.get("destPt")
        if p0m is None or p1m is None or p0d is None or p1d is None:
            return None
        vm = np.array([float(p1m[0] - p0m[0]), float(p1m[1] - p0m[1])], dtype=float)
        vd = np.array([float(p1d[0] - p0d[0]), float(p1d[1] - p0d[1])], dtype=float)
        nm = float(np.hypot(vm[0], vm[1]))
        nd = float(np.hypot(vd[0], vd[1]))
        if nm <= 1e-12 or nd <= 1e-12:
            return None
        ang_m = math.atan2(float(vm[1]), float(vm[0]))
        ang_d = math.atan2(float(vd[1]), float(vd[0]))
        dtheta = ang_d - ang_m
        c = math.cos(dtheta)
        s = math.sin(dtheta)
        R = np.array([[c, -s], [s, c]], dtype=float)
        T = np.array([float(p0d[0]), float(p0d[1])], dtype=float) - (R @ np.array([float(p0m[0]), float(p0m[1])], dtype=float))
        return (R, T)

    def _injectSplitByEdgeDir(
        self,
        nodes: list[str],
        pts: list[np.ndarray],
        index: dict[str, int],
        refNode: str,
        splitPt: np.ndarray,
        edgeDirRot: np.ndarray,
    ) -> tuple[list[str], list[np.ndarray], dict[str, int], str] | None:
        """Injecte un pseudo-nœud sur un ring selon la direction d'arête."""
        if refNode not in index:
            return None
        n = len(nodes)
        if n < 3:
            return None
        i = int(index[refNode])
        i_prev = (i - 1) % n
        i_next = (i + 1) % n
        p_ref = pts[i]
        p_prev = pts[i_prev]
        p_next = pts[i_next]

        v_prev = np.array([float(p_prev[0] - p_ref[0]), float(p_prev[1] - p_ref[1])], dtype=float)
        v_next = np.array([float(p_next[0] - p_ref[0]), float(p_next[1] - p_ref[1])], dtype=float)
        n_prev = float(np.hypot(v_prev[0], v_prev[1]))
        n_next = float(np.hypot(v_next[0], v_next[1]))
        if n_prev <= 1e-12 or n_next <= 1e-12:
            return None
        v_prev /= n_prev
        v_next /= n_next

        s_prev = float(np.dot(edgeDirRot, v_prev))
        s_next = float(np.dot(edgeDirRot, v_next))
        insert_after = bool(s_next >= s_prev)

        out_nodes = list(nodes)
        out_pts = list(pts)
        pseudo = "__EDGE__1"
        insert_at = i + 1 if insert_after else i
        out_nodes.insert(insert_at, pseudo)
        out_pts.insert(insert_at, np.array([float(splitPt[0]), float(splitPt[1])], dtype=float))
        out_index = {str(nid): k for k, nid in enumerate(out_nodes)}
        if len(out_index) != len(out_nodes):
            return None
        return (out_nodes, out_pts, out_index, pseudo)

    def _isValidPolygon(self, ringPts: list[np.ndarray]) -> bool:
        """Valide un contour simple: LineString simple + Polygon valide + aire > EPS_AREA."""
        if not ringPts or len(ringPts) < 3:
            return False
        ring = [(float(p[0]), float(p[1])) for p in ringPts]
        ls = _ShLine(ring)
        if not ls.is_simple:
            return False
        poly = _ShPoly(ring)
        if getattr(poly, "is_empty", False) or not poly.is_valid:
            return False
        eps_area = float(getattr(self, "overlap_eps_area", 1e-12))
        if abs(float(poly.area)) <= eps_area:
            return False
        return True

    def simulateOverlapTopologique(self,
                                   groupDest : TopologyGroup, groupMob: TopologyGroup,
                                   attachments : list[TopologyAttachment],
                                   debug : bool = False) -> bool:
        """
        Overview
        --------
        Simule une tentative d’assemblage topologique entre deux groupes (mobile et destination)
        à partir d’attachments candidats, et détecte un chevauchement via la qualité du contour
        résultant. L’idée est de reconstruire un ring "résultat" en recollant deux arcs de boundary,
        puis de valider que ce contour est un polygone simple (pas d’auto-intersection, aire non nulle).

        Entrées
        -------
        groupDest : TopologyGroup | str
            Groupe destination (ou identifiant) sur lequel on tente de coller.
        groupMob : TopologyGroup | str
            Groupe mobile (ou identifiant) que l’on tente de poser.
        attachments : list[TopologyAttachment]
            Attachments candidats décrivant la jonction (edge-edge ou vertex-edge).
        debug : bool
            Active un mode de trace interne (si implémenté dans la version courante).

        Sorties
        -------
        bool
            True  : chevauchement détecté / placement rejeté.
            False : placement admissible (contour final simple et valide).

        Traitement
        ----------
        1) Vérifications et normalisations (groupes, ids, attachments).
        2) Reconstruit les rings géométriques (monde) des deux groupes depuis leurs boundary cycles
           via `_build_ring_from_boundary_cycle(...)`.
        3) Analyse les attachments pour construire deux jonctions J0/J1 (types, ids, points, directions)
           via `_resolveJunctionsMobDest(...)`.
        4) Calcule la pose rigide (R, T) qui aligne la partie mobile sur la destination à partir de J0/J1,
           puis applique (R, T) au ring mobile.
        5) Si J1 implique une jonction sur arête (type "edge"), injecte un pseudo-nœud (split)
           sur le ring concerné, en choisissant le côté d’insertion selon la direction de l’arête.
        6) Résout les indices de coupe (positions de J0 et J1 dans les rings) et construit deux arcs
           complémentaires entre ces jonctions. Le choix Forward/Reverse est fait pour respecter le
           sens du boundary (CW dans le modèle courant).
        7) Concatène les deux arcs pour former le ring de sortie (`ring_out`).
        8) Valide le ring final avec `_isValidPolygon` (simple + polygon valide + aire > eps).
           Ring invalide ⇒ placement rejeté (chevauchement).
        """
        def _dbg(msg: str) -> None:
            if debug:
                print(str(msg))

        def _fmt_feature(f) -> str:
            if f is None:
                return "None"
            return f"{str(getattr(f, 'feature_type', ''))}:{getattr(f, 'element_id', '')}:{getattr(f, 'index', '')}"

        def _fmt_atts(att_list: list[TopologyAttachment]) -> str:
            out = []
            for a in att_list:
                kind = str(getattr(a, "kind", ""))
                fa = _fmt_feature(getattr(a, "feature_a", None))
                fb = _fmt_feature(getattr(a, "feature_b", None))
                params = getattr(a, "params", {}) or {}
                keys = {}
                for k in ("t", "edgeFrom", "mapping"):
                    if k in params:
                        keys[k] = params.get(k)
                out.append(f"{kind}({fa}<->{fb}){keys}")
            return "[" + ", ".join(out) + "]"

        # --- checks ---
        if groupMob is None or groupDest is None:
            return True

        gid_mob = self.find_group(str(groupMob.group_id))
        gid_dest = self.find_group(str(groupDest.group_id))
        if not gid_mob or not gid_dest or gid_mob == gid_dest:
            return True

        atts = list(attachments or [])
        if not atts:
            return True

        _dbg(f"[TOPO-OVERLAP][START] mob={gid_mob} dest={gid_dest} atts={_fmt_atts(atts)}")

        # --- build rings from boundary cycles ---
        eps_world = float(getattr(self, "overlap_eps_world", 1e-9))
        ring_mob = self._build_ring_from_boundary_cycle(gid_mob, eps_world=eps_world)
        ring_dest = self._build_ring_from_boundary_cycle(gid_dest, eps_world=eps_world)
        if ring_mob is None or ring_dest is None:
            return True
        nodes_mob, pts_mob, index_mob = ring_mob[1], ring_mob[0], ring_mob[2]
        nodes_dest, pts_dest, index_dest = ring_dest[1], ring_dest[0], ring_dest[2]

        _dbg(f"[TOPO-OVERLAP][RINGS] nMob={len(pts_mob)} nDest={len(pts_dest)}")

        # --- resolve junctions ---
        res = self._resolveJunctionsMobDest(atts, nodes_mob, pts_mob, index_mob, nodes_dest, pts_dest, index_dest)
        if res is None:
            return True
        J0, J1 = res
        _dbg(f"[TOPO-OVERLAP][J] J0={J0}")
        _dbg(f"[TOPO-OVERLAP][J] J1={J1}")

        # --- compute rigid transform (mob -> dest) ---
        RT = self._computeRigidTransform(J0, J1)
        if RT is None:
            return True
        R, T = RT
        _dbg(f"[TOPO-OVERLAP][RT] R={R.tolist()} T={[float(T[0]), float(T[1])]}")

        # Apply transform to mob ring
        pts_mob = [(R @ np.array([float(p[0]), float(p[1])], dtype=float)) + T for p in pts_mob]

        # --- inject split by edge direction (only for J1 edge) ---
        if J1.get("mobType") == "edge":
            edge_dir_rot = R @ np.array(J1.get("mobDir"), dtype=float)
            split_pt = (R @ np.array([float(J1["mobPt"][0]), float(J1["mobPt"][1])], dtype=float)) + T
            res_split = self._injectSplitByEdgeDir(nodes_mob, pts_mob, index_mob, J0.get("mobId"), split_pt, edge_dir_rot)
            if res_split is None:
                return True
            nodes_mob, pts_mob, index_mob, edge_node_id = res_split
            J1["mobType"] = "node"
            J1["mobId"] = edge_node_id
            J1["mobPt"] = split_pt

        if J1.get("destType") == "edge":
            edge_dir_rot = np.array(J1.get("destDir"), dtype=float)
            split_pt = np.array([float(J1["destPt"][0]), float(J1["destPt"][1])], dtype=float)
            res_split = self._injectSplitByEdgeDir(nodes_dest, pts_dest, index_dest, J0.get("destId"), split_pt, edge_dir_rot)
            if res_split is None:
                return True
            nodes_dest, pts_dest, index_dest, edge_node_id = res_split
            J1["destType"] = "node"
            J1["destId"] = edge_node_id
            J1["destPt"] = split_pt

        # --- resolve cut indices ---
        if J0.get("mobId") not in index_mob or J1.get("mobId") not in index_mob:
            return True
        if J0.get("destId") not in index_dest or J1.get("destId") not in index_dest:
            return True
        i_m0 = int(index_mob[J0.get("mobId")])
        i_m1 = int(index_mob[J1.get("mobId")])
        i_d0 = int(index_dest[J0.get("destId")])
        i_d1 = int(index_dest[J1.get("destId")])
        if i_m0 == i_m1 or i_d0 == i_d1:
            return True
        _dbg(f"[TOPO-OVERLAP][CUT] iMob=({i_m0},{i_m1}) iDest=({i_d0},{i_d1})")

        # --- build output ring ---
        # on part du point __EDGE_ pour revenir à refNode
        def arcEntreJonctionsSelonEdge(pts, j0, j1):
            """
            Renvoie l'arc entre a et b en choisissant Forward/Reverse selon le sens de l'edge j0->j1
            par rapport au boundary CW.
            """
            n = len(pts)
            if j1 == (j0 + 1) % n:       # edge suit CW
                return self._arcReverse(pts, j0, j1)
            elif j1 == (j0 - 1) % n:     # edge inversé
                return self._arcForward(pts, j0, j1)

        arc_mob = arcEntreJonctionsSelonEdge(pts_mob, i_m0, i_m1)
        arc_dest = arcEntreJonctionsSelonEdge(pts_dest, i_d0, i_d1)
        ring_out = self._concatArcs(arc_mob, arc_dest)
        ring_xy = [(float(p[0]), float(p[1])) for p in ring_out]
        poly = _ShPoly(ring_xy) if ring_xy else _ShPoly()
        area = float(getattr(poly, "area", 0.0) or 0.0)
        _dbg(f"[TOPO-OVERLAP][RINGOUT] n={len(ring_xy)} area={area:.6f}")

        # --- validate final ring ---
        is_valid = self._isValidPolygon(ring_out)
        overlap = not is_valid
        _dbg(f"[TOPO-OVERLAP][VALID] overlap={overlap}")
        return overlap

    # --- Parcours du graphe conceptuel (par groupe) ---
    def getConceptEdgesForNode(self, node_id: str, group_id: str | None = None) -> list[dict]:
        """Liste des edges conceptuels incidentes à un concept node (par groupe)."""
        gid = group_id if group_id is not None else self._infer_group_for_node(node_id)
        if gid is None:
            return []
        cn = self.find_node(str(node_id))
        c = self.ensureConceptGraph(gid)
        out: list[dict] = []
        for (a, b), ce in c.edges.items():
            if cn != a and cn != b:
                continue
            other = b if cn == a else a
            out.append({
                "a": a, "b": b,
                "this": cn,
                "other": other,
                "occurrences": list(ce.occurrences),
            })
        return out

    def getConceptNeighborNodes(self, node_id: str, group_id: str | None = None) -> list[str]:
        """Retourne les voisins conceptuels directs d'un node.
        Renvoie une liste d'IDs canoniques, uniques et stables.
        Garantit que les voisins sont limités au groupe résolu.
        Ne calcule pas d'azimut ni de géométrie.
        Ne modifie pas la topologie ni les caches.
        """
        edges = self.getConceptEdgesForNode(node_id, group_id)
        neigh = [e["other"] for e in edges]
        # unique + stable
        seen = set()
        out = []
        for n in neigh:
            if n in seen:
                continue
            seen.add(n)
            out.append(n)
        return out

    def getConceptNodeType(self, node_id: str) -> str:
        """Type conceptuel: 'M' si mélange de types, sinon type unique."""
        cn = self.find_node(str(node_id))
        mem = self.node_members(cn)
        if not mem:
            return str(self.node_type(cn))
        types = {str(self.node_type(m)) for m in mem}
        return "M" if len(types) > 1 else next(iter(types))

    # --- IDs helpers ---
    @staticmethod
    def format_element_id(tri_rank_1based: int) -> str:
        return f"T{int(tri_rank_1based):02d}"

    def format_edge_id(self, element_id: str, edge_index: int) -> str:
        return f"{str(element_id)}:E{int(edge_index)}"

    def format_node_id(self, element_id: str, vertex_index: int) -> str:
        return f"{str(element_id)}:N{int(vertex_index)}"

    @staticmethod
    def parse_tri_rank_from_element_id(element_id: str) -> int | None:
        s = str(element_id or "").strip()
        m = re.fullmatch(r"T(\d+)", s) or re.search(r":T(\d+)$", s)
        if not m:
            return None
        return int(m.group(1))

    # Azimut (C7) – Référence unique

    def azimutDegFromDxDy(self, dx: float, dy: float) -> float:
        """Calcule l'azimut standard (0°=Nord, sens horaire) à partir d'un delta.

        Convention :
        - dx = Est  (x2 - x1)
        - dy = variation sur l'axe Y du repère monde courant.

        IMPORTANT : si `self.worldYAxisDown` est True (repère image/écran), alors
        `dy` pointe vers le Sud quand il est positif. On convertit donc en dyNord = -dy
        avant d'appliquer la formule de référence.

        Retour : angle normalisé dans [0..360).
        """
        dx = float(dx)
        dy = float(dy)
        dy_north = -dy if bool(getattr(self, "worldYAxisDown", True)) else dy
        a = math.degrees(math.atan2(dx, dy_north))  # atan2(E, N)
        return float((a + 360.0) % 360.0)

    def new_group_id(self) -> str:
        self._created_counter_groups += 1
        gid = f"G{self._created_counter_groups:03d}"
        self._group_created_order[gid] = self._created_counter_groups
        return gid

    def new_attachment_id(self) -> str:
        self._created_counter_attachments += 1
        return f"A{self._created_counter_attachments:03d}"

    # --- DSU nodes ---
    def create_node_atomic(self, node_id: str, node_type: str) -> str:
        node_id = str(node_id)
        self._created_counter_nodes += 1
        self._node_parent[node_id] = node_id
        self._node_type[node_id] = str(node_type)
        self._node_members[node_id] = [node_id]
        self._node_created_order[node_id] = self._created_counter_nodes
        return node_id

    def find_node(self, node_id: str) -> str:
        node_id = str(node_id)
        parent = self._node_parent.get(node_id, node_id)
        if parent != node_id:
            self._node_parent[node_id] = self.find_node(parent)
        return self._node_parent.get(node_id, node_id)

    def _node_rank_key(self, node_id: str) -> tuple[int, int, str]:
        c = self.find_node(node_id)
        t = self._node_type.get(c, TopologyNodeType.OUVERTURE)
        created = self._node_created_order.get(c, 0)
        return (TopologyNodeType.rank(t), created, c)

    def union_nodes(self, a: str, b: str) -> str:
        ra = self.find_node(a)
        rb = self.find_node(b)
        if ra == rb:
            return ra
        canonical = ra if self._node_rank_key(ra) >= self._node_rank_key(rb) else rb
        other = rb if canonical == ra else ra
        self._node_parent[other] = canonical
        mem = self._node_members.get(canonical, [canonical])
        mem_other = self._node_members.get(other, [other])
        self._node_members[canonical] = mem + [x for x in mem_other if x not in mem]
        self.invalidateConceptGraph(None)
        return canonical

    def node_members(self, node_id: str) -> list[str]:
        c = self.find_node(node_id)
        return list(self._node_members.get(c, [c]))

    def node_type(self, node_id: str) -> str:
        c = self.find_node(node_id)
        return self._node_type.get(c, TopologyNodeType.OUVERTURE)

    # --- DSU groups ---
    def create_group_atomic(self) -> str:
        gid = self.new_group_id()
        self._group_parent[gid] = gid
        self._group_members[gid] = [gid]
        self.groups[gid] = TopologyGroup(gid)
        return gid

    # API camelCase (spec UI/Core)
    def createAtomicGroup(self) -> str:
        return self.create_group_atomic()

    def rebuildGroupElementLists(self) -> None:
        for g in self.groups.values():
            g.element_ids = []
        for eid, gid in self.element_to_group.items():
            gidc = self.find_group(gid)
            if gidc in self.groups:
                if eid not in self.groups[gidc].element_ids:
                    self.groups[gidc].element_ids.append(eid)

    def find_group(self, group_id: str) -> str:
        group_id = str(group_id)
        parent = self._group_parent.get(group_id, group_id)
        if parent != group_id:
            self._group_parent[group_id] = self.find_group(parent)
        return self._group_parent.get(group_id, group_id)

    def _group_rank_key(self, group_id: str) -> tuple[int, str]:
        c = self.find_group(group_id)
        created = self._group_created_order.get(c, 0)
        return (created, c)

    def union_groups(self, a: str, b: str) -> str:
        ra = self.find_group(a)
        rb = self.find_group(b)
        if ra == rb:
            return ra
        canonical = ra if self._group_rank_key(ra) >= self._group_rank_key(rb) else rb
        other = rb if canonical == ra else ra
        self._group_parent[other] = canonical
        mem = self._group_members.get(canonical, [canonical])
        mem_other = self._group_members.get(other, [other])
        self._group_members[canonical] = mem + [x for x in mem_other if x not in mem]
        if canonical in self.groups and other in self.groups:
            self.groups[canonical].element_ids.extend(self.groups[other].element_ids)
            self.groups[canonical].attachment_ids.extend(self.groups[other].attachment_ids)
        self.invalidateConceptGraph(canonical)
        self._markTopoTouched(canonical)
        return canonical

    def group_members(self, group_id: str) -> list[str]:
        c = self.find_group(group_id)
        return list(self._group_members.get(c, [c]))

    # --- Elements ---
    def add_element_as_new_group(self, element: TopologyElement) -> str:
        if element.element_id in self.elements:
            raise ValueError(f"TopologyWorld: elementId déjà présent: {element.element_id}")
        gid = self.create_group_atomic()
        self.elements[element.element_id] = element
        self.element_to_group[element.element_id] = gid
        self.groups[gid].element_ids.append(element.element_id)

        element.vertexes = []
        for i, (lab, typ) in enumerate(zip(element.vertex_labels, element.vertex_types)):
            nid = self.format_node_id(element.element_id, i)
            self.create_node_atomic(nid, typ)
            element.vertexes.append(TopologyVertex(element.element_id, i, lab, nid, typ))

        n = element.n_vertices()
        element.edges = []
        for i in range(n):
            v0 = element.vertexes[i]
            v1 = element.vertexes[(i + 1) % n]
            element.edges.append(TopologyEdge(element.element_id, i, v0, v1, element.edge_lengths_km[i]))

        self.invalidateConceptGraph(gid)
        self._markTopoTouched(gid)
        return gid

    def get_edge(self, element_id: str, edge_index: int) -> TopologyEdge:
        return self.elements[str(element_id)].edges[int(edge_index)]

    def get_vertex(self, element_id: str, vertex_index: int) -> TopologyVertex:
        return self.elements[str(element_id)].vertexes[int(vertex_index)]

    def get_group_of_element(self, element_id: str) -> str:
        gid0 = self.element_to_group.get(str(element_id))
        if gid0 is None:
            raise ValueError(f"Element sans groupe: {element_id}")
        return self.find_group(gid0)

    # --- API pose par élément (session) ---
    def getElementPose(self, element_id: str) -> tuple[np.ndarray, np.ndarray, bool]:
        """Retourne (R, T, mirrored) pour un élément."""
        el = self.elements.get(str(element_id))
        if el is None:
            raise ValueError(f"Element inconnu: {element_id}")
        return el.get_pose()

    def setElementPose(self, element_id: str, R: np.ndarray, T: np.ndarray, mirrored: bool | None = None) -> None:
        """Assigne la pose d'un élément (R,T,j'ai )."""
        el = self.elements.get(str(element_id))
        if el is None:
            raise ValueError(f"Element inconnu: {element_id}")
        el.set_pose(R=R, T=T, mirrored=mirrored)
        # pose change -> positions monde changent -> géométrie conceptuelle invalide
        self.invalidateConceptGeom(self.element_to_group.get(str(element_id)))

    # ------------------------------------------------------------------
    # Flip (symétrie axiale) au niveau GROUPE (mais appliquée aux poses d'éléments)
    # - groupes = topologiques
    # - LocalCoords intacts
    # - on met à jour (R,T,mirrored) pour chaque élément du groupe
    # ------------------------------------------------------------------
    @staticmethod
    def _closest_rotation(R: np.ndarray) -> np.ndarray:
        """Projette une matrice 2x2 proche d'une rotation sur SO(2) (det=+1)."""
        U, _S, Vt = np.linalg.svd(np.array(R, float))
        R2 = U @ Vt
        # Forcer det=+1
        if float(np.linalg.det(R2)) < 0.0:
            U[:, -1] *= -1.0
            R2 = U @ Vt
        return np.array(R2, float)

    @staticmethod
    def _reflection_matrix(axis_dir: np.ndarray) -> np.ndarray:
        """Matrice de réflexion (2D) par rapport à une droite passant par l'origine, dirigée par axis_dir."""
        u = np.array(axis_dir, float)
        n = float(np.hypot(u[0], u[1]))
        if n < 1e-12:
            raise ValueError("flipGroup: axisDir dégénéré")
        u /= n
        # S = 2*u*u^T - I
        I = np.eye(2, dtype=float)
        uuT = np.outer(u, u)
        return (2.0 * uuT) - I

    def flipGroup(self, group_id: str, axisPoint: np.ndarray, axisDir: np.ndarray) -> None:
        """Applique une symétrie axiale à tous les éléments du groupe.

        Contrat:
        - géométrie monde: x' = S*(x - p) + p  (p = axisPoint, S = reflection(axisDir))
        - mise à jour des poses éléments uniquement
        - LocalCoords inchangés
        """
        gid = self.find_group(str(group_id))
        g = self.groups.get(gid)
        if g is None:
            return

        p = np.array(axisPoint, float).reshape((2,))
        S = self._reflection_matrix(axisDir)
        b = p - (S @ p)  # translation affine pour une droite passant par p

        M = TopologyElementPose2D.mirror_matrix()

        for eid in list(g.element_ids):
            el = self.elements.get(str(eid))
            if el is None:
                continue
            R, T, mirrored = el.get_pose()
            R = np.array(R, float)
            T = np.array(T, float).reshape((2,))

            # A = rotation * (miroir local si mirrored)
            A = R @ (M if mirrored else np.eye(2, dtype=float))

            # Après réflexion monde: x' = S*(A p + T) + b = (S A) p + (S T + b)
            A2 = S @ A
            T2 = (S @ T) + b

            # Re-factorisation A2 en (R', mirrored') avec R' rotation pure det=+1
            detA2 = float(np.linalg.det(A2))
            if detA2 < 0.0:
                mirrored2 = True
                R2_raw = A2 @ M  # car A2 = R2 * M  => R2 = A2 * M
            else:
                mirrored2 = False
                R2_raw = A2

            R2 = self._closest_rotation(R2_raw)
            el.set_pose(R=R2, T=T2, mirrored=mirrored2)

    def elementLocalToWorld(self, element_id: str, p_local: np.ndarray | tuple[float, float]) -> np.ndarray:
        el = self.elements.get(str(element_id))
        if el is None:
            raise ValueError(f"Element inconnu: {element_id}")
        return el.localToWorld(p_local)

    def get_element_world_pose(self, element_id: str) -> TopologyPose2D:
        """Pose monde dérivée d'un élément.

        NOTE compat : TopologyPose2D ne porte pas mirrored.
        - si mirrored==False : retourne (R,T) tel quel
        - si mirrored==True  : retourne (R@M, T) (det négatif)
        """
        R, T, mirrored = self.getElementPose(str(element_id))
        if mirrored:
            R = np.array(R, float) @ TopologyElementPose2D.mirror_matrix()
        return TopologyPose2D(R=R, T=T)

    # ------------------------------------------------------------------
    # Pose Edge↔Edge (V1) : calcule la pose ABSOLUE (R,T) du groupe mobile
    # pour aligner une arête mobile sur une arête cible.
    # - mapping="direct" : (start_m -> start_t) et (end_m -> end_t)
    # - mapping="reverse": (start_m -> end_t) et (end_m -> start_t)
    # Contrat :
    #   - retourne (R_abs, T_abs) : Monde <- GroupeMobile
    #   - aucune modif de state (pur calcul)
    # ------------------------------------------------------------------
    def compute_pose_edge_to_edge(
        self,
        group_id_m: str,
        element_id_m: str,
        edge_index_m: int,
        group_id_t: str,
        element_id_t: str,
        edge_index_t: int,
        mapping: str = "direct",
    ) -> tuple[np.ndarray, np.ndarray]:
        """Calcule une pose ABSOLUE (R,T) pour l'ÉLÉMENT mobile afin d'aligner une arête mobile sur une arête cible.

        NOTE: les groupes sont topologiques (pas de pose groupe).
        - Cette fonction sert uniquement de helper géométrique éventuel côté UI/tests.
        - Elle ne modifie aucun state.

        Contrat:
        - retourne (R_abs, T_abs) tel que:
            world = (R_abs @ (M? @ p_local)) + T_abs
            avec mirrored supposé False (si mirrored est nécessaire, c'est à l'UI de le gérer).

        mapping:
        - "direct"  : start_m -> start_t  et end_m -> end_t
        - "reverse" : start_m -> end_t    et end_m -> start_t
        """
        # On ignore group_id_* (compat API), tout est porté par les poses d'éléments.
        e_m = self.get_edge(str(element_id_m), int(edge_index_m))
        e_t = self.get_edge(str(element_id_t), int(edge_index_t))

        el_m = self.elements[str(element_id_m)]
        el_t = self.elements[str(element_id_t)]

        def _v_local(el: TopologyElement, vidx: int) -> np.ndarray:
            xy = el.vertex_local_xy.get(int(vidx))
            if xy is None:
                raise ValueError(
                    f"compute_pose_edge_to_edge: vertex_local_xy manquant (element={el.element_id} idx={vidx})"
                )
            return np.array([float(xy[0]), float(xy[1])], dtype=float)

        # Endpoints LOCAL (mobile)
        p0m = _v_local(el_m, e_m.v_start.vertex_index)
        p1m = _v_local(el_m, e_m.v_end.vertex_index)

        # Endpoints WORLD (cible) via pose élément
        p0t = _v_local(el_t, e_t.v_start.vertex_index)
        p1t = _v_local(el_t, e_t.v_end.vertex_index)
        w0t = self.elementLocalToWorld(str(element_id_t), p0t)
        w1t = self.elementLocalToWorld(str(element_id_t), p1t)

        if str(mapping).lower() == "direct":
            W0, W1 = w0t, w1t
        elif str(mapping).lower() == "reverse":
            W0, W1 = w1t, w0t
        else:
            raise ValueError(f"compute_pose_edge_to_edge: mapping invalide: {mapping}")

        vL = (p1m - p0m)
        vW = (W1 - W0)

        nL = float(np.linalg.norm(vL))
        nW = float(np.linalg.norm(vW))
        if nL <= 1e-12 or nW <= 1e-12:
            raise ValueError("compute_pose_edge_to_edge: arête dégénérée (norme ~0)")

        angL = float(np.arctan2(vL[1], vL[0]))
        angW = float(np.arctan2(vW[1], vW[0]))
        dtheta = angW - angL

        c = float(np.cos(dtheta))
        s = float(np.sin(dtheta))
        R_abs = np.array([[c, -s], [s,  c]], dtype=float)

        # T : amener p0m sur W0
        T_abs = np.array(W0, dtype=float) - (R_abs @ np.array(p0m, dtype=float))
        return (R_abs, T_abs)

    def apply_attachments(self, attachments: list[TopologyAttachment]) -> str:
        """Applique une liste d'attachments (transaction simple V1) et retourne le gid canonique final."""
        gid_last: str | None = None
        for att in list(attachments or []):
            # Si ID vide OU collision => on régénère ici.
            if (att.attachment_id is None) or (str(att.attachment_id) in self.attachments):
                att.attachment_id = self.new_attachment_id()
            gid_last = self.apply_attachment(att)
        return str(gid_last) if gid_last is not None else ""

    def _edge_endpoints_atomic_nodes(self, edge: TopologyEdge) -> tuple[str, str]:
        return (edge.v_start.node_id, edge.v_end.node_id)

    # --- points on edge ---
    # --- attachments ---
    def _validate_attachment_p2(self, attachment: TopologyAttachment, eps_t: float = 1e-9) -> None:
        if self._isImportingSnapshot:
            return
        kind = str(getattr(attachment, "kind", "") or "")
        if kind not in ("vertex-vertex", "vertex-edge", "edge-edge"):
            raise ValueError(f"[P2][Attachment] invalid type={kind} id={attachment.attachment_id}")

        if attachment.feature_a is None or attachment.feature_b is None:
            raise ValueError(f"[P2][Attachment] missing featureA/featureB type={kind} id={attachment.attachment_id}")

        def require_vertex_ref(f, side: str) -> None:
            if f.feature_type != TopologyFeatureType.VERTEX:
                raise ValueError(
                    f"[P2][Attachment] expected VertexRef for {side} got={f.feature_type} id={attachment.attachment_id}"
                )
            element = self.elements.get(str(f.element_id))
            if element is None:
                raise ValueError(
                    f"[P2][Attachment] unknown element elementId={f.element_id} in feature{side} id={attachment.attachment_id}"
                )
            vidx = int(f.index)
            if vidx < 0 or vidx >= len(element.vertexes):
                attid = attachment.attachment_id
                eid = element.element_id
                raise ValueError(
                    f"[P2][Attachment] unknown vertex vertexId=N{vidx} in elementId={eid} feature{side} id={attid}"
                )

        def require_edge_ref(f, side: str) -> None:
            if f.feature_type != TopologyFeatureType.EDGE:
                raise ValueError(
                    f"[P2][Attachment] expected EdgeRef for {side} got={f.feature_type} id={attachment.attachment_id}"
                )
            eid = f.element_id
            attid = attachment.attachment_id
            element = self.elements.get(str(eid))
            if element is None:
                raise ValueError(
                    f"[P2][Attachment] unknown element elementId={eid} in feature{side} id={attid}"
                )
            eidx = int(f.index)
            if eidx < 0 or eidx >= len(element.edges):
                raise ValueError(
                    f"[P2][Attachment] unknown edge edgeId=E{eidx} in elementId={element.element_id} feature{side} id={attid}"
                )

        if kind == "vertex-vertex":
            require_vertex_ref(attachment.feature_a, "A")
            require_vertex_ref(attachment.feature_b, "B")
            if attachment.params.get("t", None) is not None:
                raise ValueError(
                    f"[P2][Attachment] vertex-vertex must not have t (t={attachment.params.get('t')}) id={attachment.attachment_id}"
                )
            return

        if kind == "edge-edge":
            require_edge_ref(attachment.feature_a, "A")
            require_edge_ref(attachment.feature_b, "B")
            if attachment.params.get("t", None) is not None:
                raise ValueError(
                    f"[P2][Attachment] edge-edge must not have t (t={attachment.params.get('t')}) id={attachment.attachment_id}"
                )
            return

        # vertex-edge
        require_vertex_ref(attachment.feature_a, "A")
        require_edge_ref(attachment.feature_b, "B")
        if "t" not in attachment.params or attachment.params.get("t", None) is None:
            raise ValueError(f"[P2][Attachment] vertex-edge missing t id={attachment.attachment_id}")
        try:
            t_val = float(attachment.params.get("t"))
        except Exception:
            raise ValueError(
                f"[P2][Attachment] vertex-edge t not float (t={attachment.params.get('t')}) id={attachment.attachment_id}"
            )
        if t_val < -eps_t or t_val > 1.0 + eps_t:
            raise ValueError(
                f"[P2][Attachment] vertex-edge t out of [0,1] (t={t_val:.6f}) id={attachment.attachment_id}"
            )
        if abs(t_val) <= eps_t:
            t_val = 0.0
        elif abs(t_val - 1.0) <= eps_t:
            t_val = 1.0
        attachment.params["t"] = t_val
        edge_from = attachment.params.get("edgeFrom", None)
        if edge_from is None or str(edge_from).strip() == "":
            raise ValueError(f"[P2][Attachment] vertex-edge missing edgeFrom id={attachment.attachment_id}")
        edge_from = str(edge_from)
        e = self.get_edge(attachment.feature_b.element_id, attachment.feature_b.index)
        s = str(e.v_start.node_id)
        t = str(e.v_end.node_id)
        if edge_from != s and edge_from != t:
            raise ValueError(
                f"[P2][Attachment] vertex-edge edgeFrom not on edge (edgeFrom={edge_from}) id={attachment.attachment_id}"
            )
        attachment.params["edgeFrom"] = edge_from

    def record_attachment(self, attachment: TopologyAttachment, group_id: str | None = None):
        self.attachments[attachment.attachment_id] = attachment
        if group_id is not None:
            self.groups[self.find_group(group_id)].attachment_ids.append(attachment.attachment_id)

    def _apply_attachment_internal(self, attachment: TopologyAttachment) -> str:
        gA = self.get_group_of_element(attachment.feature_a.element_id)
        gB = self.get_group_of_element(attachment.feature_b.element_id)
        gC = self.union_groups(gA, gB) if gA != gB else gA

        kind = str(attachment.kind)
        if kind == "vertex-vertex":
            vA = self.get_vertex(attachment.feature_a.element_id, attachment.feature_a.index)
            vB = self.get_vertex(attachment.feature_b.element_id, attachment.feature_b.index)
            self.union_nodes(vA.node_id, vB.node_id)

        elif kind == "vertex-edge":
            a, b = attachment.feature_a, attachment.feature_b
            ta, tb = a.feature_type, b.feature_type

            if ta == TopologyFeatureType.VERTEX and tb == TopologyFeatureType.EDGE:
                vRef, eRef = a, b
            elif ta == TopologyFeatureType.EDGE and tb == TopologyFeatureType.VERTEX:
                vRef, eRef = b, a
            else:
                raise ValueError("vertex-edge attend un vertexRef et un edgeRef")

            t = attachment.params.get("t", None)
            if t is None:
                t = vRef.t if vRef.t is not None else eRef.t
            if t is None:
                raise ValueError("vertex-edge: parametre t manquant")
            edge_from = attachment.params.get("edgeFrom", None)
            if edge_from is None or str(edge_from).strip() == "":
                raise ValueError("vertex-edge: parametre edgeFrom manquant")

            v = self.get_vertex(vRef.element_id, vRef.index)
            e = self.get_edge(eRef.element_id, eRef.index)
            edge_from = str(edge_from)
            if edge_from == str(e.v_start.node_id):
                t_val = float(t)
            elif edge_from == str(e.v_end.node_id):
                t_val = 1.0 - float(t)
            else:
                raise ValueError(f"vertex-edge: edgeFrom incoherent (edgeFrom={edge_from})")
            eps = float(getattr(self, "vertex_edge_endpoint_epsilon", 1e-6))
            if abs(t_val) <= eps:
                self.union_nodes(v.node_id, e.v_start.node_id)
            elif abs(t_val - 1.0) <= eps:
                self.union_nodes(v.node_id, e.v_end.node_id)

        elif kind == "edge-edge":
            eA = self.get_edge(attachment.feature_a.element_id, attachment.feature_a.index)
            eB = self.get_edge(attachment.feature_b.element_id, attachment.feature_b.index)

            mapping = str(attachment.params.get("mapping", "direct"))
            a0, a1 = self._edge_endpoints_atomic_nodes(eA)
            b0, b1 = self._edge_endpoints_atomic_nodes(eB)

            if mapping == "direct":
                self.union_nodes(a0, b0)
                self.union_nodes(a1, b1)
            elif mapping == "reverse":
                self.union_nodes(a0, b1)
                self.union_nodes(a1, b0)
            else:
                raise ValueError(f"edge-edge: mapping invalide: {mapping}")

            eA.coverages.append(TopologyCoverageInterval(0.0, 1.0))
            eB.coverages.append(TopologyCoverageInterval(0.0, 1.0))

        else:
            raise ValueError(f"Attachment kind non supporte: {kind}")

        self.record_attachment(attachment, group_id=gC)
        self.invalidateConceptGraph(gC)
        self._markTopoTouched(gC)
        return gC

    def loadAttachment(self, attachment: TopologyAttachment) -> str:
        """Load one attachment from snapshot in strict passive mode."""
        gA = self.get_group_of_element(attachment.feature_a.element_id)
        gB = self.get_group_of_element(attachment.feature_b.element_id)
        gC = self.union_groups(gA, gB) if gA != gB else gA

        kind = str(attachment.kind)
        if kind == "vertex-vertex":
            vA = self.get_vertex(attachment.feature_a.element_id, attachment.feature_a.index)
            vB = self.get_vertex(attachment.feature_b.element_id, attachment.feature_b.index)
            self.union_nodes(vA.node_id, vB.node_id)

        elif kind == "vertex-edge":
            a, b = attachment.feature_a, attachment.feature_b
            ta, tb = a.feature_type, b.feature_type
            if ta == TopologyFeatureType.VERTEX and tb == TopologyFeatureType.EDGE:
                vRef, eRef = a, b
            elif ta == TopologyFeatureType.EDGE and tb == TopologyFeatureType.VERTEX:
                vRef, eRef = b, a
            else:
                raise ValueError("Snapshot invalide : vertex-edge attend un vertexRef et un edgeRef")

            v = self.get_vertex(vRef.element_id, vRef.index)
            e = self.get_edge(eRef.element_id, eRef.index)
            edge_from = attachment.params.get("edgeFrom", None)
            if edge_from is None:
                raise ValueError("Snapshot invalide : edgeFrom non porte par l'arete referencee")
            edge_from = str(edge_from)
            s = str(e.v_start.node_id)
            tnode = str(e.v_end.node_id)
            if edge_from != s and edge_from != tnode:
                raise ValueError("Snapshot invalide : edgeFrom non porte par l'arete referencee")

            t_raw = attachment.params.get("t", None)
            if t_raw is not None:
                t_val = float(t_raw)
                if edge_from == s:
                    if t_val == 0.0:
                        self.union_nodes(v.node_id, e.v_start.node_id)
                    elif t_val == 1.0:
                        self.union_nodes(v.node_id, e.v_end.node_id)
                else:
                    if t_val == 0.0:
                        self.union_nodes(v.node_id, e.v_end.node_id)
                    elif t_val == 1.0:
                        self.union_nodes(v.node_id, e.v_start.node_id)

        elif kind == "edge-edge":
            eA = self.get_edge(attachment.feature_a.element_id, attachment.feature_a.index)
            eB = self.get_edge(attachment.feature_b.element_id, attachment.feature_b.index)
            mapping = str(attachment.params.get("mapping", "direct"))
            a0, a1 = self._edge_endpoints_atomic_nodes(eA)
            b0, b1 = self._edge_endpoints_atomic_nodes(eB)
            if mapping == "direct":
                self.union_nodes(a0, b0)
                self.union_nodes(a1, b1)
            elif mapping == "reverse":
                self.union_nodes(a0, b1)
                self.union_nodes(a1, b0)
            else:
                raise ValueError(f"Snapshot invalide : mapping edge-edge invalide ({mapping})")
            eA.coverages.append(TopologyCoverageInterval(0.0, 1.0))
            eB.coverages.append(TopologyCoverageInterval(0.0, 1.0))

        else:
            raise ValueError(f"Snapshot invalide : kind d'attachment non supporte ({kind})")

        self.record_attachment(attachment, group_id=gC)
        self.invalidateConceptGraph(gC)
        self._markTopoTouched(gC)
        return gC

    def apply_attachment(self, attachment: TopologyAttachment) -> str:
        self._validate_attachment_p2(attachment)
        return self._apply_attachment_internal(attachment)

    def rebuild_from_attachments(self, attachments: list[TopologyAttachment]):
        self.attachments = {}
        for g in self.groups.values():
            g.attachment_ids = []

        for nid in list(self._node_parent.keys()):
            self._node_parent[nid] = nid
            self._node_members[nid] = [nid]

        for gid in list(self._group_parent.keys()):
            self._group_parent[gid] = gid
            self._group_members[gid] = [gid]

        for el in self.elements.values():
            for edge in el.edges:
                edge.coverages = []

        for att in sorted(attachments, key=lambda a: a.attachment_id):
            self.apply_attachment(att)

    def removeElementsAndRebuild(self, elementIds: list[str]) -> None:
        """Supprime des éléments du monde (core) et reconstruit la topologie à partir des attachments restants.

        Utilisation typique : suppression UI d'un groupe (plusieurs triangles) côté Tk.
        Contrat:
        - supprime les éléments + leurs noeuds atomiques
        - purge tous les attachments qui référencent un des éléments supprimés
        - reconstruit les DSU nodes/groups + coverages via rebuild_from_attachments()
        - invalide la géométrie conceptuelle des groupes touchés

        Note:
        - on ne renumérote pas les elementId.
        - on évite de changer les groupId existants des éléments restants.
        """
        if not elementIds:
            return
        if int(getattr(self, "_topoTxDepth", 0)) > 0:
            raise ValueError("TopologyWorld.removeElementsAndRebuild: transaction ouverte")

        removed: set[str] = {str(e).strip() for e in (elementIds or []) if str(e).strip()}
        if not removed:
            return

        # Snapshot des groupes concernés (pour invalider / nettoyer des sous-systèmes, ex: chemins)
        removedGroups: set[str] = set()
        for eid in list(removed):
            gid0 = self.element_to_group.get(str(eid))
            if gid0 is not None:
                removedGroups.add(str(self.find_group(str(gid0))))

        # 1) Filtrer les attachments conservés (avant de modifier self.attachments)
        keptAttachments: list[TopologyAttachment] = []
        for att in list(getattr(self, "attachments", {}).values()):
            try:
                a_eid = str(att.feature_a.element_id)
                b_eid = str(att.feature_b.element_id)
            except Exception as e:
                raise ValueError("TopologyWorld.removeElementsAndRebuild: attachment corrompu (feature)") from e
            if a_eid in removed or b_eid in removed:
                continue
            keptAttachments.append(att)

        # 2) Supprimer les éléments + noeuds atomiques
        for eid in list(removed):
            el = self.elements.get(str(eid))
            if el is None:
                # Suppression idempotente : si l'élément n'existe déjà plus, on continue.
                continue

            # Noeuds atomiques de l'élément
            for v in list(getattr(el, "vertexes", []) or []):
                nid = str(getattr(v, "node_id", ""))
                if not nid:
                    continue
                # DSU nodes / metadata
                if nid in self._node_parent:
                    self._node_parent.pop(nid, None)
                if nid in self._node_members:
                    self._node_members.pop(nid, None)
                if nid in self._node_type:
                    self._node_type.pop(nid, None)
                if nid in self._node_created_order:
                    self._node_created_order.pop(nid, None)

            # Retirer l'élément du mapping element->group
            gid0 = self.element_to_group.pop(str(eid), None)
            if gid0 is not None:
                gidc = str(self.find_group(str(gid0)))
                g = self.groups.get(gidc)
                if g is not None and str(eid) in list(getattr(g, "element_ids", []) or []):
                    g.element_ids = [x for x in list(g.element_ids) if str(x) != str(eid)]

            # Retirer l'élément
            self.elements.pop(str(eid), None)

        # 3) Purger les groupes vides / orphelins
        usedRaw = {str(g) for g in (self.element_to_group.values() or [])}
        usedCanon = {str(self.find_group(g)) for g in list(usedRaw)} if usedRaw else set()

        for gid in list(self.groups.keys()):
            if str(self.find_group(str(gid))) not in usedCanon:
                self.groups.pop(gid, None)

        for gid in list(self._group_parent.keys()):
            if str(self.find_group(str(gid))) not in usedCanon:
                self._group_parent.pop(gid, None)
                self._group_members.pop(gid, None)
                self._group_created_order.pop(gid, None)

        # 4) Reconstruire la topologie depuis les attachments conservés
        self.rebuild_from_attachments(keptAttachments)
        self.rebuildGroupElementLists()

        # 5) Chemins : si le chemin courant porte sur un groupe supprimé, on le supprime.
        tc = getattr(self, "topologyChemins", None)
        if tc is not None and bool(getattr(tc, "isDefined", False)):
            gidChemin = str(getattr(tc, "groupId", "") or "")
            if gidChemin and str(self.find_group(gidChemin)) in removedGroups:
                tc.supprimerChemin()

        # 6) Invalidation conceptuelle des groupes touchés
        for gid in (usedCanon | removedGroups):
            if gid:
                self.invalidateConceptGraph(str(gid))

    # ------------------------------------------------------------------
    # Dégrouper par nœud (suppression d'attachments + rebuild complet)
    # ------------------------------------------------------------------

    def _degrouperAttachmentElementPair(self, att: TopologyAttachment) -> tuple[str, str]:
        eA = str(getattr(att.feature_a, "element_id", "") or "")
        eB = str(getattr(att.feature_b, "element_id", "") or "")
        if not eA or not eB:
            raise ValueError(f"[degrouper] attachment sans elementId (aid={att.attachment_id})")
        return (eA, eB)

    def _degrouperIterGroupScopedAttachments(self, elementIds: set[str]) -> list[TopologyAttachment]:
        out: list[TopologyAttachment] = []
        for aid in sorted(self.attachments.keys()):
            att = self.attachments[aid]
            eA, eB = self._degrouperAttachmentElementPair(att)
            if eA in elementIds and eB in elementIds:
                out.append(att)
        return out

    def _degrouperIsVertexFeatureOnNode(self, f: TopologyFeatureRef, nodeId: str) -> bool:
        if str(f.feature_type) != TopologyFeatureType.VERTEX:
            raise ValueError(f"[degrouper] feature attendue vertex, reçu={f.feature_type}")
        v = self.get_vertex(str(f.element_id), int(f.index))
        return str(self.find_node(str(v.node_id))) == str(nodeId)

    def _degrouperIsEdgeFeatureOnNode(self, f: TopologyFeatureRef, nodeId: str) -> bool:
        if str(f.feature_type) != TopologyFeatureType.EDGE:
            raise ValueError(f"[degrouper] feature attendue edge, reçu={f.feature_type}")
        e = self.get_edge(str(f.element_id), int(f.index))
        a = str(self.find_node(str(e.v_start.node_id)))
        b = str(self.find_node(str(e.v_end.node_id)))
        return (a == str(nodeId)) or (b == str(nodeId))

    def _degrouperIsVertexVertexIncident(self, att: TopologyAttachment, nodeId: str) -> bool:
        if str(att.kind) != "vertex-vertex":
            return False
        return (
            self._degrouperIsVertexFeatureOnNode(att.feature_a, nodeId)
            or self._degrouperIsVertexFeatureOnNode(att.feature_b, nodeId)
        )

    def _degrouperIsEdgeEdgeIncident(self, att: TopologyAttachment, nodeId: str) -> bool:
        if str(att.kind) != "edge-edge":
            return False
        return (
            self._degrouperIsEdgeFeatureOnNode(att.feature_a, nodeId)
            or self._degrouperIsEdgeFeatureOnNode(att.feature_b, nodeId)
        )

    def _degrouperVertexEdgeTrianglePair(self, att: TopologyAttachment) -> tuple[str, str]:
        if str(att.kind) != "vertex-edge":
            raise ValueError(f"[degrouper] kind inattendu pour vertex-edge pair: {att.kind}")

        a = att.feature_a
        b = att.feature_b
        ta = str(a.feature_type)
        tb = str(b.feature_type)

        if ta == TopologyFeatureType.VERTEX and tb == TopologyFeatureType.EDGE:
            return (str(a.element_id), str(b.element_id))
        if ta == TopologyFeatureType.EDGE and tb == TopologyFeatureType.VERTEX:
            return (str(b.element_id), str(a.element_id))

        raise ValueError(f"[degrouper] vertex-edge invalide (aid={att.attachment_id})")

    def canDegrouperAtNode(self, groupId: str, nodeId: str) -> bool:
        gid = self.find_group(str(groupId))
        nid = self.find_node(str(nodeId))
        self.rebuildGroupElementLists()
        g = self.groups.get(gid)
        if g is None:
            return False

        elementIds = {str(eid) for eid in list(getattr(g, "element_ids", []) or [])}
        if not elementIds:
            return False

        for att in self._degrouperIterGroupScopedAttachments(elementIds):
            if self._degrouperIsVertexVertexIncident(att, nid):
                return True
            if self._degrouperIsEdgeEdgeIncident(att, nid):
                return True
        return False

    def pruneOrphanGroups(self) -> None:
        # groupes réellement utilisés par au moins un elementId
        used = {self.find_group(gid) for gid in self.element_to_group.values()}

        # 1) self.groups + cache conceptuel
        for gid in list(self.groups.keys()):
            gidc = self.find_group(gid)
            if gidc not in used:
                self.groups.pop(gid, None)
                self._concept_by_gid.pop(gid, None)

        # 2) DSU group structures
        keep = set(self.groups.keys()) | used
        for gid in list(self._group_parent.keys()):
            if gid not in keep:
                self._group_parent.pop(gid, None)
                self._group_members.pop(gid, None)
                self._group_created_order.pop(gid, None)
                self._concept_by_gid.pop(gid, None)

    def degrouperAtNode(self, groupId: str, nodeId: str) -> dict:
        if int(getattr(self, "_topoTxDepth", 0)) > 0:
            raise ValueError("TopologyWorld.degrouperAtNode: transaction ouverte")

        gid = self.find_group(str(groupId))
        nid = self.find_node(str(nodeId))
        assert gid in self.groups, f"[degrouper] groupe introuvable: {groupId}"

        # A) Entrées du groupe source
        self.rebuildGroupElementLists()
        g = self.groups.get(gid)
        assert g is not None, f"[degrouper] groupe canonique introuvable: {gid}"
        sourceElementIds = sorted({str(eid) for eid in list(getattr(g, "element_ids", []) or [])})
        sourceSet = set(sourceElementIds)
        assert sourceSet, f"[degrouper] groupe vide: {gid}"

        # B) Collecte des attachments à supprimer
        groupAttachments = self._degrouperIterGroupScopedAttachments(sourceSet)
        vvIncident: list[TopologyAttachment] = []
        eeIncident: list[TopologyAttachment] = []
        for att in groupAttachments:
            if self._degrouperIsVertexVertexIncident(att, nid):
                vvIncident.append(att)
            if self._degrouperIsEdgeEdgeIncident(att, nid):
                eeIncident.append(att)

        if not vvIncident and not eeIncident:
            assert False, "[degrouper] aucun attachment incident eligible au noeud cible"

        toRemove: set[str] = set()
        for att in (vvIncident + eeIncident):
            aid = str(getattr(att, "attachment_id", "") or "")
            assert aid, "[degrouper] attachment sans id"
            toRemove.add(aid)

        vvTrianglePairsOriented: set[tuple[str, str]] = set()
        for att in vvIncident:
            ti, tj = self._degrouperAttachmentElementPair(att)
            vvTrianglePairsOriented.add((ti, tj))
            vvTrianglePairsOriented.add((tj, ti))

        if vvTrianglePairsOriented:
            for att in groupAttachments:
                if str(att.kind) != "vertex-edge":
                    continue
                tv, te = self._degrouperVertexEdgeTrianglePair(att)
                if (tv, te) in vvTrianglePairsOriented:
                    aid = str(getattr(att, "attachment_id", "") or "")
                    assert aid, "[degrouper] attachment sans id"
                    toRemove.add(aid)

        allAttachmentsSorted = [self.attachments[aid] for aid in sorted(self.attachments.keys())]
        keptAttachments = [att for att in allAttachmentsSorted if str(att.attachment_id) not in toRemove]

        # C) Rebuild conceptuel intermédiaire
        self.rebuild_from_attachments(keptAttachments)
        self.rebuildGroupElementLists()

        # D) Composantes connexes triangle<->triangle sur les éléments du groupe source
        adjacency: dict[str, set[str]] = {eid: set() for eid in sourceElementIds}
        for att in keptAttachments:
            eA, eB = self._degrouperAttachmentElementPair(att)
            if eA in sourceSet and eB in sourceSet:
                adjacency[eA].add(eB)
                adjacency[eB].add(eA)

        components: list[list[str]] = []
        visited: set[str] = set()
        for seed in sourceElementIds:
            if seed in visited:
                continue
            stack: list[str] = [seed]
            visited.add(seed)
            comp: list[str] = []
            while stack:
                cur = stack.pop()
                comp.append(cur)
                for nxt in sorted(adjacency.get(cur, set()), reverse=True):
                    if nxt in visited:
                        continue
                    visited.add(nxt)
                    stack.append(nxt)
            components.append(sorted(comp))

        components = sorted(components, key=lambda c: (-len(c), c))
        assert len(components) >= 2, "[degrouper] K<2 apres suppression d'attachments"

        # E) Réassignation des groupes : la plus grande composante garde gid
        mainGroupId = gid
        mainComponent = components[0]
        for eid in mainComponent:
            self.element_to_group[str(eid)] = mainGroupId

        newGroupIds: list[str] = []
        movedElementIdsByGroup: dict[str, list[str]] = {}
        for comp in components[1:]:
            newGid = self.createAtomicGroup()
            newGroupIds.append(newGid)
            moved = sorted({str(eid) for eid in comp})
            movedElementIdsByGroup[newGid] = moved
            for eid in moved:
                self.element_to_group[str(eid)] = newGid

        self.rebuild_from_attachments(keptAttachments)
        self.rebuildGroupElementLists()
        self.pruneOrphanGroups()
        
        allResultGroupIds = [mainGroupId] + list(newGroupIds)

        # Invariant: aucune attache ne relie 2 groupes différents.
        for att in self.attachments.values():
            gA = self.get_group_of_element(str(att.feature_a.element_id))
            gB = self.get_group_of_element(str(att.feature_b.element_id))
            assert gA == gB, (
                f"[degrouper] attachment inter-groupes interdit: aid={att.attachment_id} gA={gA} gB={gB}"
            )

        # F) Recompute concept + contour
        for rid in allResultGroupIds:
            ridc = self.find_group(str(rid))
            assert ridc == str(rid), f"[degrouper] groupId canonique inattendu: req={rid} got={ridc}"
            self.recomputeConceptAndBoundary(rid)
            boundary = self.getBoundarySegments(rid)
            assert len(boundary) >= 3, f"[degrouper] contour invalide pour {rid}"

        # G) Chemins: casser si le groupe touché portait le chemin
        tc = getattr(self, "topologyChemins", None)
        if tc is not None and bool(getattr(tc, "isDefined", False)):
            tcGroupId = str(getattr(tc, "groupId", "") or "")
            if tcGroupId and self.find_group(tcGroupId) == mainGroupId:
                tc.supprimerChemin()

        # H) Retour normatif
        return {
            "mainGroupId": str(mainGroupId),
            "newGroupIds": [str(x) for x in newGroupIds],
            "movedElementIdsByGroup": {str(k): [str(eid) for eid in v] for k, v in movedElementIdsByGroup.items()},
            "allResultGroupIds": [str(x) for x in allResultGroupIds],
        }

    # ------------------------------------------------------------------
    # Clone physique deterministe (sans derives)
    # ------------------------------------------------------------------
    def _exportPhysicalSnapshot(self) -> dict:
        if self._topoTxDepth > 0:
            raise ValueError("TopologyWorld.clonePhysicalState: transaction ouverte")

        elements_payload: list[dict] = []
        for eid in sorted(self.elements.keys()):
            el = self.elements[eid]
            R, T, mirrored = el.get_pose()
            elements_payload.append({
                "element_id": str(el.element_id),
                "name": str(el.name),
                "vertex_labels": list(el.vertex_labels),
                "vertex_types": list(el.vertex_types),
                "edge_lengths_km": [float(x) for x in el.edge_lengths_km],
                "intrinsic_sides_km": dict(getattr(el, "intrinsic_sides_km", {}) or {}),
                "local_frame": dict(getattr(el, "local_frame", {}) or {}),
                "vertex_local_xy": {int(k): (float(v[0]), float(v[1]))
                                    for k, v in dict(getattr(el, "vertex_local_xy", {}) or {}).items()},
                "meta": dict(getattr(el, "meta", {}) or {}),
                "pose": {
                    "R": np.array(R, float).tolist(),
                    "T": np.array(T, float).tolist(),
                    "mirrored": bool(mirrored),
                },
            })

        attachments_payload: list[dict] = []
        for att in self.attachments.values():
            attachments_payload.append({
                "attachment_id": str(att.attachment_id),
                "kind": str(att.kind),
                "feature_a": {
                    "feature_type": str(att.feature_a.feature_type),
                    "element_id": str(att.feature_a.element_id),
                    "index": int(att.feature_a.index),
                    "t": None if att.feature_a.t is None else float(att.feature_a.t),
                },
                "feature_b": {
                    "feature_type": str(att.feature_b.feature_type),
                    "element_id": str(att.feature_b.element_id),
                    "index": int(att.feature_b.index),
                    "t": None if att.feature_b.t is None else float(att.feature_b.t),
                },
                "params": dict(att.params) if att.params is not None else {},
                "source": str(att.source),
            })

        config_payload = {
            "fusion_distance_km": float(getattr(self, "fusion_distance_km", 1.0)),
            "worldYAxisDown": bool(getattr(self, "worldYAxisDown", False)),
        }
        if hasattr(self, "vertex_edge_endpoint_epsilon"):
            config_payload["vertex_edge_endpoint_epsilon"] = float(getattr(self, "vertex_edge_endpoint_epsilon"))

        return {
            "config": config_payload,
            "elements": elements_payload,
            "attachments": attachments_payload,
        }

    def _importPhysicalSnapshot(self, snapshot: dict) -> None:
        if snapshot is None:
            raise ValueError("TopologyWorld._importPhysicalSnapshot: snapshot absent")
        self._isImportingSnapshot = True
        try:
            config = snapshot.get("config", {}) or {}
            if "fusion_distance_km" in config:
                self.fusion_distance_km = float(config.get("fusion_distance_km", 1.0))
            if "worldYAxisDown" in config:
                self.worldYAxisDown = bool(config.get("worldYAxisDown", False))
            if "vertex_edge_endpoint_epsilon" in config:
                self.vertex_edge_endpoint_epsilon = float(config.get("vertex_edge_endpoint_epsilon", 1e-6))

            elements_payload = list(snapshot.get("elements", []) or [])
            for el in elements_payload:
                eid = str(el.get("element_id"))
                raw_vxy = el.get("vertex_local_xy", {}) or {}
                vertex_local_xy = {}
                for k, v in dict(raw_vxy).items():
                    kk = int(k)
                    vertex_local_xy[kk] = (float(v[0]), float(v[1]))
                clone = TopologyElement(
                    element_id=eid,
                    name=str(el.get("name", "")),
                    vertex_labels=list(el.get("vertex_labels", [])),
                    vertex_types=list(el.get("vertex_types", [])),
                    edge_lengths_km=list(el.get("edge_lengths_km", [])),
                    intrinsic_sides_km=dict(el.get("intrinsic_sides_km", {}) or {}),
                    local_frame=dict(el.get("local_frame", {}) or {}),
                    vertex_local_xy=vertex_local_xy,
                    meta=dict(el.get("meta", {}) or {}),
                )
                self.add_element_as_new_group(clone)
                pose = el.get("pose", {}) or {}
                if pose:
                    self.setElementPose(
                        eid,
                        R=np.array(pose.get("R", np.eye(2)), float),
                        T=np.array(pose.get("T", np.zeros(2)), float),
                        mirrored=bool(pose.get("mirrored", False)),
                    )

            attachments_payload = list(snapshot.get("attachments", []) or [])
            max_att_num = 0
            for att in attachments_payload:
                fa = att.get("feature_a", {}) or {}
                fb = att.get("feature_b", {}) or {}
                feature_a = TopologyFeatureRef(
                    feature_type=fa.get("feature_type"),
                    element_id=fa.get("element_id"),
                    index=int(fa.get("index", 0)),
                    t=fa.get("t", None),
                )
                feature_b = TopologyFeatureRef(
                    feature_type=fb.get("feature_type"),
                    element_id=fb.get("element_id"),
                    index=int(fb.get("index", 0)),
                    t=fb.get("t", None),
                )
                attachment = TopologyAttachment(
                    attachment_id=str(att.get("attachment_id")),
                    kind=str(att.get("kind")),
                    feature_a=feature_a,
                    feature_b=feature_b,
                    params=dict(att.get("params", {}) or {}),
                    source=str(att.get("source", "manual")),
                )
                self.loadAttachment(attachment)
                m = re.search(r"(?:^A|:A)(\d+)$", str(attachment.attachment_id))
                if m:
                    max_att_num = max(max_att_num, int(m.group(1)))

            if max_att_num > 0:
                self._created_counter_attachments = max_att_num

            self.rebuildGroupElementLists()
            for gid in sorted(self.groups.keys()):
                if gid == self.find_group(gid):
                    self.recomputeConceptAndBoundary(gid)
        finally:
            self._isImportingSnapshot = False

    def clonePhysicalState(self) -> "TopologyWorld":
        """Clone deterministe de l'etat physique (sans derives)."""
        if self._topoTxDepth > 0:
            raise ValueError("TopologyWorld.clonePhysicalState: transaction ouverte")

        snapshot = self._exportPhysicalSnapshot()
        target = TopologyWorld()
        target._importPhysicalSnapshot(snapshot)
        return target

    # ------------------------------------------------------------------
    # Helpers "Tooltip" (Core)
    # ------------------------------------------------------------------
    def _parseElementAndVertexIndexFromNodeId(self, node_id: str) -> tuple[str, int] | None:
        """
        Parse un node_id atomique 'Txx:Nn' -> (element_id, vertex_index).
        Exemples valides :
            'T01:N0' -> ('T01', 0)
            'T12:N2' -> ('T12', 2)
        """
        if node_id is None:
            return None

        m = re.fullmatch(r"(T\d+):N(\d+)", str(node_id))
        if not m:
            return None

        element_id = m.group(1)
        vertex_index = int(m.group(2))
        return (element_id, vertex_index)

    def getAtomicNodeLabel(self, node_id: str) -> str:
        """Retourne le label métier (ville) d'un node atomique.
        Fallback: 'N0', 'N1', ...
        """
        parsed = self._parseElementAndVertexIndexFromNodeId(str(node_id))
        if parsed is None:
            return str(node_id)

        element_id, vidx = parsed
        el = self.elements.get(element_id)
        if el is None:
            return f"N{vidx}"

        # v4.3 : labels portés par TopologyVertex / vertex_labels
        if hasattr(el, "vertexes") and el.vertexes and 0 <= vidx < len(el.vertexes):
            return str(el.vertexes[vidx].label)
        if hasattr(el, "vertex_labels") and 0 <= vidx < len(el.vertex_labels):
            return str(el.vertex_labels[vidx])

        return f"N{vidx}"

    def getNodeType(self, nodeId: str) -> str:
        elementId, vidx = self._parseElementAndVertexIndexFromNodeId(nodeId)
        el = self.elements[elementId]
        return el.vertex_types[vidx]

    def getNodeLabel(self, node_id: str) -> str:
        """Label métier d'un node (canonique ou atomique).
        - si node_id est canonique (DSU), agrège les labels de ses membres.
        """
        nid = str(node_id)
        if nid not in self._node_parent:
            return self.getAtomicNodeLabel(nid)
        c = self.find_node(nid)
        uniq = sorted({self.getAtomicNodeLabel(m) for m in self.node_members(c)})
        if len(uniq) == 0:
            return self.getAtomicNodeLabel(nid)
        if len(uniq) == 1:
            return uniq[0]
        return "|".join([str(x) for x in uniq])

    def getConceptNodeLabel(self, node_id: str) -> str:
        """Label métier d'un concept node (DSU canonique)."""
        return self.getNodeLabel(self.find_node(str(node_id)))

    def getPhysicalNodeName(self, node_id: str) -> str:
        node_type = self.getNodeTypeAtomic(node_id)
        label = self.getAtomicNodeLabel(str(node_id))
        tri = self._parseTriangleIdFromNodeId(str(node_id))
        return f"{node_type}:{label}({tri})"

    def getConceptNodeName(self, node_id: str) -> str:
        """Retourne le libellé conceptuel pour un node.
        Renvoie un texte "Type:Labels" basé sur le node canonique.
        Garantit une agrégation des labels des membres.
        Ne choisit pas une vue physique et ne retourne pas d'ID.
        Ne modifie pas la topologie ni les caches.
        """
        cn = self.find_node(str(node_id))
        member_ids = list(self.node_members(cn))
        node_type = self.computeConceptNodeTypeOptionB(member_ids)
        label = self.getNodeLabel(cn)
        return f"{node_type}:{label}"

    def getPhysicalNodesForConceptNode(self, node_id: str) -> list[str]:
        cn = self.find_node(str(node_id))
        return list(self.node_members(cn))

    def getNodeTypeAtomic(self, node_id: str) -> str:
        return str(self._node_type[str(node_id)])

    def resolvePhysicalNodeIdDeterministic(self, nodeId: str) -> str | None:
        """Retourne un nodeId physique 'Txx:Nn' depuis un nodeId (DSU canon ou déjà physique).
        Stratégie: si DSU -> getPhysicalNodesForConceptNode(canon) triés, prendre le 1er."""
        if nodeId is None:
            return None

        s = str(nodeId)

        # déjà physique ?
        if self._parseElementAndVertexIndexFromNodeId(s) is not None:
            return s

        canon = str(self.find_node(s))
        phys = self.getPhysicalNodesForConceptNode(canon) or []   # retourne node_members(canon)
        if not phys:
            return None
        return sorted(str(p) for p in phys)[0]

    def getElementVertexFromAnyNodeId(self, nodeId: str, groupId: str | None = None) -> dict | None:
        """Retourne une ref exploitable par Tk:
        - elementId, vertexIndex, vkey
        - wbest (position monde du concept node)
        """
        physId = self.resolvePhysicalNodeIdDeterministic(nodeId)
        if physId is None:
            return None
        parsed = self._parseElementAndVertexIndexFromNodeId(physId)
        if parsed is None:
            return None

        elementId, vertexIndex = parsed

        # vkey = type atomique ("O","B","L") du noeud physique
        vkey = str(self.getNodeTypeAtomic(physId))
        gid = str(groupId) if groupId is not None else str(self._infer_group_for_node(physId))

        canon = str(self.find_node(physId))
        wbest = self.getConceptNodeWorldXY(canon, gid)

        return {
            "nodeCanon": canon,
            "nodePhys": str(physId),
            "elementId": str(elementId),
            "vertexIndex": int(vertexIndex),
            "vkey": str(vkey),
            "wbest": (float(wbest[0]), float(wbest[1])),
        }

    def computeConceptNodeTypeOptionB(self, member_ids: list[str]) -> str:
        types = {self.getNodeTypeAtomic(mid) for mid in member_ids}
        if not types:
            raise ValueError("[ConceptNode] empty members list")
        if len(types) == 1:
            return next(iter(types))
        return "M"

    def _parseTriangleIdFromNodeId(self, node_id: str) -> str:
        """
        Retourne 'Txx' à partir d'un node_id atomique 'Txx:Nn'.

        Exemples :
            'T01:N0' -> 'T01'
            'T12:N2' -> 'T12'
        """
        if node_id is None:
            return "T??"

        m = re.fullmatch(r"(T\d+):N\d+", str(node_id))
        if not m:
            return "T??"
        return m.group(1)

    def _localPointOnEdge(self, el: "TopologyElement", edge: "TopologyEdge", t: float) -> np.ndarray:
        """Point local sur une arête (interpolation linéaire) via vertex_local_xy."""
        v0 = int(edge.v_start.vertex_index)
        v1 = int(edge.v_end.vertex_index)
        p0 = el.vertex_local_xy.get(v0) if hasattr(el, "vertex_local_xy") else None
        p1 = el.vertex_local_xy.get(v1) if hasattr(el, "vertex_local_xy") else None
        if p0 is None or p1 is None:
            raise ValueError(
                f"_localPointOnEdge: vertex_local_xy manquant (element={el.element_id} v0={v0} v1={v1})"
            )
        t = float(t)
        x = (1.0 - t) * float(p0[0]) + t * float(p1[0])
        y = (1.0 - t) * float(p0[1]) + t * float(p1[1])
        return np.array([x, y], dtype=float)

    # --- export ---

    def validate_world(self) -> list[str]:
        errors: list[str] = []
        try:
            self.assertNoPhysicalSplitPoints()
        except RuntimeError as exc:
            errors.append(str(exc))
        for aid in sorted(self.attachments.keys()):
            try:
                self._validate_attachment_p2(self.attachments[aid])
            except ValueError as exc:
                errors.append(str(exc))
        eps_t = float(getattr(self, "concept_split_epsilon_t", 1e-9))
        for el in self.elements.values():
            for edge in el.edges:
                for c in edge.coverages:
                    if c.t0 >= c.t1:
                        errors.append(f"Edge {edge.edge_id()}: coverage dégénéré [{c.t0},{c.t1}]")

                gid = self.element_to_group.get(str(el.element_id))
                if gid is None:
                    continue
                try:
                    ccache = self.ensureConceptGraph(gid)
                except ValueError as exc:
                    errors.append(str(exc))
                    continue
                edge_id = str(edge.edge_id())
                occs = [
                    occ for ce in ccache.edges.values()
                    for occ in ce.occurrences
                    if str(occ.get("elementId")) == str(el.element_id)
                    and str(occ.get("edgeId")) == edge_id
                ]
                if not occs:
                    errors.append(f"[C2][{edge_id}] no concept occurrences (no coverage)")
                    continue

                norm_occs: list[tuple[float, float]] = []
                for k, occ in enumerate(occs):
                    try:
                        t0 = float(occ.get("t0", None))
                        t1 = float(occ.get("t1", None))
                    except Exception:
                        errors.append(
                            f"[C2][{edge_id}] occ#{k} t0/t1 not float t0={occ.get('t0')} t1={occ.get('t1')}"
                        )
                        continue
                    if t0 < -eps_t or t0 > 1.0 + eps_t or t1 < -eps_t or t1 > 1.0 + eps_t:
                        errors.append(
                            f"[C2][{edge_id}] occ#{k} out of [0,1] t0={t0} t1={t1}"
                        )
                        continue
                    if abs(t0) <= eps_t:
                        t0 = 0.0
                    if abs(t1 - 1.0) <= eps_t:
                        t1 = 1.0
                    if t1 <= t0 + eps_t:
                        errors.append(
                            f"[C2][{edge_id}] occ#{k} null/neg segment t0={t0} t1={t1}"
                        )
                        continue
                    norm_occs.append((t0, t1))

                norm_occs.sort(key=lambda v: (v[0], v[1]))
                if norm_occs:
                    if abs(norm_occs[0][0] - 0.0) > eps_t:
                        errors.append(
                            f"[C2][{edge_id}] coverage does not start at 0 (t0={norm_occs[0][0]})"
                        )
                    if abs(norm_occs[-1][1] - 1.0) > eps_t:
                        errors.append(
                            f"[C2][{edge_id}] coverage does not end at 1 (t1={norm_occs[-1][1]})"
                        )
                    for i in range(1, len(norm_occs)):
                        prev_t1 = norm_occs[i - 1][1]
                        cur_t0 = norm_occs[i][0]
                        gap = cur_t0 - prev_t1
                        if gap > eps_t:
                            errors.append(
                                f"[C2][{edge_id}] GAP between occ#{i-1} and occ#{i} prev.t1={prev_t1} cur.t0={cur_t0}"
                            )
                        elif gap < -eps_t:
                            errors.append(
                                f"[C2][{edge_id}] OVERLAP between occ#{i-1} and occ#{i} prev.t1={prev_t1} cur.t0={cur_t0}"
                            )

                    try:
                        dp_points = self.buildDerivedSplitPointsForPhysEdge(el.element_id, edge.edge_index, eps_t=eps_t)
                    except ValueError as exc:
                        errors.append(str(exc))
                        continue
                    dp = [float(sp["t"]) for sp in dp_points]
                    if len(dp) < 2:
                        errors.append(f"[C2+][{edge_id}] occ count mismatch occs={len(norm_occs)} dp={len(dp)}")
                    else:
                        if len(norm_occs) != len(dp) - 1:
                            errors.append(
                                f"[C2+][{edge_id}] occ count mismatch occs={len(norm_occs)} dp={len(dp)}"
                            )
                        else:
                            for i in range(len(norm_occs)):
                                t0, t1 = norm_occs[i]
                                if abs(t0 - dp[i]) > eps_t or abs(t1 - dp[i + 1]) > eps_t:
                                    errors.append(
                                        f"[C2+][{edge_id}] split mismatch i={i} occ=({t0},{t1}) dp=({dp[i]},{dp[i+1]})"
                                    )
        return errors

    @staticmethod
    def _format_chemin_angle(value: float) -> str:
        """Stable decimal format for azimuths and angles."""
        v = float(value)
        if not math.isfinite(v):
            raise RuntimeError(f"[CheminsDump] valeur non finie: {value}")
        return f"{v:.2f}"

    @staticmethod
    def _format_chemin_distance(value: float) -> str:
        """Stable decimal format for distances in km."""
        v = float(value)
        if not math.isfinite(v):
            raise RuntimeError(f"[CheminsDump] valeur non finie: {value}")
        return f"{v:.3f}"

    def _dump_chemins_to_xml(self, parent: _ET.Element) -> None:
        """Serialise topologyChemins in a deterministic and diff-friendly XML section."""
        tc = self.topologyChemins
        chemins_el = _ET.SubElement(parent, "chemins", {
            "isDefined": "1" if bool(tc.isDefined) else "0",
        })

        if not tc.isDefined:
            return

        if tc.groupId is None or tc.startNodeId is None:
            raise RuntimeError("[CheminsDump] chemin defini incomplet: groupId/startNodeId")
        orient = TopologyChemins._normalizeOrientation(tc.orientationUser)
        if len(tc.selectionMask) != len(tc.borderSnapshotNodes):
            raise RuntimeError("[CheminsDump] borderSnapshot/selectionMask tailles differentes")

        chemins_el.set("groupId", str(tc.groupId))
        chemins_el.set("startNodeId", str(tc.startNodeId))
        chemins_el.set("orientationUser", orient)

        border_el = _ET.SubElement(chemins_el, "borderSnapshot")
        for nid in tc.borderSnapshotNodes:
            _ET.SubElement(border_el, "node", {"id": str(nid)})

        mask_el = _ET.SubElement(chemins_el, "selectionMask")
        for selected in tc.selectionMask:
            _ET.SubElement(mask_el, "m", {"v": "1" if bool(selected) else "0"})

        path_el = _ET.SubElement(chemins_el, "pathNodesOrdered")
        for nid in tc.pathNodesOrdered:
            _ET.SubElement(path_el, "node", {"id": str(nid)})

        triplets_el = _ET.SubElement(chemins_el, "triplets")
        for t in tc.triplets:
            if not bool(t.isGeometrieValide):
                raise RuntimeError(
                    f"[CheminsDump] triplet sans geometrie valide: ({t.nodeA},{t.nodeO},{t.nodeB})"
                )
            _ET.SubElement(triplets_el, "triplet", {
                "A": str(t.nodeA),
                "O": str(t.nodeO),
                "B": str(t.nodeB),
                "azOA": self._format_chemin_angle(t.azOA),
                "azOB": self._format_chemin_angle(t.azOB),
                "distOA_km": self._format_chemin_distance(t.distOA_km),
                "distOB_km": self._format_chemin_distance(t.distOB_km),
                "angleDeg": self._format_chemin_angle(t.angleDeg),
            })

    def export_topo_dump_xml(self, path: str, orientation: str = "cw") -> str:
        self.assertNoPhysicalSplitPoints()
        root = _ET.Element("TopoDump", {
            "version": "4.3",
            "units": "km",
            "tFrame": "world",
        })

        groups_el = _ET.SubElement(root, "Groups")
        for gid in sorted(self.groups.keys()):
            if gid != self.find_group(gid):
                continue
            g = self.groups[gid]
            grp = _ET.SubElement(groups_el, "Group", {
                "id": gid,
                "elements": str(len(g.element_ids)),
                "attachments": str(len(g.attachment_ids)),
                "originGroups": ",".join(sorted(self.group_members(gid))),
            })
            errs = self.validate_world()
            _ET.SubElement(grp, "Validate", {"errors": str(len(errs)), "warnings": "0"})

        nodes_el = _ET.SubElement(root, "Nodes")
        canonical_nodes = sorted({self.find_node(nid) for nid in self._node_parent.keys()})
        for cn in canonical_nodes:
            _ET.SubElement(nodes_el, "Node", {
                "id": cn,
                "type": self.node_type(cn),
                "origins": ",".join(sorted(self.node_members(cn))),
                "createdOrder": str(self._node_created_order.get(self.find_node(cn), 0)),
            })

        # --- ConceptModels (debug) : un par GROUPE ---
        # DSU nodes = global, mais on exporte un sous-graphe conceptuel par groupe canonique,
        # indispensable pour calculer/valider un contour de groupe.
        cm_all = _ET.SubElement(root, "ConceptModels")
        for gid in sorted(self.groups.keys()):
            if gid != self.find_group(gid):
                continue
            ccache = self.ensureConceptGeom(gid)

            cm = _ET.SubElement(cm_all, "ConceptModel", {"version": "1.0", "group": str(gid)})
            cns = _ET.SubElement(cm, "ConceptNodes")
            for cid in sorted(ccache.nodes.keys()):
                xy = ccache.nodes[cid].world_xy()
                member_ids = sorted(self.node_members(cid))
                concept_type = self.computeConceptNodeTypeOptionB(member_ids)
                cn_el = _ET.SubElement(cns, "ConceptNode", {
                    "id": str(cid),
                    "type": str(concept_type),
                    "x": f"{float(xy[0]):.6f}",
                    "y": f"{float(xy[1]):.6f}",
                })
                mem_el = _ET.SubElement(cn_el, "Members")
                for mid in member_ids:
                    _ET.SubElement(mem_el, "M", {
                        "id": str(mid),
                        "type": str(self.getNodeTypeAtomic(mid)),
                        "tri": str(self._parseTriangleIdFromNodeId(str(mid))),
                        "label": str(self.getAtomicNodeLabel(str(mid))),
                    })

            ces = _ET.SubElement(cm, "ConceptEdges")
            for (a, b) in sorted(ccache.edges.keys()):
                p0 = ccache.nodes[a].world_xy() if a in ccache.nodes else (0.0, 0.0)
                p1 = ccache.nodes[b].world_xy() if b in ccache.nodes else (0.0, 0.0)
                dx = float(p1[0] - p0[0])
                dy = float(p1[1] - p0[1])
                dist = float(math.hypot(dx, dy))
                az = self.azimutDegFromDxDy(dx, dy)
                ce_el = _ET.SubElement(ces, "ConceptEdge", {
                    "a": str(a),
                    "b": str(b),
                    "distanceKm": f"{dist:.6f}",
                    "azimutDeg": f"{az:.6f}",
                })
                occ_el = _ET.SubElement(ce_el, "Occurrences")
                for occ in ccache.edges[(a, b)].occurrences:
                    _ET.SubElement(occ_el, "O", {
                        "element": str(occ.get("elementId", "")),
                        "edge": str(occ.get("edgeId", "")),
                        "from": str(occ.get("fromNodeId", "")),
                        "to": str(occ.get("toNodeId", "")),
                        "t0": f"{float(occ.get('t0', 0.0)):.6f}",
                        "t1": f"{float(occ.get('t1', 0.0)):.6f}",
                    })
            if ccache.boundaryCycle:
                bnd = _ET.SubElement(cm, "Boundary", {"orientation": str(ccache.boundaryOrientation)})
                cycle_el = _ET.SubElement(bnd, "Cycle")
                for nid in ccache.boundaryCycle:
                    _ET.SubElement(cycle_el, "N", {"id": str(nid)})

        self._dump_chemins_to_xml(root)

        elements_el = _ET.SubElement(root, "Elements")
        for eid in sorted(self.elements.keys()):
            el = self.elements[eid]
            # Lien explicite Element -> Groupe (dump-only, info-only)
            gid = self.get_group_of_element(el.element_id)
            elx = _ET.SubElement(elements_el, "Element", {
                "id": el.element_id,
                "name": el.name,
                "n": str(el.n_vertices()),
                "group": str(gid),
            })

            # --- Pose monde (par élément) ---
            # Fallback implicite: identité si non renseigné.
            R, T, mirrored = el.get_pose()
            ep = _ET.SubElement(elx, "ElementPose", {
                "mirrored": "1" if mirrored else "0",
                "tx": f"{float(T[0]):.6f}",
                "ty": f"{float(T[1]):.6f}",
            })
            _ET.SubElement(ep, "R", {
                "r00": f"{float(R[0, 0]):.9f}", "r01": f"{float(R[0, 1]):.9f}",
                "r10": f"{float(R[1, 0]):.9f}", "r11": f"{float(R[1, 1]):.9f}",
            })

            # --- Coordonnées locales (relatives) ---
            orient = str(el.meta.get("orient", "")).strip().upper()
            lc = _ET.SubElement(elx, "LocalCoords", {
                "frameOrigin": str(el.local_frame.get("origin", TopologyNodeType.OUVERTURE)),
                "xAxis": str(el.local_frame.get("xAxis", "O->B")),
                "units": str(el.local_frame.get("units", "km")),
                "orient": orient if orient else "CCW",
            })
            for vidx in sorted(el.vertex_local_xy.keys()):
                x, y = el.vertex_local_xy[vidx]
                _ET.SubElement(lc, "P", {"idx": str(vidx), "x": f"{float(x):.6f}", "y": f"{float(y):.6f}"})

            vx = _ET.SubElement(elx, "Vertices")
            for v in el.vertexes:
                _ET.SubElement(vx, "V", {
                    "idx": str(v.vertex_index),
                    "label": v.label,
                    "node": self.find_node(v.node_id),
                    "nodeOrigin": v.node_id,
                    "type": v.node_type,
                })

            ex = _ET.SubElement(elx, "Edges")
            for e in el.edges:
                a, b = e.edge_label_key()
                edge_el = _ET.SubElement(ex, "Edge", {
                    "id": e.edge_id(),
                    "idx": str(e.edge_index),
                    "from": str(e.v_start.vertex_index),
                    "to": str(e.v_end.vertex_index),
                    "labels": e.labels_display(),
                    "edgeLabelKey": f"{a}|{b}",
                    "lengthKm": f"{e.edge_length_km:.3f}",
                })
                pts = _ET.SubElement(edge_el, "Points")
                phys_points = [
                    (0.0, e.v_start.node_id),
                    (1.0, e.v_end.node_id),
                ]
                t_list = [float(t) for t, _ in phys_points]
                if len(phys_points) != 2 or any(t not in (0.0, 1.0) for t in t_list):
                    raise ValueError(
                        f"Edge {e.edge_id()}: invalid physical points (n={len(phys_points)} t={t_list})"
                    )
                for t_val, node_id in phys_points:
                    _ET.SubElement(pts, "P", {
                        "t": f"{float(t_val):.6f}",
                        "node": self.find_node(node_id),
                        "nodeOrigin": node_id,
                    })
                covs = _ET.SubElement(edge_el, "Coverages")
                for c in sorted(e.coverages, key=lambda c: (c.t0, c.t1)):
                    _ET.SubElement(covs, "C", {"t0": f"{c.t0:.6f}", "t1": f"{c.t1:.6f}"})

        atts_el = _ET.SubElement(root, "Attachments")
        for aid in sorted(self.attachments.keys()):
            a = self.attachments[aid]
            att = _ET.SubElement(atts_el, "Attachment", {
                "id": a.attachment_id,
                "kind": a.kind,
                "source": a.source,
                "featureA": a.feature_a.to_string(),
                "featureB": a.feature_b.to_string(),
            })
            params_el = _ET.SubElement(att, "Params")
            for k in sorted(a.params.keys()):
                _ET.SubElement(params_el, "Param", {"name": str(k), "value": str(a.params[k])})

        _ET.ElementTree(root).write(path, encoding="utf-8", xml_declaration=True)
        return path

    @staticmethod
    def _distPointToSegment2D(p: np.ndarray, a: np.ndarray, b: np.ndarray) -> tuple[float, np.ndarray, float]:
        """Distance point-segment (2D).
        Retourne (dist, proj, t) avec proj = a + t*(b-a), t clampé dans [0,1]."""
        p = np.array(p, dtype=float)
        a = np.array(a, dtype=float)
        b = np.array(b, dtype=float)
        ab = b - a
        ab2 = float(ab[0] * ab[0] + ab[1] * ab[1])
        if ab2 <= 1e-24:
            # segment dégénéré
            d = float(np.linalg.norm(p - a))
            return (d, a, 0.0)
        t = float(((p - a)[0] * ab[0] + (p - a)[1] * ab[1]) / ab2)
        if t < 0.0:
            t = 0.0
        elif t > 1.0:
            t = 1.0
        proj = a + ab * t
        d = float(np.linalg.norm(p - proj))
        return (d, proj, t)

    def findNearestBoundaryNode(self, groupId: str | None, pWorld: tuple[float, float] | np.ndarray) -> dict | None:
        """Trouve le node DSU (canonique) le plus proche d'un point monde,
        en ne considérant QUE la boundary.

        - Si groupId est fourni -> comportement actuel (1 groupe)
        - Si groupId est None -> scanne tous les groupes disponibles et renvoie le meilleur hit

        Retour:
            None si pas de boundary exploitable.
            Sinon dict:
            {
                "groupId": <gid canonique>,
                "nodeId": <node canonique>,
                "distKm": <distance>,
                "ptOnBoundary": (x,y),
                "segment": (a,b),
                "t": <0..1>
            }
        """
        p = np.array(pWorld, dtype=float)

        def _hit_for_gid(gid_in: str) -> dict | None:
            gid = self.find_group(str(gid_in))

            # boundary + géométrie conceptuelle requises
            self.computeBoundary(gid)  # no-op si déjà calculé/caché
            c = self.ensureConceptGeom(gid)

            if not c.boundaryCycle or len(c.boundaryCycle) < 2:
                return None

            segs = self.getBoundarySegments(gid)
            if not segs:
                return None

            best = None  # (dist, proj, t, a, b)
            for s in segs:
                a = str(s.conceptA)
                b = str(s.conceptB)

                pa = np.array(self.getConceptNodeWorldXY(a, gid), dtype=float)
                pb = np.array(self.getConceptNodeWorldXY(b, gid), dtype=float)

                d, proj, t = self._distPointToSegment2D(p, pa, pb)
                if (best is None) or (d < best[0]):
                    best = (d, proj, t, a, b)

            if best is None:
                return None

            d, proj, t, a, b = best

            # Retourner un *node DSU* => choisir l'endpoint le plus proche du point
            pa = np.array(self.getConceptNodeWorldXY(a, gid), dtype=float)
            pb = np.array(self.getConceptNodeWorldXY(b, gid), dtype=float)
            da = float(np.linalg.norm(p - pa))
            db = float(np.linalg.norm(p - pb))
            node = a if da <= db else b

            return {
                "groupId": str(gid),
                "nodeId": str(node),
                "distKm": float(min(da, db)),
                "ptOnBoundary": (float(proj[0]), float(proj[1])),
                "segment": (str(a), str(b)),
                "t": float(t),
            }

        # ---- mode 1 groupe (compat) ----
        if groupId is not None:
            return _hit_for_gid(str(groupId))

        # ---- mode scan global ----
        bestHit = None  # (distKm, hit_dict)
        seen = set()

        for gid0 in list(self.groups.keys()):
            gid = str(self.find_group(str(gid0)))
            if gid in seen:
                continue
            seen.add(gid)

            hit = _hit_for_gid(gid)
            if hit is None:
                continue

            d = float(hit["distKm"])
            if (bestHit is None) or (d < bestHit[0]):
                bestHit = (d, hit)

        return None if bestHit is None else bestHit[1]

    def setBaliseWorldXY(self, name: str, x: float, y: float) -> None:
        self.balisesWorld[str(name)] = (float(x), float(y))

    def clearBalises(self) -> None:
        self.balisesWorld.clear()

    def hasBalise(self, name: str) -> bool:
        return str(name) in self.balisesWorld

    def getBaliseWorldXY(self, name: str) -> tuple[float, float]:
        key = str(name)
        if key not in self.balisesWorld:
            raise KeyError(f"[Balises] balise absente: {name}")
        return self.balisesWorld[key]
