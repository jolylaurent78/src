"""Fenêtre UX autonome du prototype de gestion du catalogue."""

from __future__ import annotations

from dataclasses import dataclass
import csv
import math
import unicodedata
from pathlib import Path
from typing import Callable
import tkinter as tk
from tkinter import font as tkfont
from tkinter import filedialog, messagebox, simpledialog, ttk
from tksheet import Sheet
from src.DictionnaireEnigmes import parse_book_file
from src.assembleur_catalogue import Catalogue, CatalogueBook, CatalogueCity, CatalogueTriangle as ModelCatalogueTriangle, HypothesisTemplate
from src.assembleur_catalogue_book_asset_controller import CatalogueBookAssetController
from src.assembleur_catalogue_map_assets import CatalogueMapAssetResolver, load_calibrated_catalogue_map
from src.assembleur_catalogue_map_calibration import CatalogueMapCalibrationController, STATUS_VALID
from src.assembleur_catalogue_identity import ApplicationContext, UserCatalogueIdProvider, is_system_catalogue_id
from src.assembleur_catalogue_io import save_catalogue
from src.assembleur_paths import ApplicationPaths
from src.assembleur_tooltip import attach_tooltip

from src.assembleur_dms_editor import DmsCoordinateEditor
from src.assembleur_geo_map_view import (
    GeoMapMarker,
    GeoMapPixelMarker,
    GeoMapPixelPolyline,
    GeoMapPolyline,
    GeoMapView,
)


_TEMPLATE_BASE_COLUMN_WEIGHT = 2
_TEMPLATE_RANK_COLUMN_WEIGHT = 3
_TEMPLATE_DRAG_THRESHOLD = 6
_TEMPLATE_BASE_COLUMN_WIDTH = 120
_TRIANGLE_NOTE_VALUES = ("DO", "SI", "LA", "SOL", "FA", "MI", "RE", "ZONE")
_TRIANGLE_MAP_FIT_MARGIN = 0.20
_CALIBRATION_MAP_MAXIMUM_ZOOM = 8.0


@dataclass(frozen=True)
class TemplateDropPlan:
    action: str
    valid: bool
    message: str | None
    source_index: int | None
    target_index: int
    target_triangle_id: str | None
    preview_ranks: list[str | None] | None


@dataclass(frozen=True)
class TriangleEditorResult:
    note: str
    opening_city_id: str
    base_city_id: str
    light_city_id: str


@dataclass(frozen=True)
class TriangleCsvImportResult:
    """Synthèse métier d'un import incrémental de triangles."""

    imported_triangle_ids: tuple[str, ...]
    already_present_count: int
    errors: tuple[str, ...]

    @property
    def imported_count(self) -> int:
        return len(self.imported_triangle_ids)


@dataclass(frozen=True)
class CityCsvImportResult:
    imported_city_ids: tuple[str, ...]
    updated_city_ids: tuple[str, ...]
    unchanged_count: int
    errors: tuple[str, ...]

    @property
    def imported_count(self) -> int:
        return len(self.imported_city_ids)

    @property
    def updated_count(self) -> int:
        return len(self.updated_city_ids)

def _normalize_search_text(value: str) -> str:
    return "".join(
        char
        for char in unicodedata.normalize("NFD", value)
        if unicodedata.category(char) != "Mn"
    ).casefold()

class CitySelectionDialog(tk.Toplevel):
    """Sélecteur générique d'un objet ville par recherche filtrante."""

    def __init__(self, parent, cities: list[CatalogueCity], selected_city_id: str | None = None):
        super().__init__(parent)
        self.title("Sélection d'une ville")
        self.transient(parent)
        self.resizable(True, True)
        self.minsize(280, 320)
        self.result: str | None = None
        self._cities = sorted(cities, key=lambda city: (city.name.casefold(), city.city_id))
        self._selected_city_id = selected_city_id
        self._search_var = tk.StringVar()

        root = ttk.Frame(self, padding=10)
        root.pack(fill=tk.BOTH, expand=True)
        root.rowconfigure(2, weight=1)
        root.columnconfigure(0, weight=1)
        ttk.Label(root, text="Recherche").grid(row=0, column=0, sticky="w")
        search = ttk.Entry(root, textvariable=self._search_var)
        search.grid(row=1, column=0, sticky="ew", pady=(2, 8))
        list_frame = ttk.Frame(root)
        list_frame.grid(row=2, column=0, sticky="nsew")
        self._listbox = tk.Listbox(list_frame, exportselection=False)
        self._listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self._listbox.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self._listbox.configure(yscrollcommand=scrollbar.set)
        self._listbox.bind("<Double-Button-1>", lambda _event: self._accept())
        self._listbox.bind("<Return>", lambda _event: self._accept())
        buttons = ttk.Frame(root)
        buttons.grid(row=3, column=0, sticky="e", pady=(10, 0))
        ttk.Button(buttons, text="OK", command=self._accept).pack(side=tk.LEFT)
        ttk.Button(buttons, text="Annuler", command=self.destroy).pack(side=tk.LEFT, padx=(6, 0))
        self._search_var.trace_add("write", lambda *_: self._refresh())
        self.bind("<Escape>", lambda _event: self.destroy())
        self._refresh()
        search.focus_set()
        self.grab_set()

    def show(self) -> str | None:
        self.wait_window()
        return self.result

    def _visible_cities(self) -> list[CatalogueCity]:
        search = _normalize_search_text(self._search_var.get().strip())
        return [
            city
            for city in self._cities
            if not search or search in _normalize_search_text(city.name)
        ]

    def _refresh(self):
        visible = self._visible_cities()
        self._listbox.delete(0, tk.END)
        for city in visible:
            self._listbox.insert(tk.END, city.name)
        if any(city.city_id == self._selected_city_id for city in visible):
            index = next(index for index, city in enumerate(visible) if city.city_id == self._selected_city_id)
            self._listbox.selection_set(index)
            self._listbox.activate(index)

    def _accept(self):
        selection = self._listbox.curselection()
        visible = self._visible_cities()
        if selection:
            self.result = visible[selection[0]].city_id
            self.destroy()


class TriangleEditorDialog(tk.Toplevel):
    """Dialogue modal d'édition locale d'un triangle, sans transaction Catalogue."""

    def __init__(
        self,
        parent,
        cities: list[CatalogueCity],
        triangle: ModelCatalogueTriangle | None = None,
    ):
        super().__init__(parent)
        self.title("Ajouter un triangle" if triangle is None else "Modifier un triangle")
        self.transient(parent)
        self.resizable(False, False)
        self.result: TriangleEditorResult | None = None
        self._cities = list(cities)
        self._city_by_id = {city.city_id: city for city in self._cities}
        default_opening_id = next((city.city_id for city in self._cities if not city.archived and city.name.casefold() == "bourges"), None)
        self._opening_city_id = triangle.opening_city_id if triangle else default_opening_id
        self._base_city_id = triangle.base_city_id if triangle else None
        self._light_city_id = triangle.light_city_id if triangle else None
        self._date_code_var = tk.StringVar(value=triangle.note if triangle else "")
        self._opening_var = tk.StringVar(value=self._city_name(self._opening_city_id))
        self._base_var = tk.StringVar(value=self._city_name(self._base_city_id))
        self._light_var = tk.StringVar(value=self._city_name(self._light_city_id))

        root = ttk.Frame(self, padding=12)
        root.grid(row=0, column=0, sticky="nsew")
        root.columnconfigure(1, weight=1)
        ttk.Label(root, text="Triangle", font=(None, 10, "bold")).grid(row=0, column=0, columnspan=3, sticky="w", pady=(0, 10))
        ttk.Label(root, text="Note").grid(row=1, column=0, sticky="w", pady=(0, 6))
        ttk.Combobox(root, textvariable=self._date_code_var, values=_TRIANGLE_NOTE_VALUES, width=18).grid(
            row=1, column=1, columnspan=2, sticky="ew", pady=(0, 6)
        )
        self._add_city_row(root, 2, "Ouverture", self._opening_var, "_opening_city_id")
        self._add_city_row(root, 3, "Base", self._base_var, "_base_city_id")
        self._add_city_row(root, 4, "Lumière", self._light_var, "_light_city_id")
        buttons = ttk.Frame(root)
        buttons.grid(row=5, column=0, columnspan=3, sticky="e", pady=(6, 0))
        ttk.Button(buttons, text="OK", command=self._accept).pack(side=tk.LEFT)
        ttk.Button(buttons, text="Annuler", command=self.destroy).pack(side=tk.LEFT, padx=(6, 0))
        self.bind("<Escape>", lambda _event: self.destroy())
        self.grab_set()

    def show(self) -> TriangleEditorResult | None:
        self.wait_window()
        return self.result

    def _add_city_row(self, parent, row: int, label: str, variable: tk.StringVar, attribute: str):
        ttk.Label(parent, text=label).grid(row=row, column=0, sticky="w", pady=(0, 6))
        ttk.Entry(parent, textvariable=variable, state="readonly", width=34).grid(row=row, column=1, sticky="ew", pady=(0, 6))
        ttk.Button(parent, text="...", width=3, command=lambda: self._choose_city(attribute)).grid(
            row=row, column=2, padx=(6, 0), pady=(0, 6)
        )

    def _choose_city(self, attribute: str):
        selected_city_id = getattr(self, attribute)
        selectable = [city for city in self._cities if not city.archived or city.city_id == selected_city_id]
        city_id = CitySelectionDialog(self, selectable, selected_city_id).show()
        if city_id is None:
            return
        setattr(self, attribute, city_id)
        {"_opening_city_id": self._opening_var, "_base_city_id": self._base_var, "_light_city_id": self._light_var}[attribute].set(self._city_name(city_id))

    def _city_name(self, city_id: str | None) -> str:
        return self._city_by_id[city_id].name if city_id is not None else ""

    def _accept(self):
        note = self._date_code_var.get().strip()
        city_ids = (self._opening_city_id, self._base_city_id, self._light_city_id)
        if not note or any(city_id is None for city_id in city_ids):
            messagebox.showerror("Triangle", "La note et les trois villes sont obligatoires.", parent=self)
            return
        if len(set(city_ids)) != 3:
            messagebox.showerror("Triangle", "Ouverture, Base et Lumière doivent être trois villes différentes.", parent=self)
            return
        self.result = TriangleEditorResult(note, *city_ids)
        self.destroy()


class TemplateRankSlot(tk.Frame):
    """Cible autonome d'un rang, prête à recevoir ultérieurement un glisser-déposer."""

    def __init__(self, parent, rank_number: int):
        super().__init__(parent, relief=tk.SUNKEN, borderwidth=1, background="#ffffff", padx=4, pady=3)
        self.rank_number = rank_number
        self.triangle_id: str | None = None
        self.drop_state = "normal"
        self.is_selected = False
        self._full_text = ""
        self._display_text = ""
        self.configure(takefocus=True)
        self.columnconfigure(1, weight=1)
        self._badge = tk.Label(
            self, text=str(rank_number), width=3, anchor="center", font=(None, 9, "bold"), relief=tk.RIDGE, borderwidth=1
        )
        self._badge.grid(row=0, column=0, sticky="nsw", padx=(0, 6))
        self._entry = tk.Entry(self, state="readonly", relief=tk.SUNKEN, borderwidth=1, readonlybackground="#ffffff", width=1)
        self._entry.grid(row=0, column=1, sticky="ew")
        self.bind("<Configure>", self._on_resize, add="+")
        self.set_triangle(None)

    def set_triangle(self, triangle_id: str | None, light_name: str | None = None) -> None:
        self.triangle_id = triangle_id
        self._full_text = light_name if triangle_id is not None and light_name is not None else "Déposer un triangle..."
        self._update_display_text()
        self.set_drop_state("normal")

    @property
    def tooltip_text(self) -> str:
        return self._full_text if self.triangle_id is not None else ""

    def _on_resize(self, _event=None):
        self._update_display_text()

    def _update_display_text(self):
        available_width = max(20, self._entry.winfo_width() - 8)
        font = tkfont.Font(font=self._entry.cget("font"))
        text = self._full_text
        if font.measure(text) > available_width:
            ellipsis = "..."
            while text and font.measure(text + ellipsis) > available_width:
                text = text[:-1]
            text += ellipsis
        self._display_text = text
        self._entry._catalogue_tooltip_text = self._full_text if text != self._full_text else ""
        self._entry.configure(state="normal")
        self._entry.delete(0, tk.END)
        self._entry.insert(0, text)
        self._entry.configure(state="readonly")

    def set_drop_state(self, state: str) -> None:
        if state not in {"normal", "source", "valid", "replace", "move", "swap", "invalid"}:
            raise ValueError(f"État de dépôt inconnu : {state}")
        self.drop_state = state
        background = {
            "normal": "#ffffff", "source": "#dbeafe", "valid": "#e8f5e9", "replace": "#fff8d6",
            "move": "#e7f2ff", "swap": "#eee8ff", "invalid": "#fde8e7",
        }[state]
        border = tk.RIDGE if self.is_selected or state in {"source", "valid", "replace", "move", "swap"} else tk.SUNKEN
        self.configure(background=background, relief=border)
        self._badge.configure(background=background, relief=tk.RIDGE)
        self._entry.configure(readonlybackground=background, foreground="#222222" if self.triangle_id is not None else "#777777")

    def set_selected(self, selected: bool) -> None:
        self.is_selected = selected
        self.set_drop_state(self.drop_state)


class TemplateBaseSlot(tk.Frame):
    """Champ readonly compact de la base, avec action de vidage du couple."""

    def __init__(self, parent, on_clear):
        super().__init__(parent, relief=tk.SUNKEN, borderwidth=1, background="#ffffff", padx=4, pady=3)
        self.columnconfigure(0, weight=1)
        self._full_text = "—"
        self._entry = tk.Entry(self, state="readonly", relief=tk.SUNKEN, borderwidth=1, readonlybackground="#ffffff", width=1)
        self._entry.grid(row=0, column=0, sticky="ew")
        self.clear_button = ttk.Button(self, text="×", width=2, command=on_clear)
        self.clear_button.grid(row=0, column=1, sticky="e", padx=(3, 0))

    @property
    def tooltip_text(self) -> str:
        return self._full_text if self._full_text not in {"—", "Bases différentes"} else ""

    def set_base(self, text: str, *, enabled: bool):
        self._full_text = text
        self._entry.configure(state="normal")
        self._entry.delete(0, tk.END)
        self._entry.insert(0, text)
        self._entry.configure(state="readonly", foreground="#222222" if text != "—" else "#777777")
        self.clear_button.configure(state=tk.NORMAL if enabled else tk.DISABLED)


class TemplatePairRow(ttk.Frame):
    """Ligne d'un couple de rangs, responsable de son affichage local."""

    def __init__(self, parent, pair_number: int, on_clear):
        super().__init__(parent, padding=(0, 0))
        self.pair_number = pair_number
        self.columnconfigure(0, weight=0, minsize=_TEMPLATE_BASE_COLUMN_WIDTH)
        self.columnconfigure(1, weight=1)
        self.columnconfigure(2, weight=1)
        self.base_slot = TemplateBaseSlot(self, on_clear=lambda: on_clear(self))
        self.base_slot.grid(row=0, column=0, sticky="nsew", padx=(2, 6))
        self.odd_rank_slot = TemplateRankSlot(self, pair_number * 2 - 1)
        self.odd_rank_slot.grid(row=0, column=1, sticky="nsew", padx=(0, 4))
        self.even_rank_slot = TemplateRankSlot(self, pair_number * 2)
        self.even_rank_slot.grid(row=0, column=2, sticky="nsew")

    def set_column_widths(self, base_width: int, rank_width: int) -> None:
        self.columnconfigure(0, minsize=base_width)
        self.columnconfigure(1, minsize=rank_width)
        self.columnconfigure(2, minsize=rank_width)

    def set_triangles(
        self,
        odd_triangle_id: str | None,
        odd_light_name: str | None,
        even_triangle_id: str | None,
        even_light_name: str | None,
        base_name: str | None,
    ) -> None:
        self.odd_rank_slot.set_triangle(odd_triangle_id, odd_light_name)
        self.even_rank_slot.set_triangle(even_triangle_id, even_light_name)
        self.base_slot.set_base(base_name or "—", enabled=odd_triangle_id is not None or even_triangle_id is not None)


class CatalogueWindow(tk.Toplevel):
    """Prototype non modal du catalogue et de son premier onglet Villes."""

    _CITY_CSV_HEADER = ("Nom", "Latitude", "Longitude")
    _BEACON_XLSX_HEADER = ("Nom",)
    _TRIANGLE_CSV_HEADER = ("Note", "Ouverture", "Base", "Lumiere")
    _TEMPLATE_CSV_HEADER = ("Rang", "Ouverture", "Base", "Lumiere")
    _TEMPLATE_NOTE_ORDER = {"do": 0, "si": 1, "la": 2, "sol": 3, "fa": 4, "mi": 5, "re": 6, "zone": 7}

    def __init__(
        self,
        parent,
        *,
        catalogue: Catalogue,
        catalogue_path: str | Path | None = None,
        on_catalogue_applied: Callable[[Catalogue], None] | None = None,
        is_beacon_referenced: Callable[[str], bool] | None = None,
        is_book_referenced: Callable[[str], bool] | None = None,
    ):
        super().__init__(parent)
        self.title("Gestion du catalogue")
        self.geometry("1000x700")
        self.minsize(760, 500)
        self._application_context = ApplicationContext.from_environment()
        self._catalogue_path = (
            Path(catalogue_path)
            if catalogue_path is not None
            else ApplicationPaths.from_runtime().catalogue_path_for_mode(
                self._application_context.mode
            )
        )
        self._on_catalogue_applied = on_catalogue_applied
        self._is_beacon_referenced = is_beacon_referenced
        self._is_book_referenced = is_book_referenced
        # Copie transactionnelle : le Catalogue runtime du viewer reste intact jusqu'à Appliquer.
        self.catalogue = catalogue.clone()
        self._validated_catalogue = self.catalogue.clone()
        self._selected_city_id: str | None = None
        self._selected_beacon_id: str | None = None
        self._selected_triangle_id: str | None = None
        self._selected_template_id: str | None = self.catalogue.default_template_id
        self._selected_map_id: str | None = self.catalogue.default_map_id
        self._selected_book_id: str | None = self.catalogue.default_book_id
        self._selected_calibration_city_id: str | None = None
        self._selected_template_triangle_id: str | None = None
        self._selected_template_rank_slot: TemplateRankSlot | None = None
        self._template_drag_triangle_id: str | None = None
        self._template_drag_started = False
        self._template_drag_source_iid: str | None = None
        self._template_drag_source_slot: TemplateRankSlot | None = None
        self._template_drag_target_slot: TemplateRankSlot | None = None
        self._template_drag_start_root: tuple[int, int] | None = None
        self._template_drag_ghost: tk.Toplevel | None = None
        self._updating_detail = False
        self._updating_template_detail = False
        self._is_dirty = False
        self._paths = ApplicationPaths.from_runtime()
        self._map_calibration = CatalogueMapCalibrationController(
            self.catalogue,
            self._paths,
            allow_system_map_editing=self._application_context.mode == "SYS",
        )
        self._book_assets = CatalogueBookAssetController(
            self.catalogue,
            self._paths,
            allow_system_book_editing=self._application_context.mode == "SYS",
        )

        self._search_var = tk.StringVar()
        self._show_archived_var = tk.BooleanVar(value=False)
        self._beacon_search_var = tk.StringVar()
        self._show_archived_beacons_var = tk.BooleanVar(value=False)
        self._beacon_archived_var = tk.BooleanVar(value=False)
        self._name_var = tk.StringVar()
        self._archived_var = tk.BooleanVar(value=False)
        self._triangle_date_filter_var = tk.StringVar(value="Tous")
        self._triangle_base_filter_var = tk.StringVar()
        self._triangle_light_filter_var = tk.StringVar()
        self._triangle_status_filter_var = tk.StringVar(value="Actif")
        initial_template = self._get_selected_template()
        self._template_name_var = tk.StringVar(value=initial_template.name if initial_template else "")
        self._template_default_var = tk.BooleanVar(value=initial_template is not None)
        self._template_description_var = tk.StringVar()
        self._template_note_filter_var = tk.StringVar(value="Tous")
        self._template_base_filter_var = tk.StringVar()
        self._template_light_filter_var = tk.StringVar()
        self._map_description_var = tk.StringVar()
        self._map_default_var = tk.BooleanVar(value=False)
        self._map_scale_var = tk.StringVar()
        self._map_interaction_mode = "hand"
        self._updating_map_placement = False
        self._book_name_var = tk.StringVar()
        self._book_description_var = tk.StringVar()
        self._book_default_var = tk.BooleanVar(value=False)
        self._updating_book_detail = False

        self._load_icons()
        self._build_ui()
        self._refresh_city_list()
        self._refresh_beacon_list()
        self._refresh_triangle_tree()
        self._refresh_maps()
        self._refresh_books()
        self._load_map()
        self.protocol("WM_DELETE_WINDOW", self.request_close)

    def _load_icons(self):
        """Charge une seule fois les icones utilisees par la fenetre Catalogue."""
        images_dir = ApplicationPaths.from_runtime().images_dir
        self._icon_clipboard = tk.PhotoImage(file=images_dir / "clipboard.png")
        self._icon_map_pin_plus = tk.PhotoImage(file=images_dir / "map-pin-plus.png")
        self._icon_hexagon_plus = tk.PhotoImage(file=images_dir / "hexagon-plus.png")
        self._icon_archive = tk.PhotoImage(file=images_dir / "archive.png")
        self._icon_archive_off = tk.PhotoImage(file=images_dir / "archive-off.png")
        self._icon_trash = tk.PhotoImage(file=images_dir / "trash.png")
        self._icon_template_default = tk.PhotoImage(file=images_dir / "checkbox.png")
        self._icon_duplicate = tk.PhotoImage(file=images_dir / "duplicate.png")
        self._icon_map_add = tk.PhotoImage(file=images_dir / "map-pin-plus.png")
        self._icon_focus = tk.PhotoImage(file=images_dir / "focus.png")
        self._icon_hand = tk.PhotoImage(file=images_dir / "hand-click.png")
        self._icon_book_add = tk.PhotoImage(file=images_dir / "book - plus.png")

    def _attach_tooltip(self, widget, text: str):
        return attach_tooltip(widget, text)

    @staticmethod
    def _set_tooltip_text(widget, text: str):
        tooltip = getattr(widget, "_tooltip", None)
        if tooltip is None:
            attach_tooltip(widget, text)
        else:
            tooltip.set_text(text)

    def _build_ui(self):
        root = ttk.Frame(self, padding=10)
        root.pack(fill=tk.BOTH, expand=True)
        root.rowconfigure(0, weight=1)
        root.columnconfigure(0, weight=1)

        notebook = ttk.Notebook(root)
        notebook.grid(row=0, column=0, sticky="nsew")
        self._cities_tab = ttk.Frame(notebook, padding=8)
        self._beacons_tab = ttk.Frame(notebook, padding=8)
        self._triangles_tab = ttk.Frame(notebook, padding=8)
        self._templates_tab = ttk.Frame(notebook, padding=8)
        self._maps_tab = ttk.Frame(notebook, padding=8)
        self._books_tab = ttk.Frame(notebook, padding=8)
        notebook.add(self._cities_tab, text="Villes (0)")
        notebook.add(self._beacons_tab, text="Balises (0)")
        notebook.add(self._triangles_tab, text="Triangles (0)")
        notebook.add(self._templates_tab, text="Templates (0)")
        notebook.add(self._maps_tab, text="Cartes (0)")
        notebook.add(self._books_tab, text="Livres (0)")
        self._catalogue_notebook = notebook
        self._build_cities_tab()
        self._build_beacons_tab()
        self._build_triangles_tab()
        self._build_templates_tab()
        self._build_maps_tab()
        self._build_books_tab()

        bottom = ttk.Frame(root)
        bottom.grid(row=1, column=0, sticky="ew", pady=(10, 0))
        self._import_button = ttk.Button(bottom, text="Importer...")
        self._import_button.pack(side=tk.LEFT)
        self._export_button = ttk.Button(bottom, text="Exporter...")
        self._export_button.pack(side=tk.LEFT, padx=(6, 0))
        edit_actions = ttk.Frame(bottom)
        edit_actions.pack(side=tk.RIGHT)
        self._apply_button = ttk.Button(edit_actions, text="Appliquer", command=self._apply_changes, state=tk.DISABLED)
        self._apply_button.pack(side=tk.LEFT)
        self._cancel_button = ttk.Button(edit_actions, text="Annuler", command=self._cancel_changes, state=tk.DISABLED)
        self._cancel_button.pack(side=tk.LEFT, padx=(6, 0))
        ttk.Button(edit_actions, text="Fermer", command=self.request_close).pack(side=tk.LEFT, padx=(6, 0))
        self._catalogue_notebook.bind("<<NotebookTabChanged>>", self._on_catalogue_tab_changed)
        self._update_context_actions()

    def _build_cities_tab(self):
        self._cities_tab.rowconfigure(0, weight=1)
        self._cities_tab.columnconfigure(0, weight=1)
        panes = ttk.PanedWindow(self._cities_tab, orient=tk.HORIZONTAL)
        panes.grid(row=0, column=0, sticky="nsew")
        master = ttk.Frame(panes, padding=(0, 0, 8, 0))
        detail = ttk.Frame(panes, padding=(8, 0, 0, 0))
        panes.add(master, weight=1)
        panes.add(detail, weight=2)

        ttk.Label(master, text="Rechercher").pack(anchor="w")
        ttk.Entry(master, textvariable=self._search_var).pack(fill=tk.X, pady=(2, 8))
        self._search_var.trace_add("write", lambda *_: self._refresh_city_list())
        ttk.Checkbutton(master, text="Afficher les villes archivées", variable=self._show_archived_var,
                        command=self._refresh_city_list).pack(anchor="w", pady=(0, 8))
        actions = ttk.Frame(master)
        actions.pack(fill=tk.X, pady=(0, 8))
        self._city_add_button = ttk.Button(actions, image=self._icon_map_pin_plus, command=self._add_city)
        self._city_add_button.pack(side=tk.LEFT)
        self._attach_tooltip(self._city_add_button, "Ajouter une ville")
        self._city_archive_button = ttk.Button(
            actions, image=self._icon_archive, command=self._archive_selected_city, state=tk.DISABLED
        )
        self._city_archive_button.pack(side=tk.LEFT, padx=4)
        self._attach_tooltip(self._city_archive_button, "Archiver la ville")
        self._city_delete_button = ttk.Button(
            actions, image=self._icon_trash, command=self._delete_selected_city, state=tk.DISABLED
        )
        self._city_delete_button.pack(side=tk.LEFT)
        self._attach_tooltip(self._city_delete_button, "Supprimer la ville")
        list_frame = ttk.Frame(master)
        list_frame.pack(fill=tk.BOTH, expand=True)
        self._city_listbox = tk.Listbox(list_frame, exportselection=False)
        self._city_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self._city_listbox.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self._city_listbox.configure(yscrollcommand=scrollbar.set)
        self._city_listbox.bind("<<ListboxSelect>>", self._on_city_selected)

        detail.rowconfigure(4, weight=1)
        detail.columnconfigure(1, weight=1)
        ttk.Label(detail, text="Nom").grid(row=0, column=0, sticky="w", pady=(0, 6))
        ttk.Entry(detail, textvariable=self._name_var).grid(row=0, column=1, sticky="ew", pady=(0, 6))
        ttk.Label(detail, text="Coordonnées").grid(row=1, column=0, sticky="w", pady=(0, 6))
        coordinates = ttk.Frame(detail)
        coordinates.grid(row=1, column=1, sticky="w", pady=(0, 6))
        self._latitude_editor = DmsCoordinateEditor(
            coordinates, coordinate_type="latitude", on_change=self._save_detail_changes
        )
        self._latitude_editor.grid(row=0, column=0, sticky="w")
        self._longitude_editor = DmsCoordinateEditor(
            coordinates, coordinate_type="longitude", on_change=self._save_detail_changes
        )
        self._longitude_editor.grid(row=0, column=1, sticky="w", padx=(12, 0))
        ttk.Button(coordinates, image=self._icon_clipboard, command=self._paste_coordinates).grid(
            row=0, column=2, sticky="w", padx=(6, 0)
        )
        ttk.Checkbutton(detail, text="Ville archivée", variable=self._archived_var).grid(
            row=2, column=0, columnspan=2, sticky="w", pady=(2, 6)
        )
        usage = ttk.Frame(detail)
        usage.grid(row=3, column=0, columnspan=2, sticky="w", pady=(0, 8))
        ttk.Label(usage, text="Référencée par :").pack(side=tk.LEFT, padx=(0, 8))
        self._city_triangle_count_label = ttk.Label(usage, text="Triangles : 0")
        self._city_triangle_count_label.pack(side=tk.LEFT)
        self._city_triangle_references_button = ttk.Button(
            usage,
            text="...",
            width=3,
            state=tk.DISABLED,
            command=self._show_city_triangle_references,
        )
        self._city_triangle_references_button.pack(side=tk.LEFT, padx=(4, 14))
        self._city_beacon_count_label = ttk.Label(usage, text="Balises : 0")
        self._city_beacon_count_label.pack(side=tk.LEFT)
        self._city_beacon_references_button = ttk.Button(
            usage, text="...", width=3, state=tk.DISABLED, command=self._show_city_beacon_references,
        )
        self._city_beacon_references_button.pack(side=tk.LEFT, padx=(4, 0))
        self._map_view = GeoMapView(
            detail,
            on_marker_selected=self._on_map_marker_selected,
            initial_fit_zoom=2.25,
            minimum_fit_zoom=2.25,
            maximum_zoom=1.0,
        )
        self._map_view.grid(row=4, column=0, columnspan=2, sticky="nsew")

        self._name_var.trace_add("write", self._save_detail_changes)
        self._archived_var.trace_add("write", self._save_detail_changes)

    def _build_beacons_tab(self):
        self._beacons_tab.rowconfigure(0, weight=1)
        self._beacons_tab.columnconfigure(0, weight=1)
        panes = ttk.PanedWindow(self._beacons_tab, orient=tk.HORIZONTAL)
        panes.grid(row=0, column=0, sticky="nsew")
        master = ttk.Frame(panes, padding=(0, 0, 8, 0))
        detail = ttk.Frame(panes, padding=(8, 0, 0, 0))
        panes.add(master, weight=1)
        panes.add(detail, weight=2)

        ttk.Label(master, text="Rechercher").pack(anchor="w")
        ttk.Entry(master, textvariable=self._beacon_search_var).pack(fill=tk.X, pady=(2, 8))
        self._beacon_search_var.trace_add("write", lambda *_: self._refresh_beacon_list())
        ttk.Checkbutton(
            master,
            text="Afficher les balises archivées",
            variable=self._show_archived_beacons_var,
            command=self._refresh_beacon_list,
        ).pack(anchor="w", pady=(0, 8))
        actions = ttk.Frame(master)
        actions.pack(fill=tk.X, pady=(0, 8))
        self._beacon_add_button = ttk.Button(actions, image=self._icon_map_pin_plus, command=self._add_beacon)
        self._beacon_add_button.pack(side=tk.LEFT)
        self._attach_tooltip(self._beacon_add_button, "Ajouter une balise")
        self._beacon_archive_button = ttk.Button(
            actions, image=self._icon_archive, command=self._archive_selected_beacon, state=tk.DISABLED,
        )
        self._beacon_archive_button.pack(side=tk.LEFT, padx=4)
        self._attach_tooltip(self._beacon_archive_button, "Archiver la balise")
        self._beacon_delete_button = ttk.Button(
            actions, image=self._icon_trash, command=self._delete_selected_beacon, state=tk.DISABLED,
        )
        self._beacon_delete_button.pack(side=tk.LEFT)
        self._attach_tooltip(self._beacon_delete_button, "Supprimer la balise")
        list_frame = ttk.Frame(master)
        list_frame.pack(fill=tk.BOTH, expand=True)
        self._beacon_listbox = tk.Listbox(list_frame, exportselection=False)
        self._beacon_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self._beacon_listbox.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self._beacon_listbox.configure(yscrollcommand=scrollbar.set)
        self._beacon_listbox.bind("<<ListboxSelect>>", self._on_beacon_selected)

        detail.rowconfigure(3, weight=1)
        detail.columnconfigure(1, weight=1)
        ttk.Label(detail, text="Ville").grid(row=0, column=0, sticky="w", pady=(0, 6))
        self._beacon_city_label = ttk.Label(detail, text="")
        self._beacon_city_label.grid(row=0, column=1, sticky="w", pady=(0, 6))
        ttk.Label(detail, text="Coordonnées").grid(row=1, column=0, sticky="w", pady=(0, 6))
        coordinates = ttk.Frame(detail)
        coordinates.grid(row=1, column=1, sticky="w", pady=(0, 6))
        self._beacon_latitude_editor = DmsCoordinateEditor(coordinates, coordinate_type="latitude")
        self._beacon_latitude_editor.grid(row=0, column=0, sticky="w")
        self._beacon_longitude_editor = DmsCoordinateEditor(coordinates, coordinate_type="longitude")
        self._beacon_longitude_editor.grid(row=0, column=1, sticky="w", padx=(12, 0))
        self._set_dms_editor_readonly(self._beacon_latitude_editor)
        self._set_dms_editor_readonly(self._beacon_longitude_editor)
        ttk.Checkbutton(detail, text="Balise archivée", variable=self._beacon_archived_var, state=tk.DISABLED).grid(
            row=2, column=0, columnspan=2, sticky="w", pady=(2, 6)
        )
        self._beacon_map_view = GeoMapView(
            detail,
            on_marker_selected=self._on_beacon_map_marker_selected,
            initial_fit_zoom=2.25,
            minimum_fit_zoom=2.25,
            maximum_zoom=1.0,
        )
        self._beacon_map_view.grid(row=3, column=0, columnspan=2, sticky="nsew")

    @staticmethod
    def _set_dms_editor_readonly(editor: DmsCoordinateEditor) -> None:
        editor._hemisphere.configure(state=tk.DISABLED)
        for spinbox in (editor._degrees, editor._minutes, editor._seconds):
            spinbox.configure(state=tk.DISABLED)

    def _load_map(self):
        try:
            map_id = self.catalogue.catalogue_reference_map_id
            if map_id is None:
                return
            calibrated_map = load_calibrated_catalogue_map(
                self.catalogue.get_map(map_id),
                CatalogueMapAssetResolver(ApplicationPaths.from_runtime()),
            )
            self._map_view.set_map(calibrated_map)
            self._beacon_map_view.set_map(calibrated_map)
            self._triangle_map_view.set_map(calibrated_map)
        except (FileNotFoundError, OSError, ValueError) as exc:
            messagebox.showwarning("Carte du catalogue", str(exc), parent=self)

    def _on_catalogue_tab_changed(self, _event=None):
        self._update_context_actions()

    def _update_context_actions(self):
        active_tab = self._catalogue_notebook.select()
        if active_tab == str(self._cities_tab):
            self._import_button.configure(command=self._import_csv, state=tk.NORMAL)
            self._export_button.configure(command=self._export_cities_csv, state=tk.NORMAL)
        elif active_tab == str(self._beacons_tab):
            self._import_button.configure(command=self._import_beacons_xlsx, state=tk.NORMAL)
            self._export_button.configure(command=self._export_beacons_xlsx, state=tk.NORMAL)
        elif active_tab == str(self._triangles_tab):
            self._import_button.configure(command=self._import_triangles_csv, state=tk.NORMAL)
            self._export_button.configure(command=self._export_triangles_csv, state=tk.NORMAL)
        elif active_tab == str(self._templates_tab):
            has_template = self._get_selected_template() is not None
            self._import_button.configure(
                command=self._import_template_csv,
                state=tk.NORMAL if has_template else tk.DISABLED,
            )
            self._export_button.configure(
                command=self._export_template_csv,
                state=tk.NORMAL if has_template else tk.DISABLED,
            )
        elif active_tab == str(self._books_tab):
            selected = self._get_selected_book()
            self._import_button.configure(
                command=self._import_selected_book,
                state=tk.NORMAL if selected is not None and not self._book_assets.is_readonly(selected) else tk.DISABLED,
            )
            self._export_button.configure(
                command=self._export_selected_book,
                state=tk.NORMAL if selected is not None else tk.DISABLED,
            )
        else:
            self._import_button.configure(command=lambda: None, state=tk.DISABLED)
            self._export_button.configure(command=lambda: None, state=tk.DISABLED)

    def _choose_export_path(self, title: str, *, initialfile: str | None = None) -> str:
        return filedialog.asksaveasfilename(
            parent=self,
            title=title,
            initialdir=self._paths.exports_dir,
            defaultextension=".csv",
            initialfile=initialfile,
            filetypes=[("Fichiers CSV", "*.csv"), ("Tous les fichiers", "*.*")],
        )

    def _export_cities_csv(self):
        path = self._choose_export_path("Exporter les villes")
        if not path:
            return
        with open(path, "w", encoding="utf-8-sig", newline="") as csv_file:
            writer = csv.writer(csv_file, delimiter=";")
            writer.writerow(self._CITY_CSV_HEADER)
            writer.writerows((city.name, city.latitude, city.longitude) for city in self.catalogue.iter_cities())

    def _export_beacons_xlsx(self):
        path = filedialog.asksaveasfilename(
            parent=self,
            title="Exporter les balises",
            initialdir=self._paths.exports_dir,
            defaultextension=".xlsx",
            initialfile="balises.xlsx",
            filetypes=[("Fichiers Excel", "*.xlsx"), ("Tous les fichiers", "*.*")],
        )
        if not path:
            return
        try:
            from openpyxl import Workbook

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Balises"
            worksheet.append(self._BEACON_XLSX_HEADER)
            for name in self._beacon_export_names():
                worksheet.append((name,))
            workbook.save(path)
            workbook.close()
        except (OSError, ValueError) as exc:
            messagebox.showerror("Exporter les balises", str(exc), parent=self)

    def _export_triangles_csv(self):
        path = self._choose_export_path("Exporter les triangles")
        if not path:
            return
        with open(path, "w", encoding="utf-8-sig", newline="") as csv_file:
            writer = csv.writer(csv_file, delimiter=";")
            writer.writerow(self._TRIANGLE_CSV_HEADER)
            writer.writerows(
                (triangle.note, opening.name, base.name, light.name)
                for triangle in sorted(self.catalogue.iter_triangles(), key=self._triangle_sort_key)
                for opening, base, light in (self._model_triangle_cities(triangle),)
            )

    @staticmethod
    def _template_export_filename(template_name: str) -> str:
        return "".join("_" if character in '<>:"/\\|?*' else character for character in template_name).strip() + ".csv"

    def _export_template_csv(self):
        template = self._get_selected_template()
        if template is None:
            return
        status = self.catalogue.get_template_validation_status(template.template_id)
        if status.state == "Incomplet":
            messagebox.showwarning(
                "Exporter un template",
                (f"Le template est incomplet : {status.filled_ranks} rangs sur 32 sont renseignés.\n\n"
                 "Complétez le template avant de l’exporter."),
                parent=self,
            )
            return
        if status.state == "Invalide":
            messagebox.showwarning(
                "Exporter un template",
                f"Le template est invalide.\n\n{status.message or ''}".rstrip(),
                parent=self,
            )
            return
        path = self._choose_export_path(
            "Exporter le template",
            initialfile=self._template_export_filename(template.name),
        )
        if not path:
            return
        try:
            with open(path, "w", encoding="utf-8-sig", newline="") as csv_file:
                writer = csv.writer(csv_file, delimiter=";")
                writer.writerow(self._TEMPLATE_CSV_HEADER)
                for rank_number, triangle_id in enumerate(template.triangle_ids_by_rank, start=1):
                    triangle = self.catalogue.get_triangle(triangle_id) if triangle_id is not None else None
                    if triangle is None:
                        raise ValueError("Le template validé contient un rang vide.")
                    opening, base, light = self._model_triangle_cities(triangle)
                    writer.writerow((
                        rank_number,
                        opening.name,
                        base.name,
                        light.name,
                    ))
        except (OSError, UnicodeError, csv.Error) as exc:
            messagebox.showerror("Exporter un template", str(exc), parent=self)

    def _build_triangles_tab(self):
        self._triangles_tab.rowconfigure(1, weight=1)
        self._triangles_tab.columnconfigure(0, weight=1)
        filters = ttk.Frame(self._triangles_tab)
        filters.grid(row=0, column=0, sticky="ew", pady=(0, 8))
        actions = ttk.Frame(filters)
        actions.grid(row=0, column=0, columnspan=6, sticky="w", pady=(0, 6))
        self._triangle_add_button = ttk.Button(actions, image=self._icon_hexagon_plus, command=self._add_triangle)
        self._triangle_add_button.pack(side=tk.LEFT)
        self._attach_tooltip(self._triangle_add_button, "Ajouter un triangle")
        self._triangle_archive_button = ttk.Button(
            actions, image=self._icon_archive, command=self._toggle_archive_selected_triangle, state=tk.DISABLED
        )
        self._triangle_archive_button.pack(side=tk.LEFT, padx=4)
        self._attach_tooltip(self._triangle_archive_button, "Archiver le triangle")
        self._triangle_delete_button = ttk.Button(
            actions, image=self._icon_trash, command=self._delete_selected_triangle, state=tk.DISABLED
        )
        self._triangle_delete_button.pack(side=tk.LEFT)
        self._attach_tooltip(self._triangle_delete_button, "Supprimer le triangle")
        ttk.Label(filters, text="Note :").grid(row=1, column=0, sticky="w")
        self._triangle_date_filter = ttk.Combobox(
            filters, textvariable=self._triangle_date_filter_var, values=("Tous",), state="readonly", width=8
        )
        self._triangle_date_filter.grid(row=1, column=1, padx=(0, 8))
        ttk.Label(filters, text="Base :").grid(row=1, column=2, sticky="w")
        ttk.Entry(filters, textvariable=self._triangle_base_filter_var, width=13).grid(row=1, column=3, padx=(0, 8))
        ttk.Label(filters, text="Lumière :").grid(row=1, column=4, sticky="w")
        ttk.Entry(filters, textvariable=self._triangle_light_filter_var, width=13).grid(row=1, column=5, padx=(0, 8))
        ttk.Label(filters, text="Statut :").grid(row=1, column=6, sticky="w")
        self._triangle_status_filter = ttk.Combobox(
            filters,
            textvariable=self._triangle_status_filter_var,
            values=("Actif", "Archivé", "Tout"),
            state="readonly",
            width=9,
        )
        self._triangle_status_filter.grid(row=1, column=7, sticky="w")
        self._triangle_date_filter.bind("<<ComboboxSelected>>", lambda _event: self._refresh_triangle_tree())
        self._triangle_status_filter.bind("<<ComboboxSelected>>", lambda _event: self._refresh_triangle_tree())
        self._triangle_base_filter_var.trace_add("write", lambda *_: self._refresh_triangle_tree())
        self._triangle_light_filter_var.trace_add("write", lambda *_: self._refresh_triangle_tree())

        panes = ttk.PanedWindow(self._triangles_tab, orient=tk.HORIZONTAL)
        panes.grid(row=1, column=0, sticky="nsew")
        list_frame = ttk.Frame(panes, padding=(0, 0, 8, 0))
        detail = ttk.Frame(panes, padding=(8, 0, 0, 0))
        panes.add(list_frame, weight=0)
        panes.add(detail, weight=1)

        columns = ("code", "opening", "base", "light")
        self._triangle_tree = ttk.Treeview(list_frame, columns=columns, show="headings", selectmode="browse")
        headings = (("code", "Note", 55), ("opening", "Ouverture", 95), ("base", "Base", 145), ("light", "Lumière", 170))
        for key, text, width in headings:
            self._triangle_tree.heading(key, text=text)
            self._triangle_tree.column(key, width=width, stretch=False)
        self._triangle_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self._triangle_tree.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self._triangle_tree.configure(yscrollcommand=scrollbar.set)
        self._triangle_tree.bind("<<TreeviewSelect>>", self._on_triangle_selected)
        self._triangle_tree.bind("<Double-Button-1>", lambda _event: self._edit_selected_triangle())

        detail.rowconfigure(4, weight=1)
        detail.columnconfigure(0, weight=1)
        self._triangle_distances_label = ttk.Label(detail, text="Distances", justify=tk.LEFT)
        self._triangle_distances_label.grid(row=0, column=0, sticky="w", pady=(0, 8))
        self._triangle_angles_label = ttk.Label(detail, text="Angles", justify=tk.LEFT)
        self._triangle_angles_label.grid(row=1, column=0, sticky="w", pady=(0, 8))
        self._triangle_status_label = ttk.Label(detail, text="Statut", justify=tk.LEFT)
        self._triangle_status_label.grid(row=2, column=0, sticky="w", pady=(0, 8))
        triangle_usage = ttk.Frame(detail)
        triangle_usage.grid(row=3, column=0, sticky="w", pady=(0, 8))
        ttk.Label(triangle_usage, text="Référencé par :").pack(side=tk.LEFT, padx=(0, 8))
        ttk.Label(triangle_usage, text="Hypothèses : 0").pack(side=tk.LEFT)
        ttk.Button(triangle_usage, text="...", width=3, state=tk.DISABLED).pack(side=tk.LEFT, padx=(4, 0))
        self._triangle_map_view = GeoMapView(
            detail,
            initial_fit_zoom=2.25,
            minimum_fit_zoom=2.25,
            maximum_zoom=1.0,
        )
        self._triangle_map_view.grid(row=4, column=0, sticky="nsew")

    def _build_templates_tab(self):
        self._templates_tab.rowconfigure(2, weight=1)
        self._templates_tab.columnconfigure(0, weight=1)

        template_header = ttk.Frame(self._templates_tab)
        template_header.grid(row=0, column=0, sticky="ew")
        ttk.Label(template_header, text="Template :").pack(side=tk.LEFT)
        self._template_selector = ttk.Combobox(
            template_header, textvariable=self._template_name_var, state="normal", width=34
        )
        self._template_selector.pack(side=tk.LEFT, padx=(6, 6))
        self._template_selector.bind("<<ComboboxSelected>>", self._on_template_selected)
        self._template_selector.bind("<ButtonPress-1>", lambda _event: self._commit_template_name(), add="+")
        self._template_selector.bind("<Return>", lambda _event: self._commit_template_name())
        self._template_selector.bind("<FocusOut>", lambda _event: self._commit_template_name(), add="+")

        self._template_add_button = ttk.Button(template_header, image=self._icon_hexagon_plus, command=self._add_template)
        self._template_add_button.pack(side=tk.LEFT)
        self._attach_tooltip(self._template_add_button, "Ajouter un template")

        self._template_duplicate_button = ttk.Button(
            template_header,
            image=self._icon_duplicate,
            command=self._duplicate_selected_template,
            state=tk.DISABLED,
        )
        self._template_duplicate_button.pack(side=tk.LEFT, padx=(4, 0))
        self._attach_tooltip(
            self._template_duplicate_button,
            "Dupliquer le template",
        )        

        self._template_default_button = ttk.Button(
            template_header,
            image=self._icon_template_default,
            command=self._set_selected_template_as_default,
            state=tk.DISABLED,
        )
        self._template_default_button.pack(side=tk.LEFT, padx=(4, 0))
        self._attach_tooltip(self._template_default_button, "Définir comme template par défaut")

        self._template_archive_button = ttk.Button(
            template_header, image=self._icon_archive, command=self._toggle_archive_selected_template, state=tk.DISABLED
        )
        self._template_archive_button.pack(side=tk.LEFT, padx=4)
        self._attach_tooltip(self._template_archive_button, "Archiver le template")

        self._template_delete_button = ttk.Button(
            template_header, image=self._icon_trash, command=self._delete_selected_template, state=tk.DISABLED
        )
        self._template_delete_button.pack(side=tk.LEFT)
        self._attach_tooltip(self._template_delete_button, "Supprimer le template")

        description = ttk.Frame(self._templates_tab)
        description.grid(row=1, column=0, sticky="ew", pady=(8, 10))
        description.columnconfigure(2, weight=1)
        self._template_default_check = ttk.Checkbutton(
            description,
            text="Défaut",
            variable=self._template_default_var,
            state=tk.DISABLED,
        )
        self._template_default_check.grid(row=0, column=0, sticky="w", padx=(0, 12))
        ttk.Label(description, text="Description :").grid(row=0, column=1, sticky="w")
        self._template_description_entry = ttk.Entry(description, textvariable=self._template_description_var)
        self._template_description_entry.grid(row=0, column=2, sticky="ew", padx=(6, 12))
        self._template_validation_status_label = ttk.Label(description, anchor="e")
        self._template_validation_status_label.grid(row=0, column=3, sticky="e")
        self._template_description_var.trace_add("write", self._save_template_description)

        panes = ttk.PanedWindow(self._templates_tab, orient=tk.HORIZONTAL)
        panes.grid(row=2, column=0, sticky="nsew")
        source = ttk.Frame(panes, padding=(0, 0, 8, 0))
        target = ttk.Frame(panes, padding=(8, 0, 0, 0))
        panes.add(source, weight=0)
        panes.add(target, weight=1)
        self._build_template_triangle_source(source)
        self._build_template_ranks_grid(target)
        self.after_idle(lambda: panes.sashpos(0, max(280, int(panes.winfo_width() * 0.36))))
        self._refresh_templates()
        self._refresh_template_triangle_tree()

    def _build_maps_tab(self):
        self._maps_tab.rowconfigure(2, weight=1)
        self._maps_tab.columnconfigure(0, weight=1)
        header = ttk.Frame(self._maps_tab)
        header.grid(row=0, column=0, sticky="ew")
        ttk.Label(header, text="Carte :").pack(side=tk.LEFT)
        self._map_selector = ttk.Combobox(header, state="readonly", width=34)
        self._map_selector.pack(side=tk.LEFT, padx=(6, 6))
        self._map_selector.bind("<<ComboboxSelected>>", self._on_map_selected)
        self._map_add_button = ttk.Button(header, image=self._icon_map_add, command=self._add_map)
        self._map_add_button.pack(side=tk.LEFT)
        self._attach_tooltip(self._map_add_button, "Ajouter une carte")
        self._map_archive_button = ttk.Button(header, image=self._icon_archive, command=self._toggle_archive_selected_map)
        self._map_archive_button.pack(side=tk.LEFT, padx=4)
        self._map_delete_button = ttk.Button(header, image=self._icon_trash, command=self._delete_selected_map)
        self._map_delete_button.pack(side=tk.LEFT)

        detail = ttk.Frame(self._maps_tab)
        detail.grid(row=1, column=0, sticky="ew", pady=(8, 10))
        detail.columnconfigure(2, weight=1)
        self._map_default_check = ttk.Checkbutton(detail, text="Défaut", variable=self._map_default_var,
                                                   command=self._set_selected_map_as_default)
        self._map_default_check.grid(row=0, column=0, sticky="w", padx=(0, 12))
        ttk.Label(detail, text="Description :").grid(row=0, column=1, sticky="w")
        self._map_description_entry = ttk.Entry(detail, textvariable=self._map_description_var)
        self._map_description_entry.grid(row=0, column=2, sticky="ew", padx=(6, 12))
        self._map_status_label = ttk.Label(detail, anchor="e")
        self._map_status_label.grid(row=0, column=3, sticky="e")
        self._map_role_label = ttk.Label(detail, foreground="#446")
        self._map_role_label.grid(row=1, column=2, columnspan=2, sticky="e", pady=(3, 0))
        self._map_description_var.trace_add("write", self._save_map_description)

        panes = ttk.PanedWindow(self._maps_tab, orient=tk.HORIZONTAL)
        panes.grid(row=2, column=0, sticky="nsew")
        cities = ttk.Frame(panes, padding=(0, 0, 8, 0))
        preview = ttk.Frame(panes, padding=(8, 0, 0, 0))
        panes.add(cities, weight=0)
        panes.add(preview, weight=1)
        cities.rowconfigure(1, weight=1)
        cities.columnconfigure(0, weight=1)
        toolbar = ttk.Frame(cities)
        toolbar.grid(row=0, column=0, sticky="ew", pady=(0, 5))
        self._calibration_city_add_button = ttk.Button(
            toolbar, image=self._icon_map_pin_plus, command=self._add_calibration_city
        )
        self._calibration_city_add_button.pack(side=tk.LEFT)
        self._attach_tooltip(self._calibration_city_add_button, "Ajouter une ville de calibration")
        self._calibration_city_remove_button = ttk.Button(
            toolbar, image=self._icon_trash, command=self._remove_calibration_city
        )
        self._calibration_city_remove_button.pack(side=tk.LEFT, padx=(5, 0))
        self._attach_tooltip(self._calibration_city_remove_button, "Supprimer la ville de calibration")
        self._calibration_city_tree = ttk.Treeview(cities, columns=("pointed", "city", "error"), show="headings", selectmode="browse")
        self._calibration_city_tree.heading("pointed", text="Pointé")
        self._calibration_city_tree.heading("city", text="Ville")
        self._calibration_city_tree.heading("error", text="Écart")
        self._calibration_city_tree.column("pointed", width=58, anchor="center", stretch=False)
        self._calibration_city_tree.column("city", width=180, anchor="w")
        self._calibration_city_tree.column("error", width=80, anchor="e", stretch=False)
        self._calibration_city_tree.grid(row=1, column=0, sticky="nsew")
        self._calibration_city_tree.bind("<<TreeviewSelect>>", self._on_calibration_city_selected)
        scroll = ttk.Scrollbar(cities, orient=tk.VERTICAL, command=self._calibration_city_tree.yview)
        scroll.grid(row=1, column=1, sticky="ns")
        self._calibration_city_tree.configure(yscrollcommand=scroll.set)

        ttk.Separator(cities, orient=tk.HORIZONTAL).grid(row=2, column=0, columnspan=2, sticky="ew", pady=(8, 6))
        placement = ttk.Frame(cities)
        placement.grid(row=3, column=0, columnspan=2, sticky="ew")
        placement.columnconfigure(2, weight=1)
        ttk.Label(placement, text="Scale").grid(row=0, column=0, sticky="w")
        self._map_scale_entry = ttk.Entry(placement, textvariable=self._map_scale_var, width=7)
        self._map_scale_entry.grid(row=0, column=1, sticky="w", padx=(6, 8))
        self._map_scale_entry.bind("<Return>", self._commit_map_scale_entry)
        self._map_scale_entry.bind("<FocusOut>", self._commit_map_scale_entry)
        self._map_scale_slider = ttk.Scale(
            placement, from_=0.0, to=20.0, orient=tk.HORIZONTAL,
            command=self._on_map_scale_slider_changed,
        )
        self._map_scale_slider.grid(row=0, column=2, sticky="ew")

        preview.rowconfigure(1, weight=1)
        preview.columnconfigure(0, weight=1)
        preview_toolbar = ttk.Frame(preview)
        preview_toolbar.grid(row=0, column=0, sticky="ew", pady=(0, 5))
        self._map_hand_button = tk.Button(
            preview_toolbar, image=self._icon_hand, command=lambda: self._set_map_interaction_mode("hand")
        )
        self._map_hand_button.pack(side=tk.LEFT)
        self._attach_tooltip(self._map_hand_button, "Mode déplacement et sélection")
        self._map_focus_button = tk.Button(
            preview_toolbar, image=self._icon_focus, command=lambda: self._set_map_interaction_mode("focus")
        )
        self._map_focus_button.pack(side=tk.LEFT, padx=(4, 0))
        self._attach_tooltip(self._map_focus_button, "Mode pointage calibration")
        self._calibration_map_view = GeoMapView(
            preview,
            on_marker_selected=self._on_calibration_marker_selected,
            maximum_zoom=_CALIBRATION_MAP_MAXIMUM_ZOOM,
        )
        self._calibration_map_view.grid(row=1, column=0, sticky="nsew")
        self._set_map_interaction_mode("hand")
        self.after_idle(lambda: panes.sashpos(0, max(250, int(panes.winfo_width() * .32))))

    def _build_books_tab(self) -> None:
        self._books_tab.rowconfigure(2, weight=1)
        self._books_tab.columnconfigure(0, weight=1)
        header = ttk.Frame(self._books_tab)
        header.grid(row=0, column=0, sticky="ew")
        ttk.Label(header, text="Livre :").pack(side=tk.LEFT)
        self._book_selector = ttk.Combobox(header, state="readonly", width=38)
        self._book_selector.pack(side=tk.LEFT, padx=(6, 6))
        self._book_selector.bind("<<ComboboxSelected>>", self._on_book_selected)
        self._book_add_button = ttk.Button(header, image=self._icon_book_add, command=self._add_book)
        self._book_add_button.pack(side=tk.LEFT)
        self._attach_tooltip(self._book_add_button, "Ajouter un livre")
        self._book_duplicate_button = ttk.Button(header, image=self._icon_duplicate, command=self._duplicate_selected_book)
        self._book_duplicate_button.pack(side=tk.LEFT, padx=4)
        self._attach_tooltip(self._book_duplicate_button, "Dupliquer le livre")
        self._book_archive_button = ttk.Button(header, image=self._icon_archive, command=self._archive_selected_book)
        self._book_archive_button.pack(side=tk.LEFT)
        self._attach_tooltip(self._book_archive_button, "Archiver le livre")
        self._book_delete_button = ttk.Button(header, image=self._icon_trash, command=self._delete_selected_book)
        self._book_delete_button.pack(side=tk.LEFT, padx=(4, 0))
        self._attach_tooltip(self._book_delete_button, "Détruire le livre")

        detail = ttk.Frame(self._books_tab)
        detail.grid(row=1, column=0, sticky="ew", pady=(8, 10))
        detail.columnconfigure(3, weight=1)
        self._book_default_check = ttk.Checkbutton(
            detail, text="Livre par défaut", variable=self._book_default_var,
            command=self._set_selected_book_as_default,
        )
        self._book_default_check.grid(row=0, column=0, sticky="w", padx=(0, 12))
        ttk.Label(detail, text="Nom :").grid(row=0, column=1, sticky="w")
        self._book_name_entry = ttk.Entry(detail, textvariable=self._book_name_var, width=22)
        self._book_name_entry.grid(row=0, column=2, sticky="ew", padx=(6, 12))
        ttk.Label(detail, text="Description :").grid(row=0, column=3, sticky="w")
        self._book_description_entry = ttk.Entry(detail, textvariable=self._book_description_var)
        self._book_description_entry.grid(row=0, column=4, sticky="ew", padx=(6, 0))
        detail.columnconfigure(4, weight=1)
        self._book_name_var.trace_add("write", self._save_book_metadata)
        self._book_description_var.trace_add("write", self._save_book_metadata)

        panes = ttk.PanedWindow(self._books_tab, orient=tk.HORIZONTAL)
        panes.grid(row=2, column=0, sticky="nsew")
        tags = ttk.Frame(panes, padding=(0, 0, 8, 0))
        grid = ttk.Frame(panes, padding=(8, 0, 0, 0))
        panes.add(tags, weight=0)
        panes.add(grid, weight=1)
        tags.rowconfigure(1, weight=1)
        tags.columnconfigure(0, weight=1)
        ttk.Label(tags, text="Tags").grid(row=0, column=0, sticky="w", pady=(0, 4))
        self._book_tags_listbox = tk.Listbox(tags, exportselection=False, selectmode=tk.BROWSE)
        self._book_tags_listbox.grid(row=1, column=0, sticky="nsew")
        tag_scroll = ttk.Scrollbar(tags, orient=tk.VERTICAL, command=self._book_tags_listbox.yview)
        tag_scroll.grid(row=1, column=1, sticky="ns")
        self._book_tags_listbox.configure(yscrollcommand=tag_scroll.set)
        self._book_tags_listbox.bind("<<ListboxSelect>>", self._on_book_tag_selected)
        grid.rowconfigure(0, weight=1)
        grid.columnconfigure(0, weight=1)
        self._book_sheet = Sheet(grid, data=[], headers=[], row_index=[], show_row_index=True)
        self._book_sheet.enable_bindings(("single_select", "row_select", "column_select", "arrowkeys", "copy"))
        self._book_sheet.pack(fill=tk.BOTH, expand=True)
        self._book_cell_tags: dict[tuple[int, int], str | None] = {}
        self._book_tags: tuple[str, ...] = ()

    def _get_selected_book(self) -> CatalogueBook | None:
        return self.catalogue.get_book(self._selected_book_id) if self._selected_book_id is not None else None

    def _refresh_books(self) -> None:
        books = [book for book in self.catalogue.iter_books() if not book.archived]
        self._book_selector_ids = [book.book_id for book in books]
        if self._selected_book_id not in self._book_selector_ids:
            self._selected_book_id = self.catalogue.default_book_id or (self._book_selector_ids[0] if self._book_selector_ids else None)
        selected = self._get_selected_book()
        self._updating_book_detail = True
        try:
            self._book_selector.configure(values=tuple(book.name for book in books))
            if selected is None:
                self._book_selector.set("")
                self._book_name_var.set("")
                self._book_description_var.set("")
            else:
                self._book_selector.current(self._book_selector_ids.index(selected.book_id))
                self._book_name_var.set(selected.name)
                self._book_description_var.set(selected.description)
            self._book_default_var.set(selected is not None and selected.book_id == self.catalogue.default_book_id)
        finally:
            self._updating_book_detail = False
        self._catalogue_notebook.tab(self._books_tab, text=f"Livres ({len(books)})")
        self._load_selected_book_source()
        self._update_book_action_buttons()

    def _on_book_selected(self, _event=None) -> None:
        index = self._book_selector.current()
        self._selected_book_id = self._book_selector_ids[index] if 0 <= index < len(self._book_selector_ids) else None
        self._refresh_books()
        self._update_context_actions()

    def _save_book_metadata(self, *_args) -> None:
        selected = self._get_selected_book()
        if self._updating_book_detail or selected is None or self._book_assets.is_readonly(selected):
            return
        try:
            self.catalogue.update_book(
                selected.book_id,
                name=self._book_name_var.get(),
                description=self._book_description_var.get(),
            )
        except ValueError:
            return
        self._mark_dirty()

    def _load_selected_book_source(self) -> None:
        selected = self._get_selected_book()
        self._book_tags_listbox.delete(0, tk.END)
        self._book_cell_tags.clear()
        self._book_tags = ()
        if selected is None:
            self._book_sheet.set_sheet_data([])
            return
        try:
            lines = parse_book_file(self._book_assets.asset_path_for(selected))
        except (OSError, ValueError) as exc:
            self._book_sheet.set_sheet_data([])
            messagebox.showerror("Livre", f"{selected.book_id} : {exc}", parent=self)
            return
        tags = tuple(sorted({token.tag for line in lines for token in line.tokens if token.tag is not None}))
        self._book_tags = tags
        for tag in tags:
            self._book_tags_listbox.insert(tk.END, f"[{tag}]")
        width = max((len(line.tokens) for line in lines), default=0)
        data = [[token.text for token in line.tokens] + [""] * (width - len(line.tokens)) for line in lines]
        self._book_sheet.set_sheet_data(data, reset_col_positions=True, reset_row_positions=True)
        self._book_sheet.headers([str(index) for index in range(1, width + 1)])
        self._book_sheet.row_index([line.title for line in lines])
        for row, line in enumerate(lines):
            for column, token in enumerate(line.tokens):
                self._book_cell_tags[(row, column)] = token.tag
        self._highlight_selected_book_tag()

    def _on_book_tag_selected(self, _event=None) -> None:
        self._highlight_selected_book_tag()

    def _highlight_selected_book_tag(self) -> None:
        self._book_sheet.dehighlight_all()
        selection = self._book_tags_listbox.curselection()
        if not selection:
            return
        tag = self._book_tags[selection[0]]
        cells = [cell for cell, token_tag in self._book_cell_tags.items() if token_tag == tag]
        if cells:
            self._book_sheet.highlight_cells(cells=cells, bg="#fff2a8")

    def _update_book_action_buttons(self) -> None:
        selected = self._get_selected_book()
        readonly = selected is None or self._book_assets.is_readonly(selected)
        self._book_name_entry.configure(state=tk.DISABLED if readonly else tk.NORMAL)
        self._book_description_entry.configure(state=tk.DISABLED if readonly else tk.NORMAL)
        self._book_default_check.configure(state=tk.NORMAL if selected is not None and not selected.archived else tk.DISABLED)
        self._book_duplicate_button.configure(state=tk.NORMAL if selected is not None else tk.DISABLED)
        self._book_archive_button.configure(state=tk.NORMAL if not readonly else tk.DISABLED, image=self._icon_archive_off if selected is not None and selected.archived else self._icon_archive)
        self._book_delete_button.configure(state=tk.NORMAL if selected is not None and not readonly and selected.book_id != self.catalogue.default_book_id else tk.DISABLED)
        self._book_add_button.configure(state=tk.NORMAL)

    def _set_selected_book_as_default(self) -> None:
        selected = self._get_selected_book()
        if selected is None:
            return
        self.catalogue.set_default_book(selected.book_id)
        self._refresh_books()
        self._mark_dirty()

    def _add_book(self) -> None:
        source = filedialog.askopenfilename(parent=self, title="Importer un livre", initialdir=self._paths.exports_dir, filetypes=[("Livres texte", "*.txt")])
        if not source:
            return
        try:
            self._selected_book_id = self._book_assets.stage_new_book(
                source,
                name=self._next_new_book_name(),
                description="",
            )
        except (OSError, ValueError) as exc:
            messagebox.showerror("Ajouter un livre", str(exc), parent=self)
            return
        self._refresh_books()
        self._mark_dirty()

    def _next_new_book_name(self) -> str:
        base = "Nouveau Livre"
        names = {book.name.casefold() for book in self.catalogue.books.values()}
        candidate = base
        suffix = 2
        while candidate.casefold() in names:
            candidate = f"{base} {suffix}"
            suffix += 1
        return candidate

    def _duplicate_selected_book(self) -> None:
        selected = self._get_selected_book()
        if selected is None:
            return
        try:
            self._selected_book_id = self._book_assets.stage_duplicate(selected)
        except (OSError, ValueError) as exc:
            messagebox.showerror("Dupliquer le livre", str(exc), parent=self)
            return
        self._refresh_books()
        self._mark_dirty()

    def _archive_selected_book(self) -> None:
        selected = self._get_selected_book()
        if selected is None:
            return
        try:
            self.catalogue.archive_book(selected.book_id)
        except ValueError as exc:
            messagebox.showerror("Archiver le livre", str(exc), parent=self)
            return
        self._selected_book_id = self.catalogue.default_book_id
        self._refresh_books()
        self._mark_dirty()

    def _delete_selected_book(self) -> None:
        selected = self._get_selected_book()
        if selected is None:
            return
        if self._book_assets.is_readonly(selected):
            messagebox.showerror(
                "Détruire le livre",
                "Les livres SYS sont consultables uniquement.",
                parent=self,
            )
            return
        if selected.book_id == self.catalogue.default_book_id:
            messagebox.showerror(
                "Détruire le livre",
                "Le livre par défaut ne peut pas être supprimé.",
                parent=self,
            )
            return
        if self._is_book_referenced is not None and self._is_book_referenced(selected.book_id):
            messagebox.showerror(
                "Détruire le livre",
                "Impossible de supprimer ce livre :\nil est utilisé par un scénario actuellement chargé.",
                parent=self,
            )
            return
        if not messagebox.askyesno("Détruire le livre", f"Détruire {selected.name} ?", parent=self):
            return
        try:
            self._book_assets.schedule_delete_asset(selected)
            self.catalogue.delete_book(selected.book_id)
        except ValueError as exc:
            messagebox.showerror("Détruire le livre", str(exc), parent=self)
            return
        self._selected_book_id = self.catalogue.default_book_id
        self._refresh_books()
        self._mark_dirty()

    def _import_selected_book(self) -> None:
        selected = self._get_selected_book()
        if selected is None or self._book_assets.is_readonly(selected):
            return
        source = filedialog.askopenfilename(parent=self, title="Importer un livre", initialdir=self._paths.exports_dir, filetypes=[("Livres texte", "*.txt")])
        if not source:
            return
        try:
            self._book_assets.stage_import(selected, source)
        except (OSError, ValueError) as exc:
            messagebox.showerror("Importer le livre", str(exc), parent=self)
            return
        self._load_selected_book_source()
        self._mark_dirty()

    def _export_selected_book(self) -> None:
        selected = self._get_selected_book()
        if selected is None:
            return
        destination = filedialog.asksaveasfilename(parent=self, title="Exporter le livre", initialdir=self._paths.exports_dir, defaultextension=".txt", initialfile=f"{selected.name}.txt", filetypes=[("Livres texte", "*.txt")])
        if not destination:
            return
        try:
            self._book_assets.export(selected, destination)
        except (OSError, ValueError) as exc:
            messagebox.showerror("Exporter le livre", str(exc), parent=self)

    def _get_selected_map(self):
        return self.catalogue.get_map(self._selected_map_id) if self._selected_map_id is not None else None

    def _refresh_maps(self, *, fit: bool = True):
        maps = list(self.catalogue.iter_maps())
        self._map_selector_ids = [item.map_id for item in maps]
        if self._selected_map_id not in self._map_selector_ids:
            self._selected_map_id = self.catalogue.default_map_id or (self._map_selector_ids[0] if self._map_selector_ids else None)
        selected = self._get_selected_map()
        self._updating_map_detail = True
        self._map_selector.configure(values=tuple(item.name for item in maps))
        if selected is not None:
            self._map_selector.current(self._map_selector_ids.index(selected.map_id))
            self._map_description_var.set(selected.description)
        else:
            self._map_selector.set("")
            self._map_description_var.set("")
        self._map_default_var.set(selected is not None and selected.map_id == self.catalogue.default_map_id)
        self._updating_map_detail = False
        self._refresh_map_placement_detail()
        self._catalogue_notebook.tab(self._maps_tab, text=f"Cartes ({len(maps)})")
        self._refresh_calibration_cities()
        self._load_selected_catalogue_map(fit=fit)
        self._update_map_action_buttons()

    def _on_map_selected(self, _event=None):
        index = self._map_selector.current()
        self._selected_map_id = self._map_selector_ids[index] if 0 <= index < len(self._map_selector_ids) else None
        self._selected_calibration_city_id = None
        self._set_map_interaction_mode("hand")
        self._refresh_maps(fit=True)
        self._update_context_actions()

    def _save_map_description(self, *_args):
        selected = self._get_selected_map()
        if self._updating_map_detail or selected is None or self._map_calibration.is_readonly(selected):
            return
        self.catalogue.update_map(selected.map_id, description=self._map_description_var.get())
        self._mark_dirty()

    def _refresh_map_placement_detail(self) -> None:
        selected = self._get_selected_map()
        self._updating_map_placement = True
        try:
            if selected is None:
                self._map_scale_var.set("")
                self._map_scale_slider.set(0.0)
                return
            self._map_scale_var.set(f"{selected.default_scale_factor:.2f}")
            self._map_scale_slider.set(selected.default_scale_factor)
        finally:
            self._updating_map_placement = False

    def _set_map_scale(self, value: float) -> None:
        selected = self._get_selected_map()
        if selected is None or self._map_calibration.is_readonly(selected):
            self._refresh_map_placement_detail()
            return
        self.catalogue.update_map(selected.map_id, default_scale_factor=value)
        self._refresh_map_placement_detail()
        self._mark_dirty()

    def _on_map_scale_slider_changed(self, value: str) -> None:
        if self._updating_map_placement:
            return
        try:
            self._set_map_scale(float(value))
        except ValueError as exc:
            messagebox.showerror("Scale", str(exc), parent=self)
            self._refresh_map_placement_detail()

    def _commit_map_scale_entry(self, _event=None) -> None:
        if self._updating_map_placement:
            return
        selected = self._get_selected_map()
        if selected is None:
            return
        if self._map_calibration.is_readonly(selected):
            self._refresh_map_placement_detail()
            return
        try:
            value = float(self._map_scale_var.get().strip())
            self.catalogue.update_map(selected.map_id, default_scale_factor=value)
        except ValueError as exc:
            messagebox.showerror("Scale", str(exc), parent=self)
            self._refresh_map_placement_detail()
            return
        self._refresh_map_placement_detail()
        self._mark_dirty()

    def _refresh_calibration_cities(self):
        tree = self._calibration_city_tree
        tree.delete(*tree.get_children())
        selected = self._get_selected_map()
        if selected is None:
            return
        points = self._map_calibration.points_for(selected)
        residuals = self._map_calibration.leave_one_out_residuals(selected)
        for city_id in selected.calibration_city_ids:
            city = self.catalogue.get_city(city_id)
            residual = residuals.get(city_id)
            error = "—" if residual is None else f"{residual.error_px:.1f} px"
            tree.insert("", tk.END, iid=city_id, values=("☑" if city_id in points else "☐", city.name, error))
        if self._selected_calibration_city_id in selected.calibration_city_ids:
            tree.selection_set(self._selected_calibration_city_id)

    def _on_calibration_city_selected(self, _event=None):
        selected = self._calibration_city_tree.selection()
        self._selected_calibration_city_id = selected[0] if selected else None
        if self._selected_calibration_city_id is not None:
            catalogue_map = self._get_selected_map()
            if catalogue_map is not None:
                points = self._map_calibration.points_for(catalogue_map)
                if self._selected_calibration_city_id in points:
                    self._calibration_map_view.set_selected_marker(
                        self._selected_calibration_city_id,
                        recenter=True,
                    )
        self._update_map_action_buttons()

    def _on_calibration_marker_selected(self, marker_id):
        if marker_id is None or not isinstance(marker_id, str):
            return
        catalogue_map = self._get_selected_map()
        if catalogue_map is None or marker_id not in catalogue_map.calibration_city_ids:
            return
        self._selected_calibration_city_id = marker_id
        self._calibration_city_tree.selection_set(marker_id)
        self._calibration_city_tree.focus(marker_id)
        self._calibration_city_tree.see(marker_id)
        self._update_map_action_buttons()

    def _load_selected_catalogue_map(self, *, fit: bool = True):
        selected = self._get_selected_map()
        if selected is None:
            return
        try:
            preview = self._map_calibration.preview_map(selected)
        except (OSError, ValueError) as exc:
            self._map_status_label.configure(text=f"Statut : {self._map_calibration.status_for(selected)} ({exc})")
            return
        view_map = getattr(self._calibration_map_view, "map", None)
        same_map = view_map is not None and view_map.map_id == preview.map_id
        self._calibration_map_view.set_map(preview, preserve_view=not fit and same_map)
        points = self._map_calibration.points_for(selected)
        residuals = self._map_calibration.leave_one_out_residuals(selected)
        projected_markers = []
        residual_lines = []
        observed_markers = []
        for city_id in selected.calibration_city_ids:
            point = points.get(city_id)
            if point is None:
                continue
            city = self.catalogue.get_city(city_id)
            projected = residuals.get(city_id)
            if projected is not None:
                projected_x = projected.predicted_x
                projected_y = projected.predicted_y
                error = projected.error_px
                dx = projected.dx
                dy = projected.dy
                projected_markers.append(
                    GeoMapPixelMarker(
                        marker_id=("projected", city_id), pixel_x=projected_x, pixel_y=projected_y,
                        fill_color="", outline_color="#6f7f8f", selectable=False,
                    )
                )
                if error > 1e-6:
                    residual_lines.append(
                        GeoMapPixelPolyline(((point.pixel_x, point.pixel_y), (projected_x, projected_y)))
                    )
                tooltip = f"Écart : {error:.1f} px\ndx : {dx:+.1f} px\ndy : {dy:+.1f} px"
            else:
                tooltip = None
            observed_markers.append(
                GeoMapPixelMarker(
                    marker_id=city_id, pixel_x=point.pixel_x, pixel_y=point.pixel_y,
                    label=city.name, always_show_label=True, fill_color="#d52b1e", tooltip=tooltip,
                )
            )
        self._calibration_map_view.set_markers(())
        self._calibration_map_view.set_pixel_polylines(residual_lines)
        self._calibration_map_view.set_pixel_markers([*projected_markers, *observed_markers])
        if fit and observed_markers:
            self._calibration_map_view.fit_to_pixel_bounds(
                [(marker.pixel_x, marker.pixel_y) for marker in observed_markers], margin=0.20
            )
        selected_city_id = getattr(self, "_selected_calibration_city_id", None)
        if selected_city_id in points:
            self._calibration_map_view.set_selected_marker(
                selected_city_id,
                recenter=False,
            )
        self._map_status_label.configure(text=f"Statut : {self._map_calibration.status_for(selected)}")

    def _update_map_action_buttons(self):
        selected = self._get_selected_map()
        readonly = selected is None or self._map_calibration.is_readonly(selected)
        state = tk.DISABLED if readonly else tk.NORMAL
        self._map_description_entry.configure(state=state)
        self._map_scale_entry.configure(state=state)
        self._map_scale_slider.configure(state=state)
        status = self._map_calibration.status_for(selected) if selected is not None else ""
        self._map_default_check.configure(state=tk.NORMAL if selected is not None and status == STATUS_VALID and not selected.archived else tk.DISABLED)
        self._map_archive_button.configure(state=state, image=self._icon_archive_off if selected is not None and selected.archived else self._icon_archive)
        self._map_delete_button.configure(state=tk.DISABLED)  # références de scénarios non encore inspectables
        self._calibration_city_add_button.configure(state=state)
        self._calibration_city_remove_button.configure(state=state if self._selected_calibration_city_id else tk.DISABLED)
        if self._selected_calibration_city_id is None and self._map_interaction_mode == "focus":
            self._set_map_interaction_mode("hand")
        self._set_map_interaction_mode(self._map_interaction_mode)
        self._map_add_button.configure(
            state=tk.NORMAL if isinstance(self.catalogue.id_provider, UserCatalogueIdProvider) else tk.DISABLED
        )
        self._map_role_label.configure(text="Carte de référence Catalogue" if selected is not None and selected.map_id == self.catalogue.catalogue_reference_map_id else "")

    def _set_selected_map_as_default(self):
        selected = self._get_selected_map()
        if selected is None or self._map_calibration.status_for(selected) != STATUS_VALID:
            return
        self.catalogue.set_default_map(selected.map_id)
        self._refresh_maps(fit=False)
        self._mark_dirty()

    def _add_calibration_city(self):
        selected = self._get_selected_map()
        if selected is None:
            return
        candidates = [city for city in self.catalogue.iter_cities() if not city.archived and city.city_id not in selected.calibration_city_ids]
        choice = CitySelectionDialog(self, candidates).show()
        if choice is None:
            return
        self._map_calibration.add_city(selected.map_id, choice)
        self._selected_calibration_city_id = choice
        self._refresh_maps(fit=True)
        self._mark_dirty()

    def _remove_calibration_city(self):
        selected = self._get_selected_map()
        city_id = self._selected_calibration_city_id
        if selected is None or city_id is None:
            return
        self._map_calibration.remove_city(selected.map_id, city_id)
        self._selected_calibration_city_id = None
        self._set_map_interaction_mode("hand")
        self._refresh_maps(fit=True)
        self._mark_dirty()

    def _set_map_interaction_mode(self, mode: str) -> None:
        if mode not in {"hand", "focus"}:
            raise ValueError(f"Mode d'interaction carte invalide : {mode!r}.")
        selected_map = self._get_selected_map()
        focus_allowed = (
            selected_map is not None
            and self._selected_calibration_city_id is not None
            and not self._map_calibration.is_readonly(selected_map)
        )
        if mode == "focus" and not focus_allowed:
            mode = "hand"
        self._map_interaction_mode = mode
        self._calibration_map_view.set_map_click_handler(
            self._on_map_focus_click if mode == "focus" else None
        )
        self._map_hand_button.configure(relief=tk.SUNKEN if mode == "hand" else tk.RAISED, state=tk.NORMAL)
        self._map_focus_button.configure(
            relief=tk.SUNKEN if mode == "focus" else tk.RAISED,
            state=tk.NORMAL if focus_allowed else tk.DISABLED,
        )

    def _on_map_focus_click(self, pixel):
        selected = self._get_selected_map()
        city_id = self._selected_calibration_city_id
        if selected is None or city_id is None:
            messagebox.showinfo("Pointage calibration", "Sélectionnez d'abord une ville de calibration.", parent=self)
            return
        self._map_calibration.set_pixel(selected.map_id, city_id, *pixel)
        self._refresh_maps(fit=False)
        self._mark_dirty()

    def _toggle_archive_selected_map(self):
        selected = self._get_selected_map()
        if selected is None:
            return
        self.catalogue.update_map(selected.map_id, archived=not selected.archived)
        self._refresh_maps(fit=False)
        self._mark_dirty()

    def _delete_selected_map(self):
        messagebox.showinfo("Supprimer une carte", "La suppression reste désactivée tant que les références de scénarios ne sont pas contrôlées.", parent=self)

    def _add_map(self):
        image_path = filedialog.askopenfilename(parent=self, title="Image de la carte", initialdir=self._paths.exports_dir, filetypes=[("Images", "*.jpg *.jpeg *.png")])
        if not image_path:
            return
        name = simpledialog.askstring("Ajouter une carte", "Nom de la carte :", parent=self)
        if name is None:
            return
        description = simpledialog.askstring("Ajouter une carte", "Description :", parent=self) or ""
        try:
            self._selected_map_id = self._map_calibration.stage_user_map(image_path, name=name, description=description)
        except (OSError, ValueError) as exc:
            messagebox.showerror("Ajouter une carte", str(exc), parent=self)
            return
        self._selected_calibration_city_id = None
        self._refresh_maps(fit=True)
        self._mark_dirty()

    def _build_template_triangle_source(self, parent):
        parent.rowconfigure(1, weight=1)
        parent.columnconfigure(0, weight=1)
        filters = ttk.Frame(parent)
        filters.grid(row=0, column=0, sticky="ew", pady=(0, 4))
        ttk.Label(filters, text="Note :").grid(row=0, column=0, sticky="w")
        self._template_note_filter = ttk.Combobox(
            filters, textvariable=self._template_note_filter_var, values=("Tous",), state="readonly", width=5
        )
        self._template_note_filter.grid(row=0, column=1, padx=(0, 6))
        ttk.Label(filters, text="Base :").grid(row=0, column=2, sticky="w")
        ttk.Entry(filters, textvariable=self._template_base_filter_var, width=8).grid(
            row=0, column=3, padx=(0, 6)
        )
        ttk.Label(filters, text="Lumière :").grid(row=0, column=4, sticky="w")
        ttk.Entry(filters, textvariable=self._template_light_filter_var, width=9).grid(
            row=0, column=5
        )
        self._template_note_filter.bind("<<ComboboxSelected>>", lambda _event: self._refresh_template_triangle_tree())
        self._template_base_filter_var.trace_add("write", lambda *_: self._refresh_template_triangle_tree())
        self._template_light_filter_var.trace_add("write", lambda *_: self._refresh_template_triangle_tree())

        self._template_triangle_tree = ttk.Treeview(parent, show="tree", selectmode="extended")
        self._template_triangle_tree.heading("#0", text="Triangles")
        self._template_triangle_tree.column("#0", stretch=True, minwidth=180)
        self._template_triangle_tree.grid(row=1, column=0, sticky="nsew")
        scrollbar = ttk.Scrollbar(parent, orient=tk.VERTICAL, command=self._template_triangle_tree.yview)
        scrollbar.grid(row=1, column=1, sticky="ns")
        self._template_triangle_tree.configure(yscrollcommand=scrollbar.set)
        self._template_triangle_by_tree_iid: dict[str, str] = {}
        self._template_triangle_tree.bind("<<TreeviewSelect>>", self._on_template_triangle_selected)
        self._template_triangle_tree.bind("<ButtonPress-1>", self._on_template_tree_press, add="+")
        self._template_triangle_tree.bind("<B1-Motion>", self._on_template_tree_drag, add="+")
        self._template_triangle_tree.bind("<ButtonRelease-1>", self._on_template_tree_release, add="+")

    def _build_template_ranks_grid(self, parent):
        parent.rowconfigure(0, weight=1)
        parent.columnconfigure(0, weight=1)
        canvas = tk.Canvas(parent, highlightthickness=0)
        canvas.grid(row=0, column=0, sticky="nsew")
        scrollbar = ttk.Scrollbar(parent, orient=tk.VERTICAL, command=canvas.yview)
        scrollbar.grid(row=0, column=1, sticky="ns")
        canvas.configure(yscrollcommand=scrollbar.set)
        grid = ttk.Frame(canvas, padding=(0, 0, 4, 0))
        self._template_ranks_grid = grid
        grid_window = canvas.create_window((0, 0), window=grid, anchor="nw")
        grid.bind("<Configure>", lambda _event: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>", lambda event: self._resize_template_ranks_columns(canvas, grid_window, event.width))
        for column in range(3):
            grid.columnconfigure(column, weight=0)
        for column, text in enumerate(("Base", "Rang impair", "Rang pair")):
            ttk.Label(grid, text=text, anchor="center", padding=(6, 5), relief=tk.RIDGE, font=(None, 9, "bold")).grid(
                row=0, column=column, sticky="nsew"
            )
        self._template_pair_rows: list[TemplatePairRow] = []
        for pair_number in range(1, 17):
            pair_row = TemplatePairRow(
                grid,
                pair_number,
                self._clear_template_pair,
            )
            pair_row.grid(row=pair_number, column=0, columnspan=3, sticky="ew", pady=(3, 0))
            self._template_pair_rows.append(pair_row)
            self._attach_tooltip(pair_row.base_slot.clear_button, "Vider ce couple")
            self._attach_tooltip(pair_row.base_slot._entry, "")
            for slot in (pair_row.odd_rank_slot, pair_row.even_rank_slot):
                self._bind_template_slot(slot)
        self.bind("<Delete>", self._delete_selected_template_rank, add="+")
        self.bind("<BackSpace>", self._delete_selected_template_rank, add="+")
        self.after_idle(lambda: self._resize_template_ranks_columns(canvas, grid_window, canvas.winfo_width()))

    def _resize_template_ranks_columns(self, canvas, grid_window, width: int):
        canvas.itemconfigure(grid_window, width=width)
        base_width = min(_TEMPLATE_BASE_COLUMN_WIDTH, max(90, width - 220))
        rank_width = max(110, (max(width, base_width + 220) - base_width) // 2)
        for column, minimum in enumerate((base_width, rank_width, rank_width)):
            self._template_ranks_grid.columnconfigure(column, minsize=minimum, weight=0)
        for pair_row in getattr(self, "_template_pair_rows", ()):
            pair_row.set_column_widths(base_width, rank_width)

    def _refresh_templates(self):
        templates = list(self.catalogue.iter_templates())
        self._template_selector_ids = [template.template_id for template in templates]
        self._template_selector.configure(values=tuple(template.name for template in templates))
        if self._selected_template_id not in self._template_selector_ids:
            self._selected_template_id = self.catalogue.default_template_id or (self._template_selector_ids[0] if self._template_selector_ids else None)
        template = self._get_selected_template()
        self._updating_template_detail = True
        if template is not None:
            self._template_selector.current(self._template_selector_ids.index(template.template_id))
            self._template_name_var.set(template.name)
        else:
            self._template_selector.set("")
            self._template_name_var.set("")
        self._template_default_var.set(template is not None and template.template_id == self.catalogue.default_template_id)
        self._template_description_var.set(template.description if template else "")
        self._template_description_entry.configure(state=tk.NORMAL if template else tk.DISABLED)
        self._updating_template_detail = False
        self._catalogue_notebook.tab(self._templates_tab, text=f"Templates ({len(templates)})")
        self._update_template_action_buttons()
        self._refresh_template_ranks_view()

    def _on_template_selected(self, _event=None):
        if self._selected_template_rank_slot is not None:
            self._selected_template_rank_slot.set_selected(False)
            self._selected_template_rank_slot = None
        selected_index = self._template_selector.current()
        self._selected_template_id = self._template_selector_ids[selected_index] if 0 <= selected_index < len(self._template_selector_ids) else None
        self._refresh_templates()
        self._update_context_actions()

    def _get_selected_template(self) -> HypothesisTemplate | None:
        return self.catalogue.get_template(self._selected_template_id) if self._selected_template_id is not None else None

    def _commit_template_name(self):
        template = self._get_selected_template()
        if template is None:
            return
        new_name = self._template_name_var.get().strip()
        if not new_name:
            messagebox.showerror("Renommer le template", "Le nom du template ne peut pas être vide.", parent=self)
            self._template_name_var.set(template.name)
            return
        if new_name == template.name:
            self._template_name_var.set(template.name)
            return
        try:
            self.catalogue.update_template(template.template_id, name=new_name)
        except ValueError as exc:
            messagebox.showerror("Renommer le template", str(exc), parent=self)
            self._template_name_var.set(template.name)
            return
        self._refresh_templates()
        self._mark_dirty()

    def _save_template_description(self, *_args):
        template = self._get_selected_template()
        if self._updating_template_detail or template is None:
            return
        self.catalogue.update_template(template.template_id, description=self._template_description_var.get())
        self._mark_dirty()

    def _add_template(self):
        names = {template.name.casefold() for template in self.catalogue.iter_templates()}
        number = 1
        while f"Nouveau template {number}".casefold() in names:
            number += 1
        new_template = self.catalogue.add_template(f"Nouveau template {number}")
        self._selected_template_id = new_template.template_id
        self._refresh_templates()
        self._update_context_actions()
        self._mark_dirty()

    def _duplicate_selected_template(self):
        source = self._get_selected_template()
        if source is None:
            return

        existing_names = {
            template.name.casefold()
            for template in self.catalogue.iter_templates()
        }

        base_name = f"{source.name} - copie"
        new_name = base_name

        number = 2
        while new_name.casefold() in existing_names:
            new_name = f"{base_name} {number}"
            number += 1

        duplicate = self.catalogue.add_template(
            new_name,
            description=source.description,
        )

        self.catalogue.set_template_ranks(
            duplicate.template_id,
            list(source.triangle_ids_by_rank),
        )

        self._selected_template_id = duplicate.template_id
        self._selected_template_rank_slot = None

        self._refresh_templates()
        self._update_context_actions()
        self._mark_dirty()

    def _toggle_archive_selected_template(self):
        template = self._get_selected_template()
        if template is None:
            return
        self.catalogue.update_template(template.template_id, archived=not template.archived)
        self._refresh_templates()
        self._mark_dirty()

    def _delete_selected_template(self):
        template = self._get_selected_template()
        if template is None:
            return
        self.catalogue.delete_template(template.template_id)
        self._selected_template_id = self.catalogue.default_template_id or next(iter(self.catalogue.templates), None)
        self._refresh_templates()
        self._update_context_actions()
        self._mark_dirty()

    def _update_template_action_buttons(self):
        template = self._get_selected_template()
        state = tk.NORMAL if template is not None else tk.DISABLED
        self._template_duplicate_button.configure(state=state)
        self._template_default_button.configure(
            state=tk.NORMAL if template is not None and template.template_id != self.catalogue.default_template_id else tk.DISABLED
        )
        self._template_delete_button.configure(state=state)
        self._template_archive_button.configure(
            state=state,
            image=self._icon_archive_off if template is not None and template.archived else self._icon_archive,
        )
        self._set_tooltip_text(
            self._template_archive_button,
            "Désarchiver le template" if template is not None and template.archived else "Archiver le template",
        )

    def _set_selected_template_as_default(self):
        template = self._get_selected_template()
        if template is None or template.template_id == self.catalogue.default_template_id:
            return
        self.catalogue.set_default_template(template.template_id)
        self._refresh_templates()
        self._mark_dirty()

    def _refresh_template_ranks_view(self):
        template = self._get_selected_template()
        ranks = template.triangle_ids_by_rank if template else [None] * 32
        for pair, pair_row in enumerate(self._template_pair_rows):
            odd_id, even_id = ranks[pair * 2], ranks[pair * 2 + 1]
            odd = self.catalogue.get_triangle(odd_id) if odd_id is not None else None
            even = self.catalogue.get_triangle(even_id) if even_id is not None else None
            odd_light = self.catalogue.get_city(odd.light_city_id).name if odd else None
            even_light = self.catalogue.get_city(even.light_city_id).name if even else None
            bases = {self.catalogue.get_city(item.base_city_id).name for item in (odd, even) if item is not None}
            base_name = bases.pop() if len(bases) == 1 else (None if not bases else "Bases différentes")
            pair_row.set_triangles(odd_id, odd_light, even_id, even_light, base_name)
            self._set_tooltip_text(pair_row.base_slot._entry, pair_row.base_slot.tooltip_text)
        if self._selected_template_rank_slot is not None:
            self._selected_template_rank_slot.set_selected(True)
        self._update_template_validation_status()

    def _update_template_validation_status(self):
        template = self._get_selected_template()
        if template is None:
            self._template_validation_status_label.configure(text="Statut :")
            return
        status = self.catalogue.get_template_validation_status(template.template_id)
        text = (
            f"Statut : Incomplet ({status.filled_ranks} / 32)"
            if status.state == "Incomplet"
            else f"Statut : {status.state}"
        )
        self._template_validation_status_label.configure(text=text)

    def _bind_template_slot(self, slot: TemplateRankSlot):
        for widget in (slot, slot._badge, slot._entry):
            widget.bind("<ButtonPress-1>", lambda _event, target=slot: self._select_template_rank_slot(target), add="+")
        self._attach_tooltip(slot._entry, "")

    def _select_template_rank_slot(self, slot: TemplateRankSlot):
        if self._selected_template_rank_slot is not None and self._selected_template_rank_slot is not slot:
            self._selected_template_rank_slot.set_selected(False)
        self._selected_template_rank_slot = slot
        slot.set_selected(True)
        slot.focus_set()
        if slot.triangle_id is None:
            return
        iid = next(
            (tree_iid for tree_iid, triangle_id in self._template_triangle_by_tree_iid.items() if triangle_id == slot.triangle_id),
            None,
        )
        if iid is None:
            return
        parent_iid = self._template_triangle_tree.parent(iid)
        if parent_iid:
            self._template_triangle_tree.item(parent_iid, open=True)
        self._template_triangle_tree.selection_set(iid)
        self._template_triangle_tree.focus(iid)
        self._template_triangle_tree.see(iid)

    def _delete_selected_template_rank(self, _event=None):
        slot = self._selected_template_rank_slot
        template = self._get_selected_template()
        if template is None or slot is None:
            return
        rank_index = slot.rank_number - 1
        if template.triangle_ids_by_rank[rank_index] is None:
            return
        preview = list(template.triangle_ids_by_rank)
        preview[rank_index] = None
        self.catalogue.set_template_ranks(template.template_id, preview)
        self._refresh_template_ranks_view()
        self._mark_dirty()

    def _clear_template_pair(self, pair_row: TemplatePairRow):
        template = self._get_selected_template()
        if template is None:
            return
        odd_index = pair_row.pair_number * 2 - 2
        even_index = odd_index + 1
        if template.triangle_ids_by_rank[odd_index] is None and template.triangle_ids_by_rank[even_index] is None:
            return
        preview = list(template.triangle_ids_by_rank)
        preview[odd_index] = preview[even_index] = None
        self.catalogue.set_template_ranks(template.template_id, preview)
        self._refresh_template_ranks_view()
        self._mark_dirty()

    def _set_dirty(self, is_dirty: bool) -> None:
        self._is_dirty = is_dirty
        state = tk.NORMAL if is_dirty else tk.DISABLED
        self._apply_button.configure(state=state)
        self._cancel_button.configure(state=state)

    def _mark_dirty(self) -> None:
        self._set_dirty(True)

    def _replace_working_catalogue(self, catalogue: Catalogue) -> None:
        """Remplace le clone de travail en conservant sa transaction cartographique."""
        self.catalogue = catalogue
        self._map_calibration.rebind_catalogue(catalogue)
        self._book_assets.rebind_catalogue(catalogue)

    def _apply_changes(self) -> None:
        """Valide uniquement l'état courant de cette session, sans persistance."""
        created_assets: list[Path] = []
        created_book_assets: list[Path] = []
        try:
            self.catalogue.validate()
            created_assets = self._map_calibration.commit()
            created_book_assets = self._book_assets.commit()
            save_catalogue(self.catalogue, self._catalogue_path)
        except (OSError, ValueError, TypeError) as exc:
            self._map_calibration.rollback(created_assets)
            self._book_assets.rollback(created_book_assets)
            messagebox.showerror("Catalogue", str(exc), parent=self)
            return
        if self._on_catalogue_applied is not None:
            self._on_catalogue_applied(self.catalogue.clone())
        self._validated_catalogue = self.catalogue.clone()
        self._map_calibration.finalize_commit()
        self._book_assets.finalize_commit()
        self._set_dirty(False)

    def _cancel_changes(self) -> None:
        """Restaure le dernier instantané validé localement."""
        if self._selected_template_rank_slot is not None:
            self._selected_template_rank_slot.set_selected(False)
        self._map_calibration.discard()
        self._book_assets.discard()
        self._replace_working_catalogue(self._validated_catalogue.clone())
        self._selected_city_id = None
        self._selected_beacon_id = None
        self._selected_triangle_id = None
        self._selected_template_id = self.catalogue.default_template_id
        self._selected_map_id = self.catalogue.default_map_id
        self._selected_book_id = self.catalogue.default_book_id
        self._selected_calibration_city_id = None
        self._selected_template_triangle_id = None
        self._selected_template_rank_slot = None
        self._refresh_city_list()
        self._refresh_beacon_list()
        self._refresh_triangle_tree()
        self._refresh_maps()
        self._refresh_books()
        self._refresh_templates()
        self._update_context_actions()
        self._load_selected_city()
        self._set_dirty(False)

    def request_close(self) -> bool:
        """Ferme après confirmation lorsqu'un état local non validé existe."""
        if self._is_dirty:
            if not messagebox.askyesno(
                "Gestion du catalogue",
                "Des modifications ne sont pas appliquées. Fermer sans les appliquer ?",
                parent=self,
            ):
                return False
            self._map_calibration.discard()
            self._book_assets.discard()
        self.destroy()
        return True

    def _import_csv(self):
        path = filedialog.askopenfilename(parent=self, title="Importer des villes", initialdir=self._paths.exports_dir,
                                          filetypes=[("Fichiers CSV", "*.csv"), ("Tous les fichiers", "*.*")])
        if not path:
            return
        try:
            imported_cities, errors = self._read_cities_csv(path)
            preview = self.catalogue.clone()
            cities_by_name = {
                city.name.strip().casefold(): city
                for city in preview.iter_cities()
            }
            imported_ids: list[str] = []
            updated_ids: list[str] = []
            unchanged_count = 0
            for name, latitude, longitude in imported_cities:
                key = name.strip().casefold()
                city = cities_by_name.get(key)
                if city is None:
                    city = preview.add_city(name, latitude, longitude)
                    cities_by_name[key] = city
                    imported_ids.append(city.city_id)
                elif math.isclose(city.latitude, latitude, abs_tol=1e-9) and math.isclose(
                    city.longitude, longitude, abs_tol=1e-9,
                ):
                    unchanged_count += 1
                else:
                    preview.update_city(city.city_id, latitude=latitude, longitude=longitude)
                    updated_ids.append(city.city_id)
            self._replace_working_catalogue(preview)
        except (OSError, UnicodeError, ValueError, csv.Error) as exc:
            messagebox.showerror("Importer des villes", str(exc), parent=self)
            return
        result = CityCsvImportResult(
            tuple(imported_ids), tuple(updated_ids), unchanged_count, tuple(errors),
        )
        if result.imported_count or result.updated_count:
            self._refresh_city_list()
            self._refresh_triangle_tree()
            self._refresh_beacon_list()
            self._mark_dirty()
        summary = self._format_cities_import_summary(result)
        if result.errors:
            messagebox.showwarning("Importer des villes", summary, parent=self)
        else:
            messagebox.showinfo("Importer des villes", summary, parent=self)

    def _import_beacons_xlsx(self):
        path = filedialog.askopenfilename(
            parent=self,
            title="Importer des balises",
            initialdir=self._paths.exports_dir,
            filetypes=[("Fichiers Excel", "*.xlsx"), ("Tous les fichiers", "*.*")],
        )
        if not path:
            return
        try:
            rows = self._read_beacons_xlsx(path)
            self._import_beacon_rows(rows)
        except (OSError, ValueError) as exc:
            messagebox.showerror("Importer des balises", str(exc), parent=self)
            return
        self._refresh_beacon_list()
        self._refresh_city_list()
        self._load_selected_city()
        self._mark_dirty()

    def _import_beacon_rows(self, rows: list[tuple[int, str]]) -> None:
        """Applique l'import dans un clone, sans mutation partielle en cas d'erreur."""
        preview = self.catalogue.clone()
        for line_number, city_id in rows:
            try:
                preview.add_beacon(city_id)
            except ValueError as exc:
                city_name = self.catalogue.get_city(city_id).name
                raise ValueError(f"Ligne {line_number} : {city_name} : {exc}") from exc
        self._replace_working_catalogue(preview)

    def _beacon_export_names(self) -> tuple[str, ...]:
        """Exporte toutes les balises, y compris archivées, dans l'ordre des IDs."""
        return tuple(
            self.catalogue.get_city(beacon.city_id).name
            for beacon in self.catalogue.iter_beacons()
        )

    def _read_beacons_xlsx(self, path: str) -> list[tuple[int, str]]:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            worksheet = workbook.active
            rows = [
                (line_number, row)
                for line_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1)
                if any(value is not None and str(value).strip() for value in row)
            ]
        finally:
            workbook.close()
        if not rows:
            raise ValueError("Le fichier Excel balises est vide.")
        _header_line, header = rows[0]
        header_values = tuple(str(value).strip() for value in header if value is not None and str(value).strip())
        if header_values != self._BEACON_XLSX_HEADER:
            raise ValueError("L'en-tête Excel doit contenir exactement la colonne : Nom")
        cities_by_name = {city.name.strip().casefold(): city for city in self.catalogue.iter_cities()}
        imported: list[tuple[int, str]] = []
        for line_number, row in rows[1:]:
            values = tuple(str(value).strip() for value in row if value is not None and str(value).strip())
            if len(values) != 1:
                raise ValueError(f"Ligne {line_number} : une seule colonne Nom est attendue.")
            city = cities_by_name.get(values[0].casefold())
            if city is None:
                raise ValueError(f"Ligne {line_number} : ville inconnue : {values[0]}.")
            imported.append((line_number, city.city_id))
        return imported

    @classmethod
    def _read_cities_csv(cls, path: str) -> tuple[list[tuple[str, float, float]], list[str]]:
        with open(path, "r", encoding="utf-8-sig", newline="") as csv_file:
            rows = [row for row in csv.reader(csv_file, delimiter=";") if any(str(value).strip() for value in row)]
        if not rows:
            raise ValueError("Le fichier CSV est vide.")
        if tuple(str(value).strip() for value in rows[0]) != cls._CITY_CSV_HEADER:
            raise ValueError("L'en-tête CSV doit être : Nom;Latitude;Longitude")
        cities = []
        errors = []
        for line_number, row in enumerate(rows[1:], start=2):
            try:
                if len(row) != 3:
                    raise ValueError("trois colonnes sont attendues.")
                name, raw_latitude, raw_longitude = (str(value).strip() for value in row)
                if not name:
                    raise ValueError("nom de ville vide.")
                cities.append((
                    name,
                    DmsCoordinateEditor.parse_coordinate(raw_latitude, "latitude"),
                    DmsCoordinateEditor.parse_coordinate(raw_longitude, "longitude"),
                ))
            except ValueError as exc:
                errors.append(f"Ligne {line_number} : {exc}")
        return cities, errors

    @staticmethod
    def _format_cities_import_summary(result: CityCsvImportResult) -> str:
        lines = ["Import terminé.", ""]
        if result.imported_count:
            lines.append(f"{result.imported_count} nouvelle(s) ville(s) importée(s).")
        if result.updated_count:
            lines.append(f"{result.updated_count} ville(s) mise(s) à jour.")
        if result.unchanged_count:
            lines.append(f"{result.unchanged_count} ville(s) déjà à jour ont été ignorée(s).")
        if not (result.imported_count or result.updated_count or result.unchanged_count):
            lines.append("Aucune ville importée.")
        if result.errors:
            lines.extend(("", f"{len(result.errors)} ligne(s) n'ont pas pu être importée(s) :", "\n".join(result.errors[:12])))
        return "\n".join(lines)

    def _import_triangles_csv(self):
        path = filedialog.askopenfilename(
            parent=self,
            title="Importer des triangles",
            initialdir=self._paths.exports_dir,
            filetypes=[("Fichiers CSV", "*.csv"), ("Tous les fichiers", "*.*")],
        )
        if not path:
            return
        try:
            result = self._read_triangles_csv(path)
        except (OSError, UnicodeError, ValueError, csv.Error) as exc:
            messagebox.showerror("Importer des triangles", str(exc), parent=self)
            return
        self._refresh_triangle_tree()
        if result.imported_count:
            self._mark_dirty()
        summary = self._format_triangles_import_summary(result)
        if result.errors:
            messagebox.showwarning(
                "Import des triangles",
                summary,
                parent=self,
            )
        else:
            messagebox.showinfo("Import des triangles", summary, parent=self)

    @staticmethod
    def _format_triangles_import_summary(result: TriangleCsvImportResult) -> str:
        lines = ["Import terminé.", ""]
        if result.imported_count:
            suffix = "" if result.imported_count == 1 else "s"
            lines.append(f"{result.imported_count} nouveau triangle importé{suffix}.")
        elif result.already_present_count or result.errors:
            lines.append("Aucun nouveau triangle.")
        if result.already_present_count:
            suffix = "" if result.already_present_count == 1 else "s"
            lines.append(
                f"{result.already_present_count} triangle{suffix} déjà présent{suffix} "
                "ont été ignorés."
                if result.already_present_count != 1
                else "1 triangle déjà présent a été ignoré."
            )
        if result.errors:
            preview = "\n".join(result.errors[:12])
            suffix = "\n…" if len(result.errors) > 12 else ""
            lines.extend(("", f"{len(result.errors)} ligne(s) n'ont pas pu être importée(s) :", preview + suffix))
        return "\n".join(lines)

    def _read_triangles_csv(self, path: str) -> TriangleCsvImportResult:
        with open(path, "r", encoding="utf-8-sig", newline="") as csv_file:
            rows = [row for row in csv.reader(csv_file, delimiter=";") if any(str(value).strip() for value in row)]
        if not rows:
            raise ValueError("Le fichier CSV triangles est vide.")
        if tuple(str(value).strip() for value in rows[0]) != self._TRIANGLE_CSV_HEADER:
            raise ValueError("L'en-tête CSV doit être : Note;Ouverture;Base;Lumiere")
        cities_by_name = {city.name.strip().casefold(): city for city in self.catalogue.iter_cities()}
        existing_triplets = {
            (triangle.opening_city_id, triangle.base_city_id, triangle.light_city_id)
            for triangle in self.catalogue.iter_triangles()
        }
        triangles: list[str] = []
        already_present_count = 0
        errors: list[str] = []
        for line_number, row in enumerate(rows[1:], start=2):
            if len(row) != 4:
                errors.append(f"Ligne {line_number} : quatre colonnes sont attendues.")
                continue
            note, opening_name, base_name, light_name = (str(value).strip() for value in row)
            cities = [cities_by_name.get(name.casefold()) for name in (opening_name, base_name, light_name)]
            if not note or any(city is None for city in cities):
                missing = [name for name, city in zip((opening_name, base_name, light_name), cities) if city is None]
                detail = f" ville(s) inconnue(s) : {', '.join(missing)}" if missing else " note vide"
                errors.append(f"Ligne {line_number} :{detail}.")
                continue
            opening, base, light = cities
            triplet = (opening.city_id, base.city_id, light.city_id)
            if triplet in existing_triplets:
                already_present_count += 1
                continue
            try:
                model_triangle = self.catalogue.add_triangle(
                    note, opening.city_id, base.city_id, light.city_id,
                )
            except (ValueError, KeyError) as exc:
                errors.append(f"Ligne {line_number} : {exc}")
                continue
            triangles.append(model_triangle.triangle_id)
            existing_triplets.add(triplet)
        return TriangleCsvImportResult(
            imported_triangle_ids=tuple(triangles),
            already_present_count=already_present_count,
            errors=tuple(errors),
        )

    def _import_template_csv(self):
        template = self._get_selected_template()
        if template is None:
            return
        path = filedialog.askopenfilename(
            parent=self,
            title="Importer un template",
            initialdir=self._paths.exports_dir,
            filetypes=[("Fichiers CSV", "*.csv"), ("Tous les fichiers", "*.*")],
        )
        if not path:
            return
        filled_count = sum(triangle_id is not None for triangle_id in template.triangle_ids_by_rank)
        if filled_count and not messagebox.askyesno(
            "Importer un template",
            (f"Le template courant contient déjà {filled_count} rang(s) renseigné(s).\n\n"
             "L’import remplacera entièrement les 32 rangs actuels.\n\nContinuer ?"),
            parent=self,
        ):
            return
        try:
            imported_ranks = self._read_template_csv(path)
        except (OSError, UnicodeError, ValueError, csv.Error) as exc:
            messagebox.showerror("Importer un template", str(exc), parent=self)
            return
        self.catalogue.set_template_ranks(template.template_id, imported_ranks)
        self._refresh_template_ranks_view()
        self._mark_dirty()

    def _read_template_csv(self, path: str) -> list[str | None]:
        with open(path, "r", encoding="utf-8-sig", newline="") as csv_file:
            rows = [
                (line_number, row)
                for line_number, row in enumerate(csv.reader(csv_file, delimiter=";"), start=1)
                if any(str(value).strip() for value in row)
            ]
        if not rows or tuple(str(value).strip() for value in rows[0][1]) != self._TEMPLATE_CSV_HEADER:
            raise ValueError("L’en-tête CSV doit être : Rang;Ouverture;Base;Lumiere")
        data_rows = rows[1:]
        if len(data_rows) != 32:
            raise ValueError("Le fichier doit contenir exactement 32 lignes de rang.")

        triangles_by_key: dict[tuple[str, str, str], list[str]] = {}
        for triangle in self.catalogue.iter_triangles():
            if not triangle.archived:
                opening, base, light = self._model_triangle_cities(triangle)
                key = tuple(
                    city.name.strip().casefold()
                    for city in (opening, base, light)
                )
                triangles_by_key.setdefault(key, []).append(triangle.triangle_id)

        imported_ranks: list[str | None] = [None] * 32
        seen_ranks: set[int] = set()
        for line_number, row in data_rows:
            if len(row) != 4:
                raise ValueError(f"Ligne {line_number} : quatre colonnes sont attendues.")
            raw_rank, raw_opening, raw_base, raw_light = (str(value).strip() for value in row)
            try:
                rank = int(raw_rank)
            except ValueError:
                raise ValueError(f"Ligne {line_number} : rang invalide « {raw_rank} ».") from None
            if not 1 <= rank <= 32:
                raise ValueError(f"Ligne {line_number} : le rang doit être compris entre 1 et 32.")
            if rank in seen_ranks:
                raise ValueError(f"Ligne {line_number} : le rang {rank} est présent plusieurs fois.")
            seen_ranks.add(rank)
            key = (raw_opening.casefold(), raw_base.casefold(), raw_light.casefold())
            matches = triangles_by_key.get(key, [])
            display_key = f"{raw_opening} / {raw_base} / {raw_light}"
            if not matches:
                raise ValueError(f"Ligne {line_number} : aucun triangle actif ne correspond à :\n{display_key}.")
            if len(matches) > 1:
                raise ValueError(f"Ligne {line_number} : plusieurs triangles actifs correspondent à :\n{display_key}.")
            imported_ranks[rank - 1] = matches[0]

        missing_ranks = [str(rank) for rank in range(1, 33) if rank not in seen_ranks]
        if missing_ranks:
            raise ValueError(f"Rangs manquants : {', '.join(missing_ranks)}.")
        template = self._get_selected_template()
        if template is None:
            raise ValueError("Aucun template n'est sélectionné.")
        message = self.catalogue.validate_template_ranks(template.template_id, imported_ranks)
        if message:
            raise ValueError(f"Le template importé est invalide : {message}")
        return imported_ranks

    def _model_triangle_cities(self, triangle: ModelCatalogueTriangle) -> tuple[CatalogueCity, CatalogueCity, CatalogueCity]:
        return (
            self.catalogue.get_city(triangle.opening_city_id),
            self.catalogue.get_city(triangle.base_city_id),
            self.catalogue.get_city(triangle.light_city_id),
        )

    def _visible_triangles(self) -> list[ModelCatalogueTriangle]:
        date_code = self._triangle_date_filter_var.get()
        base_filter = self._triangle_base_filter_var.get().strip().casefold()
        light_filter = self._triangle_light_filter_var.get().strip().casefold()
        status = self._triangle_status_filter_var.get()
        return sorted(
            (
                triangle for triangle in self.catalogue.iter_triangles()
                if (status == "Tout" or (status == "Actif" and not triangle.archived)
                    or (status == "Archivé" and triangle.archived))
                and (date_code == "Tous" or triangle.note == date_code)
                and (not base_filter or base_filter in self.catalogue.get_city(triangle.base_city_id).name.casefold())
                and (not light_filter or light_filter in self.catalogue.get_city(triangle.light_city_id).name.casefold())
            ),
            key=self._triangle_sort_key,
        )

    @classmethod
    def _template_note_order(cls, date_code: str) -> tuple[int, str]:
        date_code = date_code.strip()
        normalized = date_code.casefold()
        note_order = next(
            (order for note, order in cls._TEMPLATE_NOTE_ORDER.items()
             if normalized == note or normalized.startswith(f"{note} ")),
            len(cls._TEMPLATE_NOTE_ORDER),
        )
        return note_order, normalized

    def _triangle_sort_key(self, triangle: ModelCatalogueTriangle) -> tuple[int, str, str, str, str]:
        opening, base, light = self._model_triangle_cities(triangle)
        note_order, normalized = self._template_note_order(triangle.note)
        return (
            note_order,
            normalized,
            base.name.casefold(), light.name.casefold(), opening.name.casefold(),
        )

    def _template_note_sort_key(self, triangle: ModelCatalogueTriangle) -> tuple[int, str, str, str, str]:
        opening, base, light = self._model_triangle_cities(triangle)
        note_order, normalized = self._template_note_order(triangle.note)
        return (
            note_order,
            normalized,
            base.name.casefold(), light.name.casefold(), opening.name.casefold(),
        )

    def _visible_template_triangles(self) -> list[ModelCatalogueTriangle]:
        note = self._template_note_filter_var.get()
        base_filter = self._template_base_filter_var.get().strip().casefold()
        light_filter = self._template_light_filter_var.get().strip().casefold()
        return sorted(
            (
                triangle for triangle in self.catalogue.iter_triangles()
                if not triangle.archived
                and (note == "Tous" or triangle.note == note)
                and (not base_filter or base_filter in self.catalogue.get_city(triangle.base_city_id).name.casefold())
                and (not light_filter or light_filter in self.catalogue.get_city(triangle.light_city_id).name.casefold())
            ),
            key=self._template_note_sort_key,
        )

    def _refresh_template_triangle_tree(self):
        active_notes = sorted(
            {triangle.note for triangle in self.catalogue.iter_triangles() if not triangle.archived},
            key=self._template_note_order,
        )
        self._template_note_filter.configure(values=("Tous", *active_notes))
        if self._template_note_filter_var.get() not in {"Tous", *active_notes}:
            self._template_note_filter_var.set("Tous")
        self._template_triangle_tree.delete(*self._template_triangle_tree.get_children())
        self._template_triangle_by_tree_iid = {}
        grouped: dict[str, list[ModelCatalogueTriangle]] = {}
        for triangle in self._visible_template_triangles():
            grouped.setdefault(self.catalogue.get_city(triangle.base_city_id).name, []).append(triangle)
        ordered_bases = sorted(
            grouped,
            key=lambda base: (min(self._template_note_order(item.note) for item in grouped[base]), base.casefold()),
        )
        for base_index, base in enumerate(ordered_bases):
            notes = sorted({triangle.note for triangle in grouped[base]}, key=self._template_note_order)
            base_iid = f"template-base-{base_index}"
            self._template_triangle_tree.insert("", tk.END, iid=base_iid, text=f"{base} ({', '.join(notes)})", open=True)
            for light_index, triangle in enumerate(
                sorted(grouped[base], key=lambda item: (self.catalogue.get_city(item.light_city_id).name.casefold(), self._template_note_sort_key(item)))
            ):
                leaf_iid = f"{base_iid}-light-{light_index}"
                self._template_triangle_tree.insert(base_iid, tk.END, iid=leaf_iid, text=self.catalogue.get_city(triangle.light_city_id).name)
                self._template_triangle_by_tree_iid[leaf_iid] = triangle.triangle_id

    def _get_selected_template_triangle(self) -> str | None:
        selection = self._template_triangle_tree.selection()
        return self._template_triangle_by_tree_iid.get(selection[0]) if selection else None

    def _on_template_triangle_selected(self, _event=None):
        """Ne considère comme triangle sélectionné qu'une feuille Lumière."""
        self._selected_template_triangle_id = self._get_selected_template_triangle()

    def _on_template_tree_press(self, event):
        iid = self._template_triangle_tree.identify_row(event.y)
        triangle_id = self._template_triangle_by_tree_iid.get(iid)
        self._template_drag_triangle_id = triangle_id
        self._template_drag_source_iid = iid if triangle_id is not None else None
        self._template_drag_start_root = (event.x_root, event.y_root) if triangle_id is not None else None
        self._template_drag_started = False

    def _on_template_tree_drag(self, event):
        triangle_id = self._template_drag_triangle_id
        start = self._template_drag_start_root
        if triangle_id is None or start is None:
            return
        if not self._template_drag_started:
            distance = math.hypot(event.x_root - start[0], event.y_root - start[1])
            if distance < _TEMPLATE_DRAG_THRESHOLD:
                return
            self._template_drag_started = True
            template = self._get_selected_template()
            if template is not None:
                source_index = self._find_triangle_rank(template.triangle_ids_by_rank, triangle_id)
                if source_index is not None:
                    self._template_drag_source_slot = self._get_template_rank_slot(source_index)
                    self._template_drag_source_slot.set_drop_state("source")
            self._show_template_drag_ghost(triangle_id, event.x_root, event.y_root)
            self._set_template_drag_cursor("hand2")
        self._move_template_drag_ghost(event.x_root, event.y_root)
        slot = self._find_rank_slot_from_widget(self.winfo_containing(event.x_root, event.y_root))
        if slot is self._template_drag_target_slot:
            return
        if self._template_drag_target_slot is not None:
            self._restore_template_drag_slot_state(self._template_drag_target_slot)
        self._template_drag_target_slot = slot
        if slot is not None:
            plan = self._plan_template_drop(triangle_id, slot)
            if slot is self._template_drag_source_slot and plan.action == "noop":
                slot.set_drop_state("source")
            else:
                slot.set_drop_state(self._template_drop_visual_state(plan))

    def _on_template_tree_release(self, _event):
        if not self._template_drag_started:
            self._reset_template_drag_state()
            return
        triangle_id, slot = self._template_drag_triangle_id, self._template_drag_target_slot
        if triangle_id is not None and slot is not None:
            plan = self._plan_template_drop(triangle_id, slot)
            if plan.valid and plan.action != "noop":
                template = self._get_selected_template()
                if template is not None and plan.preview_ranks is not None:
                    self.catalogue.set_template_ranks(template.template_id, plan.preview_ranks)
                self._select_template_rank_slot(slot)
                self._refresh_template_ranks_view()
                self._mark_dirty()
            elif not plan.valid and plan.message:
                messagebox.showwarning("Template", plan.message, parent=self)
        self._reset_template_drag_state()

    @staticmethod
    def _find_rank_slot_from_widget(widget) -> TemplateRankSlot | None:
        while widget is not None:
            if isinstance(widget, TemplateRankSlot):
                return widget
            widget = widget.master
        return None

    @staticmethod
    def _find_triangle_rank(ranks: list[str | None], triangle_id: str) -> int | None:
        return next((index for index, item in enumerate(ranks) if item == triangle_id), None)

    def _get_template_rank_slot(self, rank_index: int) -> TemplateRankSlot:
        """Retourne le widget correspondant au rang 0-based indiqué."""
        pair_row = self._template_pair_rows[rank_index // 2]
        return pair_row.odd_rank_slot if rank_index % 2 == 0 else pair_row.even_rank_slot

    def _restore_template_drag_slot_state(self, slot: TemplateRankSlot) -> None:
        slot.set_drop_state("source" if slot is self._template_drag_source_slot else "normal")

    def _find_auto_companion(self, triangle_id: str, ranks: list[str | None]) -> str | None:
        """Retourne le compagnon libre d'une base ayant exactement deux triangles actifs."""
        if self._find_triangle_rank(ranks, triangle_id) is not None:
            return None
        triangle = self.catalogue.get_triangle(triangle_id)
        candidates = [
            item for item in self.catalogue.iter_triangles()
            if not item.archived and item.base_city_id == triangle.base_city_id
        ]
        if len(candidates) != 2 or triangle_id not in {item.triangle_id for item in candidates}:
            return None
        companion_id = next(item.triangle_id for item in candidates if item.triangle_id != triangle_id)
        return companion_id if self._find_triangle_rank(ranks, companion_id) is None else None

    def _plan_template_drop(self, triangle_id: str, target_slot: TemplateRankSlot) -> TemplateDropPlan:
        target_index = target_slot.rank_number - 1
        template = self._get_selected_template()
        if template is None:
            return TemplateDropPlan("invalid", False, "Aucun template n'est sélectionné.", None, target_index, None, None)
        triangle = self.catalogue.get_triangle(triangle_id)
        if triangle.archived:
            return TemplateDropPlan("invalid", False, "Un triangle archivé ne peut pas être utilisé.", None, target_index, None, None)
        ranks = template.triangle_ids_by_rank
        source_index = self._find_triangle_rank(ranks, triangle_id)
        target_triangle_id = ranks[target_index]
        if source_index == target_index:
            return TemplateDropPlan("noop", True, None, source_index, target_index, target_triangle_id, list(ranks))
        preview = list(ranks)
        if source_index is None:
            preview[target_index] = triangle_id
            action = "replace" if target_triangle_id is not None else "valid"
            other_target_index = target_index + 1 if target_index % 2 == 0 else target_index - 1
            companion = self._find_auto_companion(triangle_id, ranks)
            if companion is not None and ranks[other_target_index] is None:
                preview[other_target_index] = companion
        elif target_triangle_id is None:
            preview[source_index] = None
            preview[target_index] = triangle_id
            action = "move"
        else:
            preview[source_index] = target_triangle_id
            preview[target_index] = triangle_id
            action = "swap"
        message = self.catalogue.validate_template_ranks(template.template_id, preview)
        if message:
            return TemplateDropPlan("invalid", False, message, source_index, target_index, target_triangle_id, None)
        return TemplateDropPlan(action, True, None, source_index, target_index, target_triangle_id, preview)

    @staticmethod
    def _template_drop_visual_state(plan: TemplateDropPlan) -> str:
        return {
            "valid": "valid",
            "replace": "replace",
            "move": "move",
            "swap": "swap",
            "invalid": "invalid",
            "noop": "normal",
        }.get(plan.action, "invalid")

    def _validate_template_drop(
        self, triangle_id: str, target_slot: TemplateRankSlot
    ) -> tuple[bool, str | None, str]:
        """Compatibilité interne pour les appels existants ; délègue au plan atomique."""
        plan = self._plan_template_drop(triangle_id, target_slot)
        return plan.valid, plan.message, plan.action

    def _reset_template_drag_state(self):
        if self._template_drag_target_slot is not None:
            self._template_drag_target_slot.set_drop_state("normal")
        if (
            self._template_drag_source_slot is not None
            and self._template_drag_source_slot is not self._template_drag_target_slot
        ):
            self._template_drag_source_slot.set_drop_state("normal")
        self._template_drag_triangle_id = None
        self._template_drag_started = False
        self._template_drag_source_iid = None
        self._template_drag_source_slot = None
        self._template_drag_target_slot = None
        self._template_drag_start_root = None
        self._destroy_template_drag_ghost()
        self._set_template_drag_cursor("")

    def _show_template_drag_ghost(self, triangle_id: str, x_root: int, y_root: int):
        self._destroy_template_drag_ghost()
        ghost = tk.Toplevel(self)
        ghost.overrideredirect(True)
        try:
            ghost.attributes("-topmost", True)
            ghost.attributes("-disabled", True)
        except tk.TclError:
            pass
        ghost.configure(background="#ffffff", borderwidth=1, relief=tk.SOLID)
        content = tk.Frame(ghost, background="#ffffff", padx=8, pady=5)
        content.pack()
        tk.Label(content, image=self._icon_map_pin_plus, background="#ffffff").pack(side=tk.LEFT, padx=(0, 6))
        triangle = self.catalogue.get_triangle(triangle_id)
        light_name = self.catalogue.get_city(triangle.light_city_id).name
        tk.Label(content, text=light_name, background="#ffffff", font=(None, 9, "bold")).pack(side=tk.LEFT)
        self._template_drag_ghost = ghost
        self._move_template_drag_ghost(x_root, y_root)
        ghost.deiconify()

    def _move_template_drag_ghost(self, x_root: int, y_root: int):
        if self._template_drag_ghost is not None and self._template_drag_ghost.winfo_exists():
            self._template_drag_ghost.geometry(f"+{x_root + 16}+{y_root + 16}")

    def _destroy_template_drag_ghost(self):
        if self._template_drag_ghost is not None and self._template_drag_ghost.winfo_exists():
            self._template_drag_ghost.destroy()
        self._template_drag_ghost = None

    def _set_template_drag_cursor(self, cursor: str):
        self._template_triangle_tree.configure(cursor=cursor)
        for pair_row in getattr(self, "_template_pair_rows", ()):
            for slot in (pair_row.odd_rank_slot, pair_row.even_rank_slot):
                for widget in (slot, slot._badge, slot._entry):
                    widget.configure(cursor=cursor)

    def _refresh_triangle_tree(self):
        date_codes = sorted(
            {triangle.note for triangle in self.catalogue.iter_triangles()},
            key=self._template_note_order,
        )
        self._triangle_date_filter.configure(values=("Tous", *date_codes))
        if self._triangle_date_filter_var.get() not in {"Tous", *date_codes}:
            self._triangle_date_filter_var.set("Tous")
        selected_id = self._selected_triangle_id
        visible = self._visible_triangles()
        self._triangle_tree.delete(*self._triangle_tree.get_children())
        for triangle in visible:
            opening, base, light = self._model_triangle_cities(triangle)
            self._triangle_tree.insert(
                "", tk.END, iid=triangle.triangle_id,
                values=(triangle.note, opening.name, base.name, light.name),
            )
        visible_ids = {triangle.triangle_id for triangle in visible}
        if selected_id in visible_ids:
            self._triangle_tree.selection_set(selected_id)
            self._triangle_tree.focus(selected_id)
            self._triangle_tree.see(selected_id)
        elif selected_id is not None:
            self._selected_triangle_id = None
            self._show_triangle_details(None)
        self._catalogue_notebook.tab(self._triangles_tab, text=f"Triangles ({len(self.catalogue.triangles)})")
        self._update_triangle_action_buttons()
        self._refresh_template_triangle_tree()

    def _update_triangle_action_buttons(self):
        triangle = self._get_selected_model_triangle()
        state = tk.NORMAL if triangle is not None else tk.DISABLED
        self._triangle_delete_button.configure(state=state)
        self._triangle_archive_button.configure(
            state=state,
            image=self._icon_archive_off if triangle is not None and triangle.archived else self._icon_archive,
        )
        self._set_tooltip_text(
            self._triangle_archive_button,
            "Désarchiver le triangle" if triangle is not None and triangle.archived else "Archiver le triangle",
        )

    def _get_selected_model_triangle(self) -> ModelCatalogueTriangle | None:
        return self.catalogue.get_triangle(self._selected_triangle_id) if self._selected_triangle_id else None

    def _open_triangle_editor(self, triangle: ModelCatalogueTriangle | None = None):
        result = TriangleEditorDialog(
            self,
            list(self.catalogue.iter_cities()),
            triangle=triangle,
        ).show()
        if result is None:
            return
        try:
            if triangle is None:
                model_triangle = self.catalogue.add_triangle(result.note, result.opening_city_id, result.base_city_id, result.light_city_id)
                triangle_id = model_triangle.triangle_id
            else:
                self.catalogue.update_triangle(
                    triangle.triangle_id,
                    note=result.note,
                    opening_city_id=result.opening_city_id,
                    base_city_id=result.base_city_id,
                    light_city_id=result.light_city_id,
                )
                triangle_id = triangle.triangle_id
        except (ValueError, KeyError) as exc:
            messagebox.showerror("Triangle", str(exc), parent=self)
            return
        self._selected_triangle_id = triangle_id
        self._refresh_triangle_tree()
        self._show_triangle_details(self._get_selected_model_triangle())
        self._mark_dirty()

    def _add_triangle(self):
        self._open_triangle_editor()

    def _edit_selected_triangle(self):
        triangle = self._get_selected_model_triangle()
        if triangle is not None:
            self._open_triangle_editor(triangle)

    def _toggle_archive_selected_triangle(self):
        triangle = self._get_selected_model_triangle()
        if triangle is None:
            return
        self.catalogue.update_triangle(
            triangle.triangle_id, archived=not triangle.archived,
        )
        self._refresh_triangle_tree()
        self._show_triangle_details(self._get_selected_model_triangle())
        self._mark_dirty()

    def _delete_selected_triangle(self):
        triangle = self._get_selected_model_triangle()
        if triangle is None:
            return
        try:
            self.catalogue.delete_triangle(triangle.triangle_id)
        except ValueError as exc:
            messagebox.showerror("Supprimer le triangle", str(exc), parent=self)
            return
        self._selected_triangle_id = None
        self._refresh_triangle_tree()
        self._show_triangle_details(None)
        self._mark_dirty()

    def _on_triangle_selected(self, _event=None):
        selection = self._triangle_tree.selection()
        self._selected_triangle_id = selection[0] if selection else None
        self._show_triangle_details(self._get_selected_model_triangle())
        self._update_triangle_action_buttons()

    def _show_triangle_details(self, triangle: ModelCatalogueTriangle | None):
        if triangle is None:
            self._triangle_distances_label.configure(text="Distances")
            self._triangle_angles_label.configure(text="Angles")
            self._triangle_status_label.configure(text="Statut")
            self._triangle_map_view.set_markers(())
            self._triangle_map_view.set_polylines(())
            return
        opening, base, light = self._model_triangle_cities(triangle)
        try:
            geometry = self.catalogue.get_triangle_geometry(triangle.triangle_id)
        except ValueError:
            self._triangle_distances_label.configure(text="Distances : géométrie invalide")
            self._triangle_angles_label.configure(text="Angles : géométrie invalide")
            self._triangle_status_label.configure(text=f"Statut : {'Archivé' if triangle.archived else 'Actif'}")
            self._triangle_map_view.set_markers(())
            self._triangle_map_view.set_polylines(())
            return
        orientation = {"CW": "Horaire", "CCW": "Antihoraire"}[geometry.orientation]
        self._triangle_distances_label.configure(
            text=f"Distances : OB={geometry.distance_ob_km:.1f} km    BL={geometry.distance_bl_km:.1f} km    LO={geometry.distance_ol_km:.1f} km"
        )
        self._triangle_angles_label.configure(
            text=(f"Angles : O={geometry.angle_o_deg:.1f}°    B={geometry.angle_b_deg:.1f}°    L={geometry.angle_l_deg:.1f}°"
                  f"    Orientation : {orientation}")
        )
        self._triangle_status_label.configure(text=f"Statut : {'Archivé' if triangle.archived else 'Actif'}")
        cities = (opening, base, light)
        coordinates = tuple((city.latitude, city.longitude) for city in cities)
        marker_styles = (
            ("O", "#000000", "#000000"),
            ("B", "#1565c0", "#1565c0"),
            ("L", "#f6d32d", "#202020"),
        )
        self._triangle_map_view.set_markers(
            GeoMapMarker(
                city.city_id,
                city.latitude,
                city.longitude,
                city.name,
                always_show_label=True,
                fill_color=fill_color,
                outline_color=outline_color,
                tooltip=f"{role} : {city.name}\nAngle : {getattr(geometry, f'angle_{role.lower()}_deg'):.1f}°",
            )
            for city, (role, fill_color, outline_color) in zip(cities, marker_styles)
        )
        self._triangle_map_view.set_selected_marker(None)
        self._triangle_map_view.set_polylines((GeoMapPolyline(coordinates, color="#000000", closed=True),))
        self._triangle_map_view.fit_to_bounds(coordinates, margin=_TRIANGLE_MAP_FIT_MARGIN)

    def _visible_cities(self) -> list[CatalogueCity]:
        search = _normalize_search_text(self._search_var.get().strip())
        cities = [
            city
            for city in self.catalogue.iter_cities()
            if (self._show_archived_var.get() or not city.archived)
            and (
                not search
                or search in _normalize_search_text(city.name)
            )
        ]
        return sorted(cities, key=lambda city: city.name.casefold())

    def _visible_beacons(self):
        search = _normalize_search_text(self._beacon_search_var.get().strip())
        return [
            beacon
            for beacon in self.catalogue.iter_beacons()
            if (self._show_archived_beacons_var.get() or not beacon.archived)
            and (
                not search
                or search
                in _normalize_search_text(
                    self.catalogue.get_city(beacon.city_id).name
                )
            )
        ]

    def _refresh_beacon_list(self):
        selected_id, visible = self._selected_beacon_id, self._visible_beacons()
        self._beacon_listbox.delete(0, tk.END)
        for beacon in visible:
            self._beacon_listbox.insert(tk.END, self.catalogue.get_city(beacon.city_id).name)
        if any(beacon.beacon_id == selected_id for beacon in visible):
            index = next(index for index, beacon in enumerate(visible) if beacon.beacon_id == selected_id)
            self._beacon_listbox.selection_set(index)
            self._beacon_listbox.activate(index)
            self._beacon_listbox.see(index)
        elif selected_id is not None:
            self._selected_beacon_id = None
            self._load_selected_beacon()
        self._catalogue_notebook.tab(self._beacons_tab, text=f"Balises ({len(self.catalogue.beacons)})")
        self._beacon_map_view.set_markers(
            GeoMapMarker(
                beacon.beacon_id,
                self.catalogue.get_city(beacon.city_id).latitude,
                self.catalogue.get_city(beacon.city_id).longitude,
                self.catalogue.get_city(beacon.city_id).name,
            )
            for beacon in visible
        )
        self._beacon_map_view.set_selected_marker(self._selected_beacon_id)
        self._update_beacon_action_buttons()

    def _available_beacon_cities(self) -> list[CatalogueCity]:
        beacon_city_ids = {beacon.city_id for beacon in self.catalogue.iter_beacons()}
        return sorted(
            (
                city for city in self.catalogue.iter_cities()
                if not city.archived and city.city_id not in beacon_city_ids
            ),
            key=lambda city: (city.name.casefold(), city.city_id),
        )

    def _on_beacon_selected(self, _event=None):
        selection, visible = self._beacon_listbox.curselection(), self._visible_beacons()
        self._selected_beacon_id = visible[selection[0]].beacon_id if selection else None
        self._load_selected_beacon()
        self._beacon_map_view.set_selected_marker(self._selected_beacon_id, recenter=True)
        self._update_beacon_action_buttons()

    def _on_beacon_map_marker_selected(self, marker_id):
        self._selected_beacon_id = marker_id if marker_id in self.catalogue.beacons else None
        self._refresh_beacon_list()
        self._load_selected_beacon()

    def _load_selected_beacon(self):
        beacon = self.catalogue.get_beacon(self._selected_beacon_id) if self._selected_beacon_id else None
        city = self.catalogue.get_city(beacon.city_id) if beacon is not None else None
        self._beacon_city_label.configure(text=city.name if city else "")
        self._beacon_latitude_editor.set_decimal(city.latitude if city else 0.0)
        self._beacon_longitude_editor.set_decimal(city.longitude if city else 0.0)
        self._beacon_archived_var.set(beacon.archived if beacon else False)

    def _update_beacon_action_buttons(self):
        beacon = self.catalogue.get_beacon(self._selected_beacon_id) if self._selected_beacon_id else None
        state = tk.NORMAL if beacon is not None else tk.DISABLED
        self._beacon_delete_button.configure(state=state)
        self._beacon_archive_button.configure(
            state=state,
            image=self._icon_archive_off if beacon is not None and beacon.archived else self._icon_archive,
        )
        self._set_tooltip_text(
            self._beacon_archive_button,
            "Désarchiver la balise" if beacon is not None and beacon.archived else "Archiver la balise",
        )

    def _add_beacon(self):
        cities = self._available_beacon_cities()
        if not cities:
            messagebox.showinfo(
                "Ajouter une balise",
                "Aucune ville disponible pour créer une nouvelle balise.",
                parent=self,
            )
            return
        city_id = CitySelectionDialog(self, cities).show()
        if city_id is None:
            return
        try:
            beacon = self.catalogue.add_beacon(city_id)
        except ValueError as exc:
            messagebox.showerror("Ajouter une balise", str(exc), parent=self)
            return
        self._selected_beacon_id = beacon.beacon_id
        self._refresh_beacon_list()
        self._load_selected_beacon()
        self._refresh_city_list()
        self._load_selected_city()
        self._mark_dirty()

    def _archive_selected_beacon(self):
        if self._selected_beacon_id is None:
            return
        beacon = self.catalogue.get_beacon(self._selected_beacon_id)
        self.catalogue.update_beacon(beacon.beacon_id, archived=not beacon.archived)
        self._refresh_beacon_list()
        self._load_selected_beacon()
        self._refresh_city_list()
        self._load_selected_city()
        self._mark_dirty()

    def _delete_selected_beacon(self):
        if self._selected_beacon_id is None:
            return
        beacon = self.catalogue.get_beacon(self._selected_beacon_id)
        city = self.catalogue.get_city(beacon.city_id)
        if self._is_beacon_referenced is not None and self._is_beacon_referenced(beacon.beacon_id):
            messagebox.showerror(
                "Supprimer la balise",
                "Cette balise est utilisée par un ancrage de scénario. Archivez-la plutôt que de la supprimer.",
                parent=self,
            )
            return
        if not messagebox.askyesno(
            "Supprimer la balise", f"Supprimer la balise associée à {city.name} ?", parent=self,
        ):
            return
        self.catalogue.delete_beacon(beacon.beacon_id)
        self._selected_beacon_id = None
        self._refresh_beacon_list()
        self._load_selected_beacon()
        self._refresh_city_list()
        self._load_selected_city()
        self._mark_dirty()

    def _refresh_city_list(self):
        selected_id, visible = self._selected_city_id, self._visible_cities()
        self._city_listbox.delete(0, tk.END)
        for city in visible:
            self._city_listbox.insert(tk.END, city.name)
        if any(city.city_id == selected_id for city in visible):
            index = next(index for index, city in enumerate(visible) if city.city_id == selected_id)
            self._city_listbox.selection_set(index)
            self._city_listbox.activate(index)
        elif selected_id is not None:
            self._selected_city_id = None
            self._load_selected_city()
        self._catalogue_notebook.tab(self._cities_tab, text=f"Villes ({len(self.catalogue.cities)})")
        self._map_view.set_markers(
            GeoMapMarker(city.city_id, city.latitude, city.longitude, city.name) for city in self.catalogue.iter_cities()
        )
        self._map_view.set_selected_marker(self._selected_city_id)
        self._update_city_action_buttons()

    def _update_city_action_buttons(self):
        city = self.catalogue.get_city(self._selected_city_id) if self._selected_city_id else None
        state = tk.NORMAL if city is not None else tk.DISABLED
        self._city_delete_button.configure(state=state)
        self._city_archive_button.configure(
            state=state,
            image=self._icon_archive_off if city is not None and city.archived else self._icon_archive,
        )
        self._set_tooltip_text(
            self._city_archive_button,
            "Désarchiver la ville" if city is not None and city.archived else "Archiver la ville",
        )

    def _on_city_selected(self, _event=None):
        selection, visible = self._city_listbox.curselection(), self._visible_cities()
        self._selected_city_id = visible[selection[0]].city_id if selection else None
        self._load_selected_city()
        self._map_view.set_selected_marker(self._selected_city_id, recenter=True)
        self._update_city_action_buttons()

    def _on_map_marker_selected(self, marker_id):
        self._selected_city_id = marker_id if marker_id in self.catalogue.cities else None
        self._refresh_city_list()
        self._load_selected_city()

    def _load_selected_city(self):
        self._updating_detail = True
        city = self.catalogue.get_city(self._selected_city_id) if self._selected_city_id else None
        self._name_var.set(city.name if city else "")
        self._latitude_editor.set_decimal(city.latitude if city else 0.0)
        self._longitude_editor.set_decimal(city.longitude if city else 0.0)
        self._archived_var.set(city.archived if city else False)

        triangle_references = (
            self.catalogue.get_triangles_referencing_city(city.city_id)
            if city else ()
        )

        beacon_references = (
            self.catalogue.get_beacons_referencing_city(city.city_id)
            if city else ()
        )

        self._city_triangle_count_label.configure(
            text=f"Triangles : {len(triangle_references)}"
        )

        self._city_triangle_references_button.configure(
            state=tk.NORMAL if triangle_references else tk.DISABLED,
        )
        self._city_beacon_count_label.configure(text=f"Balises : {len(beacon_references)}")
        self._city_beacon_references_button.configure(
            state=tk.NORMAL if beacon_references else tk.DISABLED,
        )
        self._updating_detail = False

    def _show_city_triangle_references(self):
        if self._selected_city_id is None:
            return

        references = self.catalogue.get_triangles_referencing_city(
            self._selected_city_id
        )
        if not references:
            return

        city = self.catalogue.get_city(self._selected_city_id)

        window = tk.Toplevel(self)
        window.title(f"Triangles utilisant {city.name}")
        window.transient(self)
        window.resizable(True, True)
        window.geometry("650x300")

        root = ttk.Frame(window, padding=10)
        root.pack(fill=tk.BOTH, expand=True)

        columns = ("note", "opening", "base", "light")

        tree = ttk.Treeview(
            root,
            columns=columns,
            show="headings",
            selectmode="browse",
        )

        tree.heading("note", text="Note")
        tree.heading("opening", text="Ouverture")
        tree.heading("base", text="Base")
        tree.heading("light", text="Lumière")

        tree.column("note", width=70, stretch=False)
        tree.column("opening", width=160)
        tree.column("base", width=160)
        tree.column("light", width=160)

        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        scrollbar = ttk.Scrollbar(
            root,
            orient=tk.VERTICAL,
            command=tree.yview,
        )
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        tree.configure(yscrollcommand=scrollbar.set)

        for triangle in references:
            opening = self.catalogue.get_city(
                triangle.opening_city_id
            )
            base = self.catalogue.get_city(
                triangle.base_city_id
            )
            light = self.catalogue.get_city(
                triangle.light_city_id
            )

            tree.insert(
                "",
                tk.END,
                values=(
                    triangle.note,
                    opening.name,
                    base.name,
                    light.name,
                ),
            )

        window.bind("<Escape>", lambda _event: window.destroy())

    def _show_city_beacon_references(self):
        if self._selected_city_id is None:
            return
        references = self.catalogue.get_beacons_referencing_city(self._selected_city_id)
        if not references:
            return
        lines = [
            f"{beacon.beacon_id} — {'Archivée' if beacon.archived else 'Active'}"
            for beacon in references
        ]
        messagebox.showinfo("Balises référencées", "\n".join(lines), parent=self)

    def _save_detail_changes(self, *_args):
        if self._updating_detail or self._selected_city_id is None:
            return
        try:
            self.catalogue.update_city(
                self._selected_city_id,
                name=self._name_var.get().strip(),
                latitude=self._latitude_editor.get_decimal(),
                longitude=self._longitude_editor.get_decimal(),
                archived=bool(self._archived_var.get()),
            )
        except (ValueError, KeyError) as exc:
            messagebox.showerror("Ville", str(exc), parent=self)
            self._load_selected_city()
            return
        self._refresh_city_list()
        self._mark_dirty()

    def _paste_coordinates(self):
        try:
            latitude, longitude = DmsCoordinateEditor.parse_coordinate_pair(self.clipboard_get())
        except (tk.TclError, ValueError) as exc:
            messagebox.showerror("Coller coordonnées", str(exc), parent=self)
            return
        self._latitude_editor.set_decimal(latitude)
        self._longitude_editor.set_decimal(longitude)
        self._save_detail_changes()

    def _add_city(self):
        city = self.catalogue.add_city("Nouvelle ville", 0.0, 0.0)
        self._selected_city_id = city.city_id
        self._refresh_city_list()
        self._load_selected_city()
        self._mark_dirty()

    def _archive_selected_city(self):
        if self._selected_city_id is not None:
            city = self.catalogue.get_city(self._selected_city_id)
            self.catalogue.update_city(city.city_id, archived=not city.archived)
            self._refresh_city_list()
            self._load_selected_city()
            self._mark_dirty()

    def _delete_selected_city(self):
        if self._selected_city_id is not None:
            city_id = self._selected_city_id
            try:
                self.catalogue.delete_city(city_id)
            except ValueError as exc:
                messagebox.showerror("Supprimer la ville", str(exc), parent=self)
                return
            self._selected_city_id = None
            self._refresh_city_list()
            self._load_selected_city()
            self._mark_dirty()
